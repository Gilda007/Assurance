
import os
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFileDialog, QTableWidget, QTableWidgetItem, QProgressBar,
    QMessageBox, QFrame, QComboBox, QRadioButton, QTabWidget,
    QScrollArea, QSplitter, QTextEdit, QCheckBox, QHeaderView,
    QWidget, QApplication, QLineEdit, QDateEdit, QGridLayout,
    QGroupBox, QSizePolicy, QSpacerItem, QMenu, QInputDialog,
    QDialogButtonBox, QSlider
)
from PySide6.QtCore import Qt, Signal, QThread, QDate, QSettings
from PySide6.QtGui import QFont, QColor, QPalette, QAction
import pandas as pd
import traceback
from datetime import datetime
import os
import warnings

# Supprimer les avertissements Wayland
os.environ["QT_LOGGING_RULES"] = "qt.qpa.wayland.*=false"
os.environ["QT_QPA_PLATFORM"] = "xcb"
os.environ["XDG_SESSION_TYPE"] = "x11"
os.environ["GDK_BACKEND"] = "x11"
warnings.filterwarnings("ignore")


# ============================================================================
# STYLES MODERNES
# ============================================================================

class AppColors:
    PRIMARY = "#2563eb"
    PRIMARY_DARK = "#1d4ed8"
    PRIMARY_LIGHT = "#eff6ff"
    SECONDARY = "#64748b"
    SUCCESS = "#10b981"
    SUCCESS_DARK = "#059669"
    SUCCESS_LIGHT = "#d1fae5"
    WARNING = "#f59e0b"
    WARNING_LIGHT = "#fef3c7"
    DANGER = "#ef4444"
    DANGER_LIGHT = "#fee2e2"
    DARK = "#0f172a"
    GRAY = "#64748b"
    GRAY_LIGHT = "#f1f5f9"
    WHITE = "#ffffff"
    BORDER = "#e2e8f0"


STYLESHEET = f"""
    QDialog {{
        background: {AppColors.GRAY_LIGHT};
    }}
    QGroupBox {{
        font-weight: 600;
        border: 1px solid {AppColors.BORDER};
        border-radius: 12px;
        margin-top: 12px;
        padding-top: 12px;
        background: {AppColors.WHITE};
    }}
    QGroupBox::title {{
        subcontrol-origin: margin;
        left: 16px;
        padding: 0 12px;
        color: {AppColors.DARK};
    }}
    QLineEdit, QComboBox, QDateEdit, QTextEdit {{
        border: 1px solid {AppColors.BORDER};
        border-radius: 8px;
        padding: 8px 12px;
        background: {AppColors.WHITE};
        font-size: 13px;
    }}
    QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{
        border-color: {AppColors.PRIMARY};
    }}
    QPushButton {{
        border: none;
        border-radius: 8px;
        padding: 8px 16px;
        font-size: 13px;
        font-weight: 500;
    }}
    QPushButton:hover {{
        background: {AppColors.GRAY_LIGHT};
    }}
    .BtnPrimary {{
        background: {AppColors.PRIMARY};
        color: white;
    }}
    .BtnPrimary:hover {{
        background: {AppColors.PRIMARY_DARK};
    }}
    .BtnSuccess {{
        background: {AppColors.SUCCESS};
        color: white;
    }}
    .BtnSuccess:hover {{
        background: {AppColors.SUCCESS_DARK};
    }}
    .BtnSecondary {{
        background: {AppColors.WHITE};
        color: {AppColors.GRAY};
        border: 1px solid {AppColors.BORDER};
    }}
    .BtnWarning {{
        background: {AppColors.WARNING};
        color: white;
    }}
    .BtnWarning:hover {{
        background: #d97706;
    }}
    QTableWidget {{
        border: 1px solid {AppColors.BORDER};
        border-radius: 10px;
        background: {AppColors.WHITE};
        alternate-background-color: {AppColors.GRAY_LIGHT};
    }}
    QHeaderView::section {{
        background: {AppColors.GRAY_LIGHT};
        padding: 8px 4px;
        font-weight: 600;
        font-size: 10px;
        color: {AppColors.GRAY};
        border: none;
    }}
    QProgressBar {{
        border: none;
        border-radius: 6px;
        background: {AppColors.GRAY_LIGHT};
        height: 6px;
    }}
    QProgressBar::chunk {{
        background: {AppColors.PRIMARY};
        border-radius: 6px;
    }}
    QScrollArea {{
        border: none;
        background: transparent;
    }}
    QScrollBar:vertical {{
        border: none;
        background: {AppColors.GRAY_LIGHT};
        width: 8px;
        border-radius: 4px;
    }}
    QScrollBar::handle:vertical {{
        background: {AppColors.BORDER};
        border-radius: 4px;
        min-height: 20px;
    }}
    QScrollBar::handle:vertical:hover {{
        background: {AppColors.PRIMARY};
    }}
    QMenu {{
        background: white;
        border: 1px solid {AppColors.BORDER};
        border-radius: 8px;
        padding: 4px;
    }}
    QMenu::item {{
        padding: 6px 20px;
        border-radius: 4px;
    }}
    QMenu::item:selected {{
        background: {AppColors.PRIMARY_LIGHT};
        color: {AppColors.PRIMARY_DARK};
    }}
"""


# ============================================================================
# DIALOGUE DE RÃ‰DUCTION EN MASSE
# ============================================================================

class MassReductionDialog(QDialog):
    """Dialogue pour appliquer une rÃ©duction en masse sur les garanties sÃ©lectionnÃ©es"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Application de rÃ©duction en masse")
        self.setMinimumSize(550, 500)
        self.setModal(True)
        self.setStyleSheet(STYLESHEET)
        
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(14)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # En-tÃªte
        header = QLabel("ðŸ“Š Application de rÃ©duction en masse")
        header.setStyleSheet("font-size: 16px; font-weight: 700; color: #0f172a;")
        layout.addWidget(header)
        
        desc = QLabel("Appliquez un pourcentage de rÃ©duction sur les garanties sÃ©lectionnÃ©es")
        desc.setStyleSheet("color: #64748b; font-size: 12px;")
        layout.addWidget(desc)
        
        # === GARANTIES ===
        group = QGroupBox("Garanties concernÃ©es")
        group.setStyleSheet("""
            QGroupBox {
                font-weight: 600;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 6px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 8px;
                color: #1e293b;
            }
        """)
        group_layout = QVBoxLayout(group)
        group_layout.setSpacing(8)
        
        # Boutons d'action
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        select_all_btn = QPushButton("Tout sÃ©lectionner")
        select_all_btn.setFlat(True)
        select_all_btn.setStyleSheet("color: #2563eb; font-size: 11px; padding: 4px 8px;")
        select_all_btn.clicked.connect(lambda: self.toggle_all(True))
        btn_layout.addWidget(select_all_btn)
        
        deselect_all_btn = QPushButton("Tout dÃ©sÃ©lectionner")
        deselect_all_btn.setFlat(True)
        deselect_all_btn.setStyleSheet("color: #64748b; font-size: 11px; padding: 4px 8px;")
        deselect_all_btn.clicked.connect(lambda: self.toggle_all(False))
        btn_layout.addWidget(deselect_all_btn)
        group_layout.addLayout(btn_layout)
        
        # Liste des garanties
        self.garantie_checkboxes = {}
        garanties_list = [
            ('rc', "RC/RTI", "ðŸ›¡ï¸"),
            ('dr', "DÃ©fense et Recours", "âš–ï¸"),
            ('vol', "Vol/Vol partie", "ðŸš—"),
            ('vb', "Vol Braquage", "ðŸ”«"),
            ('incendie', "Incendie", "ðŸ”¥"),
            ('bris_glace', "Bris de Glaces", "ðŸªŸ"),
            ('ar', "Assistance RÃ©paration", "ðŸ”§"),
            ('dta', "Dommages Tous Accidents", "ðŸ’¥"),
            ('ipt', "IPT + Conducteur", "ðŸ‘¥"),
        ]
        
        checkbox_layout = QGridLayout()
        checkbox_layout.setSpacing(6)
        
        for i, (key, label, icon) in enumerate(garanties_list):
            row = i // 3
            col = (i % 3) * 2
            
            cb = QCheckBox(f"{icon} {label}")
            cb.setChecked(True)
            cb.setStyleSheet("""
                QCheckBox {
                    font-size: 11px;
                    font-weight: 500;
                    padding: 4px 6px;
                    spacing: 6px;
                }
                QCheckBox::indicator {
                    width: 16px;
                    height: 16px;
                }
            """)
            cb.toggled.connect(self.update_summary)
            self.garantie_checkboxes[key] = cb
            checkbox_layout.addWidget(cb, row, col)
        
        group_layout.addLayout(checkbox_layout)
        layout.addWidget(group)
        
        # === PARAMÃˆTRES ===
        reduction_group = QGroupBox("ParamÃ¨tres de rÃ©duction")
        reduction_group.setStyleSheet("""
            QGroupBox {
                font-weight: 600;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 6px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 8px;
                color: #1e293b;
            }
        """)
        reduction_layout = QGridLayout(reduction_group)
        reduction_layout.setSpacing(10)
        reduction_layout.setContentsMargins(15, 15, 15, 15)
        
        # Pourcentage
        reduction_layout.addWidget(QLabel("Pourcentage de rÃ©duction (%) :"), 0, 0)
        self.reduction_input = QLineEdit("10")
        self.reduction_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 8px 12px;
                background: white;
                font-size: 14px;
                font-weight: 600;
                min-width: 80px;
            }
            QLineEdit:focus {
                border-color: #2563eb;
            }
        """)
        self.reduction_input.textChanged.connect(self.update_summary)
        reduction_layout.addWidget(self.reduction_input, 0, 1)
        
        # Curseur
        self.reduction_slider = QSlider(Qt.Horizontal)
        self.reduction_slider.setRange(0, 100)
        self.reduction_slider.setValue(10)
        self.reduction_slider.setTickInterval(10)
        self.reduction_slider.setTickPosition(QSlider.TicksBelow)
        self.reduction_slider.valueChanged.connect(self.on_slider_changed)
        reduction_layout.addWidget(self.reduction_slider, 1, 0, 1, 2)
        
        # AperÃ§u
        preview_frame = QFrame()
        preview_frame.setStyleSheet("""
            QFrame {
                background: #f8fafc;
                border-radius: 6px;
                padding: 8px 12px;
                margin-top: 4px;
            }
        """)
        preview_layout = QHBoxLayout(preview_frame)
        
        self.preview_label = QLabel("ðŸ’¡ 10% de rÃ©duction")
        self.preview_label.setStyleSheet("color: #475569; font-size: 11px; font-weight: 500;")
        preview_layout.addWidget(self.preview_label)
        preview_layout.addStretch()
        
        self.preview_total = QLabel("0 FCFA Ã©conomisÃ©s")
        self.preview_total.setStyleSheet("color: #10b981; font-size: 11px; font-weight: 600;")
        preview_layout.addWidget(self.preview_total)
        
        reduction_layout.addWidget(preview_frame, 2, 0, 1, 2)
        
        layout.addWidget(reduction_group)
        
        # === RÃ‰SUMÃ‰ ===
        summary_frame = QFrame()
        summary_frame.setStyleSheet("""
            QFrame {
                background: #eff6ff;
                border: 1px solid #bfdbfe;
                border-radius: 8px;
                padding: 10px 14px;
            }
        """)
        summary_layout = QHBoxLayout(summary_frame)
        summary_layout.setSpacing(20)
        
        self.summary_text = QLabel("ðŸ“‹ 9 garanties sÃ©lectionnÃ©es Â· 10% de rÃ©duction")
        self.summary_text.setStyleSheet("color: #1e293b; font-size: 12px; font-weight: 500;")
        summary_layout.addWidget(self.summary_text)
        summary_layout.addStretch()
        
        self.summary_vehicles = QLabel("ðŸš— 0 vÃ©hicules")
        self.summary_vehicles.setStyleSheet("color: #2563eb; font-size: 12px; font-weight: 600;")
        summary_layout.addWidget(self.summary_vehicles)
        
        layout.addWidget(summary_frame)
        
        # === ACTIONS ===
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("Annuler")
        cancel_btn.setProperty("class", "BtnSecondary")
        cancel_btn.setFixedSize(120, 35)
        cancel_btn.clicked.connect(self.reject)
        
        apply_btn = QPushButton("âœ… Appliquer")
        apply_btn.setProperty("class", "BtnSuccess")
        apply_btn.setFixedSize(120, 35)
        apply_btn.clicked.connect(self.accept)
        
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(apply_btn)
        layout.addLayout(btn_layout)
        
        self.update_summary()
    
    def on_slider_changed(self, value):
        """Met Ã  jour le champ de pourcentage depuis le curseur"""
        self.reduction_input.setText(str(value))
    
    def toggle_all(self, checked):
        """Active ou dÃ©sactive toutes les checkboxes"""
        for cb in self.garantie_checkboxes.values():
            cb.setChecked(checked)
        self.update_summary()
    
    def update_summary(self):
        """Met Ã  jour le rÃ©sumÃ©"""
        selected = [key for key, cb in self.garantie_checkboxes.items() if cb.isChecked()]
        pct = self.get_reduction_percentage()
        
        # Mettre Ã  jour le texte du rÃ©sumÃ©
        self.summary_text.setText(
            f"ðŸ“‹ {len(selected)} garantie{'s' if len(selected) > 1 else ''} sÃ©lectionnÃ©e{'s' if len(selected) > 1 else ''} Â· {pct:.1f}% de rÃ©duction"
        )
        
        # Mettre Ã  jour l'aperÃ§u
        self.preview_label.setText(f"ðŸ’¡ {pct:.1f}% de rÃ©duction")
        
        # Compter les vÃ©hicules impactÃ©s
        if self.parent() and hasattr(self.parent(), 'vehicles_data'):
            vehicles = self.parent().vehicles_data
            impacted = 0
            for v in vehicles:
                garanties = v.get('garanties', {})
                for key in selected:
                    if garanties.get(key, 0) > 0:
                        impacted += 1
                        break
            self.summary_vehicles.setText(f"ðŸš— {impacted} vÃ©hicule{'s' if impacted > 1 else ''}")
            
            # Calculer les Ã©conomies totales
            total_economies = 0
            for v in vehicles:
                garanties = v.get('garanties', {})
                for key in selected:
                    amount = garanties.get(key, 0)
                    if amount > 0:
                        total_economies += amount * (pct / 100)
            self.preview_total.setText(f"{total_economies:,.0f}".replace(",", " ") + " FCFA Ã©conomisÃ©s")
        else:
            self.summary_vehicles.setText("ðŸš— VÃ©hicules chargÃ©s")
            self.preview_total.setText("0 FCFA Ã©conomisÃ©s")
    
    def get_selected_garanties(self):
        """Retourne la liste des garanties sÃ©lectionnÃ©es"""
        return [key for key, cb in self.garantie_checkboxes.items() if cb.isChecked()]
    
    def get_reduction_percentage(self):
        """Retourne le pourcentage de rÃ©duction"""
        try:
            value = float(self.reduction_input.text().replace(",", "."))
            if value < 0:
                return 0.0
            if value > 100:
                return 100.0
            return value
        except:
            return 0.0


def build_vehicle_import_payload(vehicle, owner_id, compagny_id, fleet_id, current_user_id, debut, fin, jours, total_garanties, tva_amount, accessoires, asac, carte_rose, vignette):
    """Construit le payload attendu par le contrÃ´leur pour crÃ©er un vÃ©hicule complet."""
    garanties = vehicle.get('garanties', {}) or {}

    categorie_value = (vehicle.get('categorie') or 'VP').strip() or 'VP'
    genre_value = (vehicle.get('genre') or 'GV04').strip() or 'GV04'
    type_value = (vehicle.get('type_vehicule') or 'TV10').strip() or 'TV10'
    usage_value = (vehicle.get('usage') or 'UV01').strip() or 'UV01'
    energie_value = (vehicle.get('energie') or 'SEE').strip() or 'SEE'

    # âœ… RÃ©cupÃ©rer les valeurs de prime depuis le vÃ©hicule (dÃ©jÃ  extraites du fichier)
    prime_brute = vehicle.get('prime_brute', 0)
    prime_nette = vehicle.get('prime_nette', 0)
    reductions = vehicle.get('reductions', 0)

    guarantee_amounts = {
        'amt_rc': float(garanties.get('rc', 0) or 0),
        'amt_dr': float(garanties.get('dr', 0) or 0),
        'amt_vol': float(garanties.get('vol', 0) or 0),
        'amt_vb': float(garanties.get('vb', 0) or 0),
        'amt_in': float(garanties.get('incendie', garanties.get('in_garantie', 0)) or 0),
        'amt_bris': float(garanties.get('bris_glace', 0) or 0),
        'amt_ar': float(garanties.get('ar', 0) or 0),
        'amt_dta': float(garanties.get('dta', 0) or 0),
        'amt_ipt': float(garanties.get('ipt', 0) or 0),
    }

    guarantee_options = {
        'check_rc': guarantee_amounts['amt_rc'] > 0,
        'check_dr': guarantee_amounts['amt_dr'] > 0,
        'check_vol': guarantee_amounts['amt_vol'] > 0,
        'check_vb': guarantee_amounts['amt_vb'] > 0,
        'check_in': guarantee_amounts['amt_in'] > 0,
        'check_bris': guarantee_amounts['amt_bris'] > 0,
        'check_ar': guarantee_amounts['amt_ar'] > 0,
        'check_dta': guarantee_amounts['amt_dta'] > 0,
        'check_ipt': guarantee_amounts['amt_ipt'] > 0,
    }

    # âœ… RÃ©cupÃ©rer les rÃ©ductions individuelles
    reductions_individuelles = {
        'red_rc': float(garanties.get('red_rc', garanties.get('reduction_rc', 0)) or 0),
        'red_dr': float(garanties.get('red_dr', garanties.get('reduction_dr', 0)) or 0),
        'red_vol': float(garanties.get('red_vol', garanties.get('reduction_vol', 0)) or 0),
        'red_vb': float(garanties.get('red_vb', garanties.get('reduction_vb', 0)) or 0),
        'red_in': float(garanties.get('red_in', garanties.get('reduction_in', 0)) or 0),
        'red_bris': float(garanties.get('red_bris', garanties.get('reduction_bris', 0)) or 0),
        'red_ar': float(garanties.get('red_ar', garanties.get('reduction_ar', 0)) or 0),
        'red_dta': float(garanties.get('red_dta', garanties.get('reduction_dta', 0)) or 0),
        'red_ipt': float(garanties.get('red_ipt', garanties.get('reduction_ipt', 0)) or 0),
    }

    # âœ… Calcul de la TVA (sur la base de la prime nette)
    # base_tva = prime_nette + accessoires + asac
    # tva = base_tva * 0.1925
    tva_rate = 0.1925
    base_tva = prime_nette + accessoires + asac
    tva_calculee = base_tva * tva_rate
    
    # âœ… Calcul du PTTC
    # pttc = prime_nette + accessoires + asac + tva + vignette + carte_rose
    pttc_calcule = prime_nette + accessoires + asac + tva_calculee + vignette + carte_rose

    return {
        'immatriculation': vehicle['immatriculation'],
        'chassis': vehicle.get('chassis') or f"CH-{vehicle['immatriculation']}",
        'marque': vehicle.get('marque', ''),
        'modele': vehicle.get('modele', ''),
        'annee': vehicle.get('annee', None),
        'puissance_fiscale': vehicle.get('puissance', 0),
        'places': vehicle.get('places', 5),
        'cylindree': 0,
        'ptac': 0,
        'charge_utile': 0,
        'valeur_neuf': vehicle.get('valeur_neuf', 0),
        'valeur_venale': vehicle.get('valeur_venale', 0),
        'has_remorque': False,
        'remorque_inflammable': False,
        'remorque_immat': '',
        'double_commande': False,
        'engin_portuaire': False,
        'rc_eleves': False,
        'code_tarif': '',
        'libele_tarif': '',
        'code_assure': '',
        'date_debut': debut,
        'date_fin': fin,
        'date_mise_circulation': vehicle.get('date_mise_circulation', None),
        'nbr_jour': jours,
        'prime_brute': prime_brute,  # âœ… Prime brute du fichier
        'reduction': reductions,      # âœ… RÃ©ductions du fichier
        'prime_nette': prime_nette,   # âœ… Prime nette du fichier
        'prime_emise': prime_nette,   # âœ… Prime Ã©mise = prime nette
        'accessoires': accessoires,
        'tva': tva_calculee,
        'fichier_asac': asac,
        'carte_rose': carte_rose,
        'vignette': vignette,
        'pttc': pttc_calcule,
        'statut': 'ACTIF',
        'is_active': True,
        'owner_id': owner_id,
        'company_id': compagny_id,
        'compagny_id': compagny_id,
        'fleet_id': fleet_id,
        'tarif_id': None,
        'created_by': current_user_id,
        'created_ip': '127.0.0.1',
        'last_ip': '127.0.0.1',
        'categorie': categorie_value,
        'genre': genre_value,
        'type_vehicule': type_value,
        'usage': usage_value,
        'energie': energie_value,
        'zone': 'A',
        **guarantee_amounts,
        **guarantee_options,
        **reductions_individuelles,
        'amt_val_red_rc': 0,
        'amt_val_red_dr': 0,
        'amt_val_red_vol': 0,
        'amt_val_red_vb': 0,
        'amt_val_red_in': 0,
        'amt_val_red_bris': 0,
        'amt_val_red_ar': 0,
        'amt_val_red_dta': 0,
        'amt_val_red_ipt': 0,
        'amt_fleet_rc_val': 0,
        'amt_fleet_dr_val': 0,
        'amt_fleet_vol_val': 0,
        'amt_fleet_vb_val': 0,
        'amt_fleet_in_val': 0,
        'amt_fleet_bris_val': 0,
        'amt_fleet_ar_val': 0,
        'amt_fleet_dta_val': 0,
        'amt_fleet_ipt_val': 0,
    }

class VehicleGarantieDialog(QDialog):
    """Dialogue de personnalisation complÃ¨te d'un vÃ©hicule"""
    
    def __init__(self, vehicle, garanties, parent=None):
        super().__init__(parent)
        self.vehicle = vehicle
        self.garanties = garanties
        self.garantie_cards = {}
        self.setWindowTitle(f"Gestion du vÃ©hicule - {vehicle.get('immatriculation', 'VÃ©hicule')}")
        self.setMinimumSize(900, 750)
        self.setModal(True)
        self.setStyleSheet(STYLESHEET)
        
        self.setup_ui()
        self.load_garanties()
        self.load_vehicle_infos()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # En-tÃªte
        header = QFrame()
        header.setStyleSheet(f"""
            background: {AppColors.PRIMARY_LIGHT};
            border-radius: 10px;
            padding: 10px;
        """)
        header_layout = QHBoxLayout(header)
        
        info = QLabel(f"ðŸš— {self.vehicle.get('immatriculation')} - {self.vehicle.get('marque')} {self.vehicle.get('modele')}")
        info.setStyleSheet(f"font-size: 14px; font-weight: 700; color: {AppColors.PRIMARY_DARK};")
        header_layout.addWidget(info)
        header_layout.addStretch()
        layout.addWidget(header)
        
        # Utiliser un QTabWidget pour organiser les informations
        tab_widget = QTabWidget()
        
        # ========== ONGLET 1 : GARANTIES ==========
        garanties_tab = self.create_garanties_tab()
        tab_widget.addTab(garanties_tab, "ðŸ›¡ï¸ Garanties")
        
        # ========== ONGLET 2 : INFORMATIONS VÃ‰HICULE ==========
        vehicle_tab = self.create_vehicle_info_tab()
        tab_widget.addTab(vehicle_tab, "ðŸš— VÃ©hicule")
        
        # ========== ONGLET 3 : FRAIS ET TAXES ==========
        frais_tab = self.create_frais_tab()
        tab_widget.addTab(frais_tab, "ðŸ’° Frais & Taxes")
        
        # ========== ONGLET 4 : DATES ET STATUT ==========
        dates_tab = self.create_dates_tab()
        tab_widget.addTab(dates_tab, "ðŸ“… Dates")
        
        layout.addWidget(tab_widget)
        
        # Total gÃ©nÃ©ral
        total_frame = QFrame()
        total_frame.setStyleSheet(f"""
            background: {AppColors.SUCCESS_LIGHT};
            border-radius: 10px;
            padding: 10px;
        """)
        total_layout = QHBoxLayout(total_frame)
        total_layout.addWidget(QLabel("<b>ðŸ’° TOTAL GÃ‰NÃ‰RAL DES GARANTIES :</b>"))
        self.total_label = QLabel("0 FCFA")
        self.total_label.setStyleSheet(f"font-size: 16px; font-weight: 800; color: {AppColors.SUCCESS_DARK};")
        total_layout.addWidget(self.total_label)
        total_layout.addStretch()
        layout.addWidget(total_frame)
        
        # Boutons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        ok_btn = QPushButton("Valider")
        ok_btn.setProperty("class", "BtnSuccess")
        ok_btn.setFixedSize(120, 35)
        ok_btn.clicked.connect(self.accept)
        
        cancel_btn = QPushButton("Annuler")
        cancel_btn.setProperty("class", "BtnSecondary")
        cancel_btn.setFixedSize(120, 35)
        cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(ok_btn)
        layout.addLayout(btn_layout)
    
    def create_garanties_tab(self):
        """CrÃ©e l'onglet des garanties"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        
        content = QWidget()
        scroll_layout = QVBoxLayout(content)
        
        self.garantie_inputs = {}
        garanties_list = [
            ('rc', "RC/RTI"),
            ('dr', "DÃ©fense et Recours"),
            ('vol', "Vol/Vol partie"),
            ('vb', "Vol Braquage"),
            ('incendie', "Incendie"),
            ('bris_glace', "Bris de Glaces"),
            ('ar', "Assistance Ã  la rÃ©paration"),
            ('dta', "Dommages Tous Accidents"),
            ('ipt', "IPT + Conducteur"),
        ]
        
        for key, label in garanties_list:
            frame = QFrame()
            frame.setStyleSheet("border: 1px solid #e2e8f0; border-radius: 8px;")
            frame_layout = QHBoxLayout(frame)
            
            checkbox = QCheckBox(label)
            checkbox.setStyleSheet("font-weight: 600;")
            checkbox.toggled.connect(lambda checked, k=key: self.on_garantie_toggle(k, checked))
            frame_layout.addWidget(checkbox, 2)
            
            amount_input = QLineEdit()
            amount_input.setPlaceholderText("Montant")
            amount_input.setEnabled(False)
            amount_input.textChanged.connect(lambda: self.update_total())
            frame_layout.addWidget(amount_input)
            
            reduction_input = QLineEdit()
            reduction_input.setPlaceholderText("RÃ©duction %")
            reduction_input.setEnabled(False)
            reduction_input.setFixedWidth(100)
            reduction_input.textChanged.connect(lambda: self.update_total())
            frame_layout.addWidget(reduction_input)
            
            net_label = QLabel("0 FCFA")
            net_label.setStyleSheet(f"color: {AppColors.SUCCESS}; font-weight: bold;")
            net_label.setFixedWidth(120)
            frame_layout.addWidget(net_label)
            
            self.garantie_inputs[key] = {
                'checkbox': checkbox,
                'amount': amount_input,
                'reduction': reduction_input,
                'net': net_label
            }
            
            scroll_layout.addWidget(frame)
        
        scroll_layout.addStretch()
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        return tab

    def create_vehicle_info_tab(self):
        """CrÃ©e l'onglet des informations du vÃ©hicule"""
        tab = QWidget()
        layout = QGridLayout(tab)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Styles
        style_field = """
            QLineEdit {
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 8px 12px;
                background: white;
                font-size: 13px;
                color: #64748b;
            }
            QLineEdit:focus {
                border-color: #2563eb;
            }
        """
        self.info_inputs = {}
        
        # Identifiants
        layout.addWidget(QLabel("Immatriculation :"), 0, 0)
        self.info_immatriculation = QLineEdit(self.vehicle.get('immatriculation', ''))
        self.info_immatriculation.setStyleSheet(style_field)
        layout.addWidget(self.info_immatriculation, 0, 1)
        
        layout.addWidget(QLabel("ChÃ¢ssis :"), 1, 0)
        self.info_chassis = QLineEdit(self.vehicle.get('chassis', ''))
        self.info_chassis.setStyleSheet(style_field)
        layout.addWidget(self.info_chassis, 1, 1)
        
        # Marque et modÃ¨le
        layout.addWidget(QLabel("Marque :"), 2, 0)
        self.info_marque = QLineEdit(self.vehicle.get('marque', ''))
        self.info_marque.setStyleSheet(style_field)
        layout.addWidget(self.info_marque, 2, 1)
        
        layout.addWidget(QLabel("ModÃ¨le :"), 3, 0)
        self.info_modele = QLineEdit(self.vehicle.get('modele', ''))
        self.info_modele.setStyleSheet(style_field)
        layout.addWidget(self.info_modele, 3, 1)
        
        # CaractÃ©ristiques
        layout.addWidget(QLabel("CatÃ©gorie :"), 4, 0)
        self.info_categorie = QLineEdit(self.vehicle.get('categorie', 'VP'))
        self.info_categorie.setStyleSheet(style_field)
        layout.addWidget(self.info_categorie, 4, 1)
        
        layout.addWidget(QLabel("AnnÃ©e :"), 5, 0)
        self.info_annee = QLineEdit(str(self.vehicle.get('annee', '')) if self.vehicle.get('annee') else '')
        self.info_annee.setStyleSheet(style_field)
        layout.addWidget(self.info_annee, 5, 1)
        
        layout.addWidget(QLabel("Ã‰nergie :"), 6, 0)
        self.info_energie = QLineEdit(self.vehicle.get('energie', 'Essence'))
        self.info_energie.setStyleSheet(style_field)
        layout.addWidget(self.info_energie, 6, 1)
        
        layout.addWidget(QLabel("Puissance (CV) :"), 7, 0)
        self.info_puissance = QLineEdit(str(self.vehicle.get('puissance', 0)))
        self.info_puissance.setStyleSheet(style_field)
        layout.addWidget(self.info_puissance, 7, 1)
        
        layout.addWidget(QLabel("Places :"), 8, 0)
        self.info_places = QLineEdit(str(self.vehicle.get('places', 5)))
        self.info_places.setStyleSheet(style_field)
        layout.addWidget(self.info_places, 8, 1)
        
        # Valeurs financiÃ¨res
        layout.addWidget(QLabel("Valeur Ã  neuf (FCFA) :"), 9, 0)
        self.info_valeur_neuf = QLineEdit(f"{self.vehicle.get('valeur_neuf', 0):,.0f}".replace(",", " "))
        self.info_valeur_neuf.setStyleSheet(style_field)
        layout.addWidget(self.info_valeur_neuf, 9, 1)
        
        layout.addWidget(QLabel("Valeur vÃ©nale (FCFA) :"), 10, 0)
        self.info_valeur_venale = QLineEdit(f"{self.vehicle.get('valeur_venale', 0):,.0f}".replace(",", " "))
        self.info_valeur_venale.setStyleSheet(style_field)
        layout.addWidget(self.info_valeur_venale, 10, 1)
        
        layout.addWidget(QLabel("Zone :"), 11, 0)
        self.info_zone = QLineEdit(self.vehicle.get('zone', 'A'))
        self.info_zone.setStyleSheet(style_field)
        layout.addWidget(self.info_zone, 11, 1)
        
        # Remplacer layout.addStretch() par rowStretch
        layout.setRowStretch(12, 1)
        
        return tab

    def create_frais_tab(self):
        """CrÃ©e l'onglet des frais et taxes"""
        tab = QWidget()
        layout = QGridLayout(tab)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        style_field = """
            QLineEdit {
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 8px 12px;
                background: white;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #2563eb;
            }
        """
        
        # Frais supplÃ©mentaires
        layout.addWidget(QLabel("Accessoires (FCFA) :"), 0, 0)
        self.frais_accessoires = QLineEdit(str(self.vehicle.get('accessoires', 0)))
        self.frais_accessoires.setStyleSheet(style_field)
        self.frais_accessoires.textChanged.connect(self.calculate_tva_pttc)
        layout.addWidget(self.frais_accessoires, 0, 1)
        
        layout.addWidget(QLabel("Fichier ASAC (FCFA) :"), 1, 0)
        self.frais_asac = QLineEdit(str(self.vehicle.get('asac', 0)))
        self.frais_asac.setStyleSheet(style_field)
        self.frais_asac.textChanged.connect(self.calculate_tva_pttc)
        layout.addWidget(self.frais_asac, 1, 1)
        
        layout.addWidget(QLabel("Carte Rose (FCFA) :"), 2, 0)
        self.frais_carte_rose = QLineEdit(str(self.vehicle.get('carte_rose', 0)))
        self.frais_carte_rose.setStyleSheet(style_field)
        self.frais_carte_rose.textChanged.connect(self.calculate_pttc)
        layout.addWidget(self.frais_carte_rose, 2, 1)
        
        layout.addWidget(QLabel("Vignette (FCFA) :"), 3, 0)
        self.frais_vignette = QLineEdit(str(self.vehicle.get('vignette', 0)))
        self.frais_vignette.setStyleSheet(style_field)
        self.frais_vignette.textChanged.connect(self.calculate_pttc)
        layout.addWidget(self.frais_vignette, 3, 1)
        
        # SÃ©parateur
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("background: #e2e8f0; margin: 10px 0;")
        layout.addWidget(sep, 4, 0, 1, 2)
        
        # RÃ©sultats des calculs
        layout.addWidget(QLabel("TVA (19.25%) :"), 5, 0)
        self.frais_tva = QLineEdit("0")
        self.frais_tva.setReadOnly(True)
        self.frais_tva.setStyleSheet("background-color: #eff6ff; border: 1px solid #bfdbfe; border-radius: 8px; padding: 8px 12px;")
        layout.addWidget(self.frais_tva, 5, 1)
        
        layout.addWidget(QLabel("PTTC (Total) :"), 6, 0)
        self.frais_pttc = QLineEdit("0")
        self.frais_pttc.setReadOnly(True)
        self.frais_pttc.setStyleSheet("background-color: #d1fae5; border: 1px solid #bbf7d0; border-radius: 8px; padding: 8px 12px; font-weight: bold;")
        layout.addWidget(self.frais_pttc, 6, 1)
        
        # Remplacer layout.addStretch() par rowStretch
        layout.setRowStretch(7, 1)
        
        return tab

    def create_dates_tab(self):
        """CrÃ©e l'onglet des dates et statut"""
        tab = QWidget()
        layout = QGridLayout(tab)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        style_field = """
            QDateEdit {
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 8px 12px;
                background: white;
                font-size: 13px;
                color: #64748b;
            }
        """
        
        # Date dÃ©but
        layout.addWidget(QLabel("Date de dÃ©but :"), 0, 0)
        self.dates_date_debut = QDateEdit()
        self.dates_date_debut.setCalendarPopup(True)
        self.dates_date_debut.setDisplayFormat("dd/MM/yyyy")
        self.dates_date_debut.dateChanged.connect(self.update_dates_calculations)
        layout.addWidget(self.dates_date_debut, 0, 1)
        
        # Date fin
        layout.addWidget(QLabel("Date de fin :"), 1, 0)
        self.dates_date_fin = QDateEdit()
        self.dates_date_fin.setCalendarPopup(True)
        self.dates_date_fin.setDisplayFormat("dd/MM/yyyy")
        self.dates_date_fin.dateChanged.connect(self.update_dates_calculations)
        layout.addWidget(self.dates_date_fin, 1, 1)
        
        # SÃ©parateur
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("background: #e2e8f0; margin: 10px 0;")
        layout.addWidget(sep, 2, 0, 1, 2)
        
        # DurÃ©e
        layout.addWidget(QLabel("DurÃ©e (jours) :"), 3, 0)
        self.dates_duree = QLabel("365")
        self.dates_duree.setStyleSheet("font-weight: bold; color: #2563eb; font-size: 14px;")
        layout.addWidget(self.dates_duree, 3, 1)
        
        # Prorata
        layout.addWidget(QLabel("Prorata :"), 4, 0)
        self.dates_prorata = QLabel("100%")
        self.dates_prorata.setStyleSheet("font-weight: bold; color: #10b981;")
        layout.addWidget(self.dates_prorata, 4, 1)
        
        # Statut
        layout.addWidget(QLabel("Statut :"), 5, 0)
        self.dates_statut = QLabel("En Circulation")
        self.dates_statut.setStyleSheet("font-weight: bold;")
        layout.addWidget(self.dates_statut, 5, 1)
        
        # Remplacer layout.addStretch() par rowStretch
        layout.setRowStretch(6, 1)
        
        # Charger les dates
        self.load_dates()
        
        return tab

    def load_dates(self):
        """Charge les dates du vÃ©hicule"""
        date_debut = self.vehicle.get('date_debut')
        date_fin = self.vehicle.get('date_fin')
        
        if date_debut:
            if isinstance(date_debut, datetime):
                self.dates_date_debut.setDate(QDate(date_debut.year, date_debut.month, date_debut.day))
            elif isinstance(date_debut, QDate):
                self.dates_date_debut.setDate(date_debut)
        else:
            self.dates_date_debut.setDate(QDate.currentDate())
        
        if date_fin:
            if isinstance(date_fin, datetime):
                self.dates_date_fin.setDate(QDate(date_fin.year, date_fin.month, date_fin.day))
            elif isinstance(date_fin, QDate):
                self.dates_date_fin.setDate(date_fin)
        else:
            self.dates_date_fin.setDate(QDate.currentDate().addYears(1))
        
        self.update_dates_calculations()
    
    def update_dates_calculations(self):
        """Met Ã  jour les calculs des dates"""
        debut = self.dates_date_debut.date()
        fin = self.dates_date_fin.date()
        
        if fin >= debut:
            jours = debut.daysTo(fin)
            prorata = (jours / 365.0) * 100
            
            self.dates_duree.setText(f"{jours} jours")
            self.dates_prorata.setText(f"{prorata:.1f}%")
            
            today = QDate.currentDate()
            if fin < today:
                self.dates_statut.setText("ExpirÃ©")
                self.dates_statut.setStyleSheet("color: #ef4444; font-weight: bold;")
            elif debut > today:
                self.dates_statut.setText("Ã€ venir")
                self.dates_statut.setStyleSheet("color: #f59e0b; font-weight: bold;")
            else:
                self.dates_statut.setText("En Circulation")
                self.dates_statut.setStyleSheet("color: #10b981; font-weight: bold;")
        else:
            self.dates_duree.setText("Invalide")
            self.dates_prorata.setText("0%")
            self.dates_statut.setText("Invalide")
            self.dates_statut.setStyleSheet("color: #ef4444; font-weight: bold;")
    
    def load_vehicle_infos(self):
        """Charge les informations du vÃ©hicule"""
        # Les valeurs sont dÃ©jÃ  chargÃ©es dans les widgets via les text()
        pass

    def load_garanties(self):
        """Charge les garanties dans les cartes"""
        for key, card in self.garantie_cards.items():
            # RÃ©cupÃ©rer le montant de la garantie
            amount = self.garanties.get(key, 0)
            
            # âœ… VÃ©rifier si le montant est valide (non None et > 0)
            if amount is None:
                amount = 0
            
            # âœ… Charger la rÃ©duction si elle existe
            reduction_key = f'reduction_{key}'
            reduction = self.garanties.get(reduction_key, 0)
            if reduction is None:
                reduction = 0
            
            # âœ… DÃ©finir les valeurs dans la carte
            card.set_values(float(amount), float(reduction))
            card.garantie_changed.connect(self.on_garantie_changed)

    def on_garantie_changed(self, key, checked, net_amount):
        self.update_total()
        self.calculate_tva_pttc()
    
    def update_total(self):
        total = sum(card.get_net_amount() for card in self.garantie_cards.values())
        self.total_label.setText(f"{total:,.0f} FCFA".replace(",", " "))
        self.calculate_tva_pttc()
    
    def calculate_tva_pttc(self):
        """Calcule la TVA et le PTTC"""
        try:
            total_garanties = sum(card.get_net_amount() for card in self.garantie_cards.values())
            accessoires = self.get_float_value(self.frais_accessoires)
            asac = self.get_float_value(self.frais_asac)
            
            base_tva = total_garanties + accessoires + asac
            tva = base_tva * 0.1925
            
            self.frais_tva.setText(f"{tva:,.0f}".replace(",", " "))
            self.calculate_pttc()
        except:
            pass
    
    def calculate_pttc(self):
        """Calcule le PTTC"""
        try:
            total_garanties = sum(card.get_net_amount() for card in self.garantie_cards.values())
            accessoires = self.get_float_value(self.frais_accessoires)
            asac = self.get_float_value(self.frais_asac)
            tva = self.get_float_value(self.frais_tva)
            vignette = self.get_float_value(self.frais_vignette)
            carte_rose = self.get_float_value(self.frais_carte_rose)
            
            pttc = total_garanties + accessoires + asac + tva + vignette + carte_rose
            self.frais_pttc.setText(f"{pttc:,.0f}".replace(",", " "))
        except:
            pass
    
    def get_float_value(self, widget):
        """RÃ©cupÃ¨re une valeur float d'un widget"""
        try:
            if not widget or not widget.text():
                return 0.0
            txt = widget.text().strip().replace(" ", "").replace(",", ".")
            return float(txt) if txt else 0.0
        except:
            return 0.0
    
    def get_data(self):
        """Retourne les donnÃ©es modifiÃ©es"""
        data = {}
        
        # Garanties
        for key, inputs in self.garantie_inputs.items():
            if inputs['checkbox'].isChecked():
                try:
                    amount_text = inputs['amount'].text().replace(" ", "").replace(",", "")
                    amount = float(amount_text) if amount_text else 0
                    data[key] = amount
                    
                    reduction_text = inputs['reduction'].text().replace("%", "").strip()
                    reduction = float(reduction_text) if reduction_text else 0
                    data[f'reduction_{key}'] = reduction
                except:
                    data[key] = 0
                    data[f'reduction_{key}'] = 0
            else:
                data[key] = 0
                data[f'reduction_{key}'] = 0
        
        # âœ… Utiliser self.info_inputs pour rÃ©cupÃ©rer les valeurs
        for key, input_widget in self.info_inputs.items():
            value = input_widget.text().strip()
            if key in ['annee', 'puissance', 'places']:
                try:
                    data[key] = int(value) if value else 0
                except:
                    data[key] = 0
            elif key in ['valeur_neuf', 'valeur_venale']:
                try:
                    data[key] = float(value.replace(" ", "").replace(",", "")) if value else 0
                except:
                    data[key] = 0
            else:
                data[key] = value
        
        # âœ… Ajouter les frais
        data['accessoires'] = self.get_float_value(self.frais_accessoires)
        data['asac'] = self.get_float_value(self.frais_asac)
        data['carte_rose'] = self.get_float_value(self.frais_carte_rose)
        data['vignette'] = self.get_float_value(self.frais_vignette)
        data['tva'] = self.get_float_value(self.frais_tva)
        data['pttc'] = self.get_float_value(self.frais_pttc)
        
        return data

    def get_garanties(self):
        result = {}
        for key, card in self.garantie_cards.items():
            result[key] = card.get_net_amount()
            result[f'brut_{key}'] = card.current_amount
            result[f'reduction_{key}'] = card.current_reduction
        result['total'] = sum(v for k, v in result.items() if not k.startswith(('brut_', 'reduction_', 'total')))
        
        # Ajouter les valeurs des frais
        result['accessoires'] = self.get_float_value(self.frais_accessoires)
        result['asac'] = self.get_float_value(self.frais_asac)
        result['carte_rose'] = self.get_float_value(self.frais_carte_rose)
        result['vignette'] = self.get_float_value(self.frais_vignette)
        result['tva'] = self.get_float_value(self.frais_tva)
        result['pttc'] = self.get_float_value(self.frais_pttc)
        
        # Ajouter les informations du vÃ©hicule
        result['immatriculation'] = self.info_immatriculation.text().strip().upper()
        result['chassis'] = self.info_chassis.text().strip()
        result['marque'] = self.info_marque.text().strip()
        result['modele'] = self.info_modele.text().strip()
        result['categorie'] = self.info_categorie.text().strip().upper()
        result['annee'] = self.get_float_value(self.info_annee)
        result['energie'] = self.info_energie.text().strip()
        result['puissance'] = self.get_float_value(self.info_puissance)
        result['places'] = int(self.get_float_value(self.info_places)) if self.get_float_value(self.info_places) else 5
        result['valeur_neuf'] = self.get_float_value(self.info_valeur_neuf)
        result['valeur_venale'] = self.get_float_value(self.info_valeur_venale)
        result['zone'] = self.info_zone.text().strip().upper()
        
        # Ajouter les dates
        result['date_debut'] = self.dates_date_debut.date().toPython()
        result['date_fin'] = self.dates_date_fin.date().toPython()
        result['nbr_jour'] = self.dates_date_debut.date().daysTo(self.dates_date_fin.date())
        result['statut'] = self.dates_statut.text()
        result['prorata'] = result['nbr_jour'] / 365.0 if result['nbr_jour'] > 0 else 0
        
        return result

# ============================================================================
# THREAD DE CALCUL
# ============================================================================

class CalculationThread(QThread):
    progress = Signal(int, int, str, dict)
    finished_signal = Signal(list)
    
    def __init__(self, controller, vehicles, params):
        super().__init__()
        self.controller = controller
        self.vehicles = vehicles
        self.params = params
        self._is_cancelled = False
    
    def cancel(self):
        self._is_cancelled = True
    
    def run(self):
        results = []
        total = len(self.vehicles)
        
        for i, vehicle in enumerate(self.vehicles):
            if self._is_cancelled:
                break
            
            try:
                garanties = self.calculate_garanties(vehicle)
                vehicle['garanties'] = garanties
                vehicle['status'] = 'success'
                results.append(vehicle)
                
                self.progress.emit(i + 1, total, vehicle['immatriculation'], garanties)
                
            except Exception as e:
                vehicle['status'] = 'error'
                vehicle['error'] = str(e)
                results.append(vehicle)
                self.progress.emit(i + 1, total, vehicle['immatriculation'], {})
        
        self.finished_signal.emit(results)

    def calculate_garanties(self, vehicle):
        """
        Calcule les garanties pour un vÃ©hicule
        Reprend la logique de calculate_garantie_amount de automobile_form_view.py
        """
        garanties = {}
        
        # RÃ©cupÃ©rer les paramÃ¨tres
        cie_id = self.params.get('compagny_id')
        zone = self.params.get('zone', 'A')
        code_tarif = self.params.get('code_tarif', '')
        avec_remorque = self.params.get('avec_remorque', False)
        
        # Valeurs du vÃ©hicule
        v_neuf = vehicle.get('valeur_neuf', 0)
        v_venale = vehicle.get('valeur_venale', 0)
        places = vehicle.get('places', 5)
        puissance = vehicle.get('puissance', 0)
        energie = vehicle.get('energie', 'Essence')
        categorie = vehicle.get('categorie', 'VP')
        
        # Calcul du prorata
        jours = self.params.get('duree_jours', 365)
        prorata = jours / 365.0
        
        # âœ… 1. CALCUL DE LA RC (ResponsabilitÃ© Civile)
        # Appel Ã  la mÃ©thode du contrÃ´leur pour obtenir la prime RC
        try:
            res = self.controller.vehicles.get_rc_premium_from_matrix(
                cie_id=cie_id,
                zone_saisie=zone,
                categorie=categorie,
                energie=energie,
                cv_saisi=puissance,
                avec_remorque=avec_remorque,
                code_tarif=code_tarif if code_tarif else None
            )
            rc_base = res.get('rc', 0) if isinstance(res, dict) else (res or 0)
        except Exception as e:
            print(f"Erreur calcul RC: {e}")
            rc_base = 0
        
        # âœ… 2. CALCUL DES GARANTIES (comme dans automobile_form_view.py)
        # RC
        garanties['rc'] = rc_base * prorata
        
        # DR (DÃ©fense et Recours) - 3% de la RC
        garanties['dr'] = (rc_base * prorata) * 0.03 * prorata
        
        # VOL - 2% de la valeur vÃ©nale
        garanties['vol'] = v_venale * 0.02 * prorata
        
        # VB (Vol Ã  main armÃ©e) - 2% de la valeur vÃ©nale
        garanties['vb'] = v_venale * 0.02 * prorata
        
        # INCENDIE - 2.5% de la valeur vÃ©nale
        garanties['incendie'] = v_venale * 0.025 * prorata
        
        # BRIS DE GLACE - 0.5% de la valeur Ã  neuf
        garanties['bris_glace'] = v_neuf * 0.005 * prorata
        
        # AR (Assistance RÃ©paration) - 3% de 75% de la valeur vÃ©nale
        garanties['ar'] = v_venale * 0.75 * 0.03 * prorata
        
        # DTA (Dommages Tous Accidents) - 5% de la valeur Ã  neuf
        garanties['dta'] = v_neuf * 0.05 * prorata
        
        # IPT (Individuelle Personnes TransportÃ©es)
        # Si places <= 5: 7500 FCFA, sinon: (7500 * places / 5)
        if places <= 5:
            garanties['ipt'] = 7500 * prorata
        else:
            garanties['ipt'] = (7500 * places / 5) * prorata
        
        # âœ… 3. TOTAL
        garanties['total'] = sum([
            garanties.get('rc', 0),
            garanties.get('dr', 0),
            garanties.get('vol', 0),
            garanties.get('vb', 0),
            garanties.get('incendie', 0),
            garanties.get('bris_glace', 0),
            garanties.get('ar', 0),
            garanties.get('dta', 0),
            garanties.get('ipt', 0)
        ])
        
        return garanties
# ============================================================================
# DIALOGUE PRINCIPAL
# ============================================================================

class VehicleActionsWidget(QWidget):
    """Widget contenant les deux boutons d'action pour un vÃ©hicule"""
    
    def __init__(self, vehicle_id, on_garanties_clicked, on_dates_clicked, on_modified_clicked, parent=None):
        super().__init__(parent)
        self.vehicle_id = vehicle_id
        self.on_garanties_clicked = on_garanties_clicked
        self.on_dates_clicked = on_dates_clicked
        self.on_modified_clicked = on_modified_clicked
        
        self.setFixedHeight(40)
        self.setMinimumWidth(110)
        
        self.setup_ui()
    
    def setup_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(4, 2, 4, 2)
        layout.setSpacing(6)
        
        # Bouton Garanties
        self.garanties_btn = QPushButton("ðŸŽ¯")
        self.garanties_btn.setToolTip("Modifier les garanties")
        self.garanties_btn.setFixedSize(38, 34)
        self.garanties_btn.setCursor(Qt.PointingHandCursor)
        self.garanties_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
        """)
        self.garanties_btn.clicked.connect(lambda: self.on_garanties_clicked(self.vehicle_id))
        
        # Bouton Dates
        self.dates_btn = QPushButton("ðŸ“…")
        self.dates_btn.setToolTip("Modifier les dates du contrat")
        self.dates_btn.setFixedSize(38, 34)
        self.dates_btn.setCursor(Qt.PointingHandCursor)
        self.dates_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        self.dates_btn.clicked.connect(lambda: self.on_dates_clicked(self.vehicle_id))

        # Bouton Dates
        self.update_btn = QPushButton("âœï¸")
        self.update_btn.setToolTip("Modifier les donnÃ©es du vÃ©hicule")
        self.update_btn.setFixedSize(38, 34)
        self.update_btn.setCursor(Qt.PointingHandCursor)
        self.update_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        self.update_btn.clicked.connect(lambda: self.on_modified_clicked(self.vehicle_id))
        
        layout.addWidget(self.garanties_btn)
        layout.addWidget(self.dates_btn)
        layout.addWidget(self.update_btn)
        layout.setAlignment(Qt.AlignCenter)


# ============================================================================
# VEHICLE DATES DIALOG
# ============================================================================

class VehicleDatesDialog(QDialog):
    """Dialogue pour modifier les dates d'un vÃ©hicule"""
    
    def __init__(self, vehicle, parent=None):
        super().__init__(parent)
        self.vehicle = vehicle
        self.setWindowTitle(f"Dates du contrat - {vehicle.get('immatriculation', 'VÃ©hicule')}")
        self.setMinimumSize(450, 350)
        self.setModal(True)
        self.setStyleSheet(STYLESHEET)
        
        self.setup_ui()
        self.load_dates()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # En-tÃªte
        header = QFrame()
        header.setStyleSheet(f"""
            background: {AppColors.PRIMARY_LIGHT};
            border-radius: 10px;
        """)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(15, 12, 15, 12)
        
        info = QLabel(f"ðŸ“… {self.vehicle.get('immatriculation')} - {self.vehicle.get('marque')} {self.vehicle.get('modele')}")
        info.setStyleSheet(f"font-size: 14px; font-weight: 700; color: {AppColors.PRIMARY_DARK};")
        header_layout.addWidget(info)
        layout.addWidget(header)
        
        # Formulaire
        form_group = QGroupBox("ParamÃ¨tres du contrat")
        form_layout = QGridLayout(form_group)
        form_layout.setSpacing(15)
        form_layout.setContentsMargins(15, 20, 15, 15)
        
        # Date dÃ©but
        form_layout.addWidget(QLabel("ðŸ“… Date de dÃ©but :"), 0, 0)
        self.date_debut = QDateEdit()
        self.date_debut.setCalendarPopup(True)
        self.date_debut.setDisplayFormat("dd/MM/yyyy")
        self.date_debut.dateChanged.connect(self.update_calculations)
        form_layout.addWidget(self.date_debut, 0, 1)
        
        # Date fin
        form_layout.addWidget(QLabel("ðŸ“… Date de fin :"), 1, 0)
        self.date_fin = QDateEdit()
        self.date_fin.setCalendarPopup(True)
        self.date_fin.setDisplayFormat("dd/MM/yyyy")
        self.date_fin.dateChanged.connect(self.update_calculations)
        form_layout.addWidget(self.date_fin, 1, 1)
        
        # SÃ©parateur
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setStyleSheet(f"background-color: {AppColors.BORDER};")
        form_layout.addWidget(line, 2, 0, 1, 2)
        
        # DurÃ©e calculÃ©e
        form_layout.addWidget(QLabel("â±ï¸ DurÃ©e :"), 3, 0)
        self.duree_label = QLabel("365 jours")
        self.duree_label.setStyleSheet(f"font-weight: bold; color: {AppColors.PRIMARY}; font-size: 14px;")
        form_layout.addWidget(self.duree_label, 3, 1)
        
        # Statut
        form_layout.addWidget(QLabel("ðŸ“Œ Statut :"), 4, 0)
        self.statut_label = QLabel("En Circulation")
        self.statut_label.setStyleSheet(f"font-weight: bold;")
        form_layout.addWidget(self.statut_label, 4, 1)
        
        # Prorata
        form_layout.addWidget(QLabel("ðŸ’° Prorata :"), 5, 0)
        self.prorata_label = QLabel("100%")
        self.prorata_label.setStyleSheet(f"color: {AppColors.SUCCESS}; font-weight: bold;")
        form_layout.addWidget(self.prorata_label, 5, 1)
        
        layout.addWidget(form_group)
        
        # Boutons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("Annuler")
        cancel_btn.setProperty("class", "BtnSecondary")
        cancel_btn.setFixedSize(120, 35)
        cancel_btn.clicked.connect(self.reject)
        
        save_btn = QPushButton("ðŸ’¾ Sauvegarder")
        save_btn.setProperty("class", "BtnSuccess")
        save_btn.setFixedSize(120, 35)
        save_btn.clicked.connect(self.accept)
        
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)
    
    def load_dates(self):
        """Charge les dates existantes du vÃ©hicule"""
        date_debut = self.vehicle.get('date_debut')
        if date_debut:
            if isinstance(date_debut, QDate):
                self.date_debut.setDate(date_debut)
            elif isinstance(date_debut, datetime):
                self.date_debut.setDate(QDate(date_debut.year, date_debut.month, date_debut.day))
            elif isinstance(date_debut, str):
                try:
                    d = datetime.strptime(date_debut, "%Y-%m-%d")
                    self.date_debut.setDate(QDate(d.year, d.month, d.day))
                except:
                    self.date_debut.setDate(QDate.currentDate())
        else:
            self.date_debut.setDate(QDate.currentDate())
        
        date_fin = self.vehicle.get('date_fin')
        if date_fin:
            if isinstance(date_fin, QDate):
                self.date_fin.setDate(date_fin)
            elif isinstance(date_fin, datetime):
                self.date_fin.setDate(QDate(date_fin.year, date_fin.month, date_fin.day))
            elif isinstance(date_fin, str):
                try:
                    d = datetime.strptime(date_fin, "%Y-%m-%d")
                    self.date_fin.setDate(QDate(d.year, d.month, d.day))
                except:
                    self.date_fin.setDate(QDate.currentDate().addYears(1))
        else:
            self.date_fin.setDate(QDate.currentDate().addYears(1))
        
        self.update_calculations()
    
    def update_calculations(self):
        """Calcule et affiche la durÃ©e, le statut et le prorata"""
        debut = self.date_debut.date()
        fin = self.date_fin.date()
        
        if fin >= debut:
            jours = debut.daysTo(fin)
            self.duree_label.setText(f"{jours} jours ({jours/30:.1f} mois)")
            
            prorata = (jours / 365.0) * 100
            self.prorata_label.setText(f"{prorata:.1f}% de la prime annuelle")
            
            today = QDate.currentDate()
            if fin < today:
                self.statut_label.setText("â° ExpirÃ©")
                self.statut_label.setStyleSheet(f"color: {AppColors.DANGER}; font-weight: bold;")
            elif debut > today:
                self.statut_label.setText("â³ Ã€ venir")
                self.statut_label.setStyleSheet(f"color: {AppColors.WARNING}; font-weight: bold;")
            else:
                self.statut_label.setText("âœ… En Circulation")
                self.statut_label.setStyleSheet(f"color: {AppColors.SUCCESS}; font-weight: bold;")
        else:
            self.duree_label.setText("Date de fin invalide")
            self.prorata_label.setText("0%")
            self.statut_label.setText("âŒ Invalide")
            self.statut_label.setStyleSheet(f"color: {AppColors.DANGER}; font-weight: bold;")
    
    def get_dates(self):
        """Retourne les dates sÃ©lectionnÃ©es"""
        debut = self.date_debut.date().toPython()
        fin = self.date_fin.date().toPython()
        jours = self.date_debut.date().daysTo(self.date_fin.date())
        prorata = jours / 365.0 if jours > 0 else 0
        
        today = datetime.now().date()
        if fin < today:
            statut = "ExpirÃ©"
        elif debut > today:
            statut = "Ã€ venir"
        else:
            statut = "En Circulation"
        
        return {
            'date_debut': debut,
            'date_fin': fin,
            'nbr_jour': jours,
            'prorata': prorata,
            'statut': statut
        }

# ============================================================================
# DIALOGUE PRINCIPAL
# ============================================================================

class FleetImportAdvancedDialog(QDialog):
    """Dialogue d'importation de flotte avec nouvelles fonctionnalitÃ©s"""
    
    # Signal pour le rafraÃ®chissement des donnÃ©es
    data_changed = Signal()
    
    def __init__(self, controller, parent=None):
        super().__init__(parent)
        self.controller = controller
        self.setWindowTitle("Importation de flotte")
        self.setMinimumSize(1600, 1000)
        self.resize(1700, 1050)
        
        self.vehicles_data = []
        self.file_path = None
        self.calculation_thread = None
        
        self.setup_ui()
        self.setStyleSheet(STYLESHEET)
        
        self._load_compagnies()
    
    def setup_ui(self):
        """Configuration de l'interface"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        # En-tÃªte
        header = self.create_header()
        main_layout.addWidget(header)
        
        # Splitter principal
        main_splitter = QSplitter(Qt.Horizontal)
        main_splitter.setHandleWidth(1)
        
        # Panneau gauche - Configuration
        left_panel = self.create_left_panel()
        main_splitter.addWidget(left_panel)
        
        # Panneau droit - RÃ©sultats
        right_panel = self.create_right_panel()
        main_splitter.addWidget(right_panel)
        
        main_splitter.setSizes([350, 1250])
        main_layout.addWidget(main_splitter, 1)
        
        # Footer
        footer = self.create_footer()
        main_layout.addWidget(footer)
    
    def create_header(self):
        """CrÃ©e l'en-tÃªte"""
        header = QWidget()
        header.setStyleSheet(f"""
            background: {AppColors.WHITE};
            border-radius: 12px;
            padding: 15px;
        """)
        
        layout = QHBoxLayout(header)
        
        title = QLabel("ðŸšš Importation de flotte")
        title.setStyleSheet("font-size: 20px; font-weight: 800; color: #0f172a;")
        
        subtitle = QLabel("Importez vos vÃ©hicules en masse avec calcul automatique des garanties")
        subtitle.setStyleSheet(f"color: {AppColors.GRAY}; font-size: 12px;")
        
        layout.addWidget(title)
        layout.addWidget(subtitle)
        layout.addStretch()
        
        return header
    
    def create_left_panel(self):
        """CrÃ©e le panneau de configuration (scrollable)"""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        
        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setSpacing(16)
        
        # 1. Fichier
        layout.addWidget(self.create_file_section())
        
        # 2. Flotte
        layout.addWidget(self.create_fleet_section())
        
        # 3. ParamÃ¨tres d'assurance (simplifiÃ©)
        layout.addWidget(self.create_insurance_section())
        
        # Bouton Appliquer les frais
        apply_frais_btn = QPushButton("ðŸ’° Appliquer les frais aux vÃ©hicules sÃ©lectionnÃ©s")
        apply_frais_btn.setProperty("class", "BtnPrimary")
        apply_frais_btn.clicked.connect(self.apply_global_frais_to_selected)
        layout.addWidget(apply_frais_btn)
        
        layout.addStretch()
        
        scroll.setWidget(content)
        return scroll
    
    def create_file_section(self):
        """Section de sÃ©lection du fichier"""
        group = QGroupBox("ðŸ“„ 1. Fichier d'importation")
        layout = QVBoxLayout(group)
        
        # Zone de dÃ©pÃ´t
        drop_zone = QFrame()
        drop_zone.setStyleSheet(f"""
            QFrame {{
                background: {AppColors.GRAY_LIGHT};
                border: 2px dashed {AppColors.BORDER};
                border-radius: 10px;
            }}
        """)
        drop_zone.setFixedHeight(80)
        drop_zone.mousePressEvent = lambda e: self.select_file()
        
        drop_layout = QVBoxLayout(drop_zone)
        drop_layout.setAlignment(Qt.AlignCenter)
        
        self.file_icon = QLabel("ðŸ“‚")
        self.file_icon.setStyleSheet("font-size: 20px;")
        self.file_label = QLabel("Cliquez pour sÃ©lectionner un fichier Excel ou CSV")
        self.file_label.setStyleSheet(f"color: {AppColors.GRAY}; font-size: 11px;")
        self.file_info = QLabel("")
        self.file_info.setStyleSheet(f"color: {AppColors.PRIMARY}; font-size: 10px;")
        
        drop_layout.addWidget(self.file_icon)
        drop_layout.addWidget(self.file_label)
        drop_layout.addWidget(self.file_info)
        
        layout.addWidget(drop_zone)
        
        # Lien template
        template_btn = QPushButton("ðŸ“¥ TÃ©lÃ©charger le modÃ¨le Excel")
        template_btn.setFlat(True)
        template_btn.setCursor(Qt.PointingHandCursor)
        template_btn.setStyleSheet(f"color: {AppColors.PRIMARY}; text-align: left; font-size: 11px;")
        template_btn.clicked.connect(self.download_template)
        layout.addWidget(template_btn)
        
        return group
    
    def create_fleet_section(self):
        """Section de configuration de la flotte"""
        group = QGroupBox("ðŸ¢ 2. Configuration de la flotte")
        layout = QVBoxLayout(group)
        
        # Mode
        mode_layout = QHBoxLayout()
        self.mode_new = QRadioButton("âœ¨ CrÃ©er une nouvelle flotte")
        self.mode_new.setStyleSheet("color: #0f172a; font-weight: bold;")
        self.mode_existing = QRadioButton("ðŸ“¦ Ajouter Ã  une flotte existante")
        self.mode_existing.setStyleSheet("color: #0f172a; font-weight: bold;")
        self.mode_new.setChecked(True)
        self.mode_new.toggled.connect(self.on_mode_changed)
        
        mode_layout.addWidget(self.mode_new)
        mode_layout.addWidget(self.mode_existing)
        mode_layout.addStretch()
        layout.addLayout(mode_layout)
        
        # Nouvelle flotte
        self.new_fleet_widget = QWidget()
        new_layout = QGridLayout(self.new_fleet_widget)
        new_layout.setSpacing(8)
        
        new_layout.addWidget(QLabel("Nom :"), 0, 0)
        self.fleet_name = QLineEdit()
        self.fleet_name.setPlaceholderText("Ex: Flotte Logistique 2024")
        self.fleet_name.setStyleSheet("color: #0f172a;")
        new_layout.addWidget(self.fleet_name, 0, 1)
        
        new_layout.addWidget(QLabel("Code :"), 1, 0)
        self.fleet_code = QLineEdit()
        self.fleet_code.setPlaceholderText("Ex: FL-2024-001")
        self.fleet_code.setStyleSheet("color: #0f172a; font-weight: bold;")
        new_layout.addWidget(self.fleet_code, 1, 1)
        
        layout.addWidget(self.new_fleet_widget)
        
        # Flotte existante
        self.existing_fleet_widget = QWidget()
        self.existing_fleet_widget.setVisible(False)
        existing_layout = QHBoxLayout(self.existing_fleet_widget)
        existing_layout.addWidget(QLabel("SÃ©lectionner :"))
        self.existing_fleet_combo = QComboBox()
        self.existing_fleet_combo.setStyleSheet("color: #0f172a; font-weight: bold;")
        existing_layout.addWidget(self.existing_fleet_combo, 1)
        layout.addWidget(self.existing_fleet_widget)
        
        return group
        
    def create_insurance_section(self):
        """Section des paramÃ¨tres d'assurance (simplifiÃ©e)"""
        group = QGroupBox("ðŸ“‹ 3. ParamÃ¨tres du contrat")
        layout = QGridLayout(group)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 20, 15, 15)
        
        # Date dÃ©but
        layout.addWidget(QLabel("ðŸ“… Date dÃ©but :"), 0, 0)
        self.date_debut = QDateEdit()
        self.date_debut.setDate(QDate.currentDate())
        self.date_debut.setCalendarPopup(True)
        self.date_debut.setDisplayFormat("dd/MM/yyyy")
        layout.addWidget(self.date_debut, 0, 1)
        
        # Date fin
        layout.addWidget(QLabel("ðŸ“… Date fin :"), 1, 0)
        self.date_fin = QDateEdit()
        self.date_fin.setDate(QDate.currentDate().addYears(1))
        self.date_fin.setCalendarPopup(True)
        self.date_fin.setDisplayFormat("dd/MM/yyyy")
        layout.addWidget(self.date_fin, 1, 1)
        
        # SÃ©parateur
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("background: #e2e8f0; margin: 8px 0;")
        layout.addWidget(sep, 2, 0, 1, 2)
        
        # Titre des frais
        frais_title = QLabel("ðŸ’° FRAIS SUPPLÃ‰MENTAIRES")
        frais_title.setStyleSheet("font-weight: 700; color: #475569; font-size: 12px;")
        layout.addWidget(frais_title, 3, 0, 1, 2)
        
        # Mode de calcul
        layout.addWidget(QLabel("Mode de calcul :"), 4, 0)
        self.frais_mode_combo = QComboBox()
        self.frais_mode_combo.addItems(["Par vÃ©hicule (unitaire)", "Global (rÃ©parti)"])
        self.frais_mode_combo.setStyleSheet("border: 1px solid #e2e8f0; border-radius: 8px; padding: 6px 10px;")
        self.frais_mode_combo.currentIndexChanged.connect(self.on_frais_mode_changed)
        layout.addWidget(self.frais_mode_combo, 4, 1)
        
        # Info mode
        self.frais_mode_info = QLabel("ðŸ’¡ Le montant sera appliquÃ© Ã  chaque vÃ©hicule sÃ©lectionnÃ©")
        self.frais_mode_info.setStyleSheet("color: #64748b; font-size: 10px; font-style: italic;")
        layout.addWidget(self.frais_mode_info, 5, 0, 1, 2)
        
        # Accessoires
        layout.addWidget(QLabel("Accessoires (FCFA) :"), 6, 0)
        self.global_accessoires = QLineEdit("0")
        self.global_accessoires.setStyleSheet("border: 1px solid #e2e8f0; border-radius: 8px; padding: 6px 10px;")
        self.global_accessoires.textChanged.connect(self.on_frais_changed)
        layout.addWidget(self.global_accessoires, 6, 1)
        
        # Fichier ASAC
        layout.addWidget(QLabel("Fichier ASAC (FCFA) :"), 7, 0)
        self.global_asac = QLineEdit("0")
        self.global_asac.setStyleSheet("border: 1px solid #e2e8f0; border-radius: 8px; padding: 6px 10px;")
        self.global_asac.textChanged.connect(self.on_frais_changed)
        layout.addWidget(self.global_asac, 7, 1)
        
        # Carte Rose
        layout.addWidget(QLabel("Carte Rose (FCFA) :"), 8, 0)
        self.global_carte_rose = QLineEdit("0")
        self.global_carte_rose.setStyleSheet("border: 1px solid #e2e8f0; border-radius: 8px; padding: 6px 10px;")
        self.global_carte_rose.textChanged.connect(self.on_frais_changed)
        layout.addWidget(self.global_carte_rose, 8, 1)
        
        # Vignette
        layout.addWidget(QLabel("Vignette (FCFA) :"), 9, 0)
        self.global_vignette = QLineEdit("0")
        self.global_vignette.setStyleSheet("border: 1px solid #e2e8f0; border-radius: 8px; padding: 6px 10px;")
        self.global_vignette.textChanged.connect(self.on_frais_changed)
        layout.addWidget(self.global_vignette, 9, 1)
        
        # AperÃ§u des frais
        preview_frame = QFrame()
        preview_frame.setStyleSheet("""
            QFrame {
                background: #f8fafc;
                border-radius: 6px;
                padding: 6px 10px;
                margin-top: 4px;
            }
        """)
        preview_layout = QHBoxLayout(preview_frame)
        self.frais_preview = QLabel("ðŸ’° Total frais: 0 FCFA (0 vÃ©hicules)")
        self.frais_preview.setStyleSheet("color: #475569; font-size: 11px; font-weight: 500;")
        preview_layout.addWidget(self.frais_preview)
        preview_layout.addStretch()
        layout.addWidget(preview_frame, 10, 0, 1, 2)
        
        return group

    def on_frais_mode_changed(self):
        """Met Ã  jour l'info du mode de calcul"""
        is_global = self.frais_mode_combo.currentIndex() == 1
        if is_global:
            self.frais_mode_info.setText("ðŸ’¡ Le montant sera rÃ©parti entre tous les vÃ©hicules sÃ©lectionnÃ©s")
        else:
            self.frais_mode_info.setText("ðŸ’¡ Le montant sera appliquÃ© Ã  chaque vÃ©hicule sÃ©lectionnÃ©")
        self.on_frais_changed()

    def on_frais_changed(self):
        """Met Ã  jour l'aperÃ§u des frais"""
        selected_count = 0
        for row in range(self.vehicles_table.rowCount()):
            item = self.vehicles_table.item(row, 0)
            if item and item.checkState() == Qt.Checked:
                selected_count += 1
        
        if selected_count == 0:
            self.frais_preview.setText("ðŸ’° Total frais: 0 FCFA (0 vÃ©hicules sÃ©lectionnÃ©s)")
            return
        
        try:
            accessoires = float(self.global_accessoires.text().replace(" ", "").replace(",", "") or 0)
            asac = float(self.global_asac.text().replace(" ", "").replace(",", "") or 0)
            carte_rose = float(self.global_carte_rose.text().replace(" ", "").replace(",", "") or 0)
            vignette = float(self.global_vignette.text().replace(" ", "").replace(",", "") or 0)
            
            total_frais = accessoires + asac + carte_rose + vignette
            
            is_global = self.frais_mode_combo.currentIndex() == 1
            
            if is_global:
                # Mode global: le montant est rÃ©parti entre tous les vÃ©hicules
                total_frais_global = total_frais
                par_vehicule = total_frais / selected_count if selected_count > 0 else 0
                self.frais_preview.setText(
                    f"ðŸ’° Total frais: {total_frais_global:,.0f}".replace(",", " ") + 
                    f" FCFA (rÃ©parti sur {selected_count} vÃ©hicules - {par_vehicule:,.0f}".replace(",", " ") + " FCFA/vÃ©hicule)"
                )
            else:
                # Mode unitaire: le montant est multipliÃ© par le nombre de vÃ©hicules
                total_frais_global = total_frais * selected_count
                self.frais_preview.setText(
                    f"ðŸ’° Total frais: {total_frais_global:,.0f}".replace(",", " ") + 
                    f" FCFA ({total_frais:,.0f}".replace(",", " ") + f" FCFA Ã— {selected_count} vÃ©hicules)"
                )
        except:
            pass

    def create_right_panel(self):
        """CrÃ©e le panneau des rÃ©sultats"""
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        
        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setSpacing(16)
        
        # Statut
        self.status_frame = QFrame()
        self.status_frame.setStyleSheet(f"""
            QFrame {{
                background: {AppColors.PRIMARY_LIGHT};
                border-radius: 10px;
                padding: 10px;
            }}
        """)
        status_layout = QHBoxLayout(self.status_frame)
        self.status_icon = QLabel("â³")
        self.status_icon.setStyleSheet("font-size: 18px;")
        self.status_text = QLabel("SÃ©lectionnez un fichier pour commencer")
        self.status_text.setStyleSheet(f"color: {AppColors.PRIMARY_DARK}; font-weight: 500; font-size: 13px;")
        status_layout.addWidget(self.status_icon)
        status_layout.addWidget(self.status_text)
        status_layout.addStretch()
        
        self.calc_btn = QPushButton("ðŸ”¢ Calculer les garanties")
        self.calc_btn.setProperty("class", "BtnPrimary")
        self.calc_btn.setEnabled(False)
        self.calc_btn.clicked.connect(self.start_calculation)
        status_layout.addWidget(self.calc_btn)

        self.refresh_btn = QPushButton("ðŸ”„ RafraÃ®chir")
        self.refresh_btn.setProperty("class", "BtnSecondary")
        self.refresh_btn.setEnabled(False)
        self.refresh_btn.clicked.connect(self.refresh_fleet_guarantees)
        status_layout.addWidget(self.refresh_btn)
        
        # Bouton de rÃ©duction en masse
        self.mass_reduction_btn = QPushButton("ðŸ“‰ RÃ©duction en masse")
        self.mass_reduction_btn.setProperty("class", "BtnWarning")
        self.mass_reduction_btn.setEnabled(True)
        self.mass_reduction_btn.clicked.connect(self.apply_mass_reduction)
        status_layout.addWidget(self.mass_reduction_btn)
        
        layout.addWidget(self.status_frame)
        
        # Progression
        self.progress_widget = QWidget()
        self.progress_widget.setVisible(False)
        progress_layout = QVBoxLayout(self.progress_widget)
        progress_layout.setSpacing(6)
        
        self.progress_bar = QProgressBar()
        self.progress_label = QLabel("")
        self.progress_label.setStyleSheet(f"color: {AppColors.GRAY}; font-size: 11px;")
        
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.progress_label)
        layout.addWidget(self.progress_widget)
        
        # Tableau des vÃ©hicules avec nouvelles colonnes
        self.setup_vehicle_table()
        layout.addWidget(self.vehicles_table, 1)
        
        # RÃ©capitulatif
        recap_group = self.create_recap_section()
        layout.addWidget(recap_group)
        
        scroll.setWidget(content)
        return scroll
          
    def create_recap_section(self):
        """CrÃ©e la section rÃ©capitulative avec des cartes Ã©lÃ©gantes"""
        recap_group = QGroupBox("ðŸ“Š RÃ©capitulatif")
        recap_group.setStyleSheet("""
            QGroupBox {
                font-weight: 600;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                margin-top: 10px;
                padding-top: 15px;
                background: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 16px;
                padding: 0 10px;
                color: #1e293b;
                font-size: 13px;
            }
        """)
        
        recap_layout = QGridLayout(recap_group)
        recap_layout.setSpacing(8)
        recap_layout.setContentsMargins(12, 12, 12, 12)
        
        self.recap_cards = {}
        
        # Configuration des cartes
        recap_items = [
            # Informations gÃ©nÃ©rales
            {'key': 'total', 'label': 'Total', 'icon': 'ðŸš—', 'type': 'info'},
            {'key': 'selected', 'label': 'SÃ©lectionnÃ©s', 'icon': 'âœ…', 'type': 'info'},
            
            # Garanties
            {'key': 'rc', 'label': 'RC/RTI', 'icon': 'ðŸ›¡ï¸', 'type': 'garantie'},
            {'key': 'dr', 'label': 'DR', 'icon': 'âš–ï¸', 'type': 'garantie'},
            {'key': 'vol', 'label': 'Vol', 'icon': 'ðŸš—', 'type': 'garantie'},
            {'key': 'vb', 'label': 'VB', 'icon': 'ðŸ”«', 'type': 'garantie'},
            {'key': 'incendie', 'label': 'Incendie', 'icon': 'ðŸ”¥', 'type': 'garantie'},
            {'key': 'bris_glace', 'label': 'Bris Glace', 'icon': 'ðŸªŸ', 'type': 'garantie'},
            {'key': 'ar', 'label': 'AR', 'icon': 'ðŸ”§', 'type': 'garantie'},
            {'key': 'dta', 'label': 'DTA', 'icon': 'ðŸ’¥', 'type': 'garantie'},
            {'key': 'ipt', 'label': 'IPT', 'icon': 'ðŸ‘¥', 'type': 'garantie'},
            
            # Totaux
            {'key': 'prime_brute', 'label': 'Prime brute', 'icon': 'ðŸ’°', 'type': 'total'},
            {'key': 'reductions_total', 'label': 'RÃ©ductions', 'icon': 'ðŸ“‰', 'type': 'total'},
            {'key': 'prime_nette', 'label': 'Prime nette', 'icon': 'ðŸ’³', 'type': 'total'},
            
            # Frais
            {'key': 'carte_rose', 'label': 'Carte Rose', 'icon': 'ðŸ“„', 'type': 'frais'},
            {'key': 'accessoires', 'label': 'Accessoires', 'icon': 'ðŸ”§', 'type': 'frais'},
            {'key': 'asac', 'label': 'ASAC', 'icon': 'ðŸ“', 'type': 'frais'},
            {'key': 'vignette', 'label': 'Vignette', 'icon': 'ðŸ·ï¸', 'type': 'frais'},
            {'key': 'tva', 'label': 'TVA', 'icon': 'ðŸ“Š', 'type': 'frais'},
            {'key': 'pttc', 'label': 'PTTC Total', 'icon': 'ðŸ’°', 'type': 'frais'},
        ]
        
        row = 0
        col = 0
        max_cols = 4
        
        for item in recap_items:
            key = item['key']
            label = item['label']
            icon = item['icon']
            card_type = item['type']
            
            # Carte avec style Ã©purÃ©
            card = QFrame()
            card.setStyleSheet("""
                QFrame {
                    background: #f8fafc;
                    border: 1px solid #e2e8f0;
                    border-radius: 8px;
                    padding: 8px 10px;
                }
            """)
            
            card_layout = QVBoxLayout(card)
            card_layout.setSpacing(4)
            card_layout.setContentsMargins(8, 6, 8, 6)
            
            # En-tÃªte
            header_layout = QHBoxLayout()
            header_layout.setSpacing(6)
            
            icon_label = QLabel(icon)
            icon_label.setStyleSheet("font-size: 13px;")
            
            label_widget = QLabel(label)
            label_widget.setStyleSheet("""
                color: #64748b;
                font-size: 10px;
                font-weight: 600;
            """)
            
            header_layout.addWidget(icon_label)
            header_layout.addWidget(label_widget)
            header_layout.addStretch()
            card_layout.addLayout(header_layout)
            
            if card_type == 'garantie':
                # Valeur principale
                value_widget = QLabel("0")
                value_widget.setStyleSheet("""
                    color: #0f172a;
                    font-size: 16px;
                    font-weight: 700;
                    padding: 2px 0;
                """)
                value_widget.setAlignment(Qt.AlignCenter)
                card_layout.addWidget(value_widget)
                
                # Ligne rÃ©duction
                reduction_layout = QHBoxLayout()
                reduction_layout.setSpacing(6)
                
                # Pourcentage
                pct_widget = QLabel("0%")
                pct_widget.setStyleSheet("color: #f59e0b; font-size: 11px; font-weight: 600;")
                pct_widget.setAlignment(Qt.AlignCenter)
                
                # Montant dÃ©duit
                deduit_widget = QLabel("0")
                deduit_widget.setStyleSheet("color: #ef4444; font-size: 11px; font-weight: 600;")
                deduit_widget.setAlignment(Qt.AlignCenter)
                
                reduction_layout.addWidget(pct_widget)
                reduction_layout.addStretch()
                reduction_layout.addWidget(deduit_widget)
                card_layout.addLayout(reduction_layout)
                
                self.recap_cards[key] = value_widget
                self.recap_cards[f'{key}_red'] = pct_widget
                self.recap_cards[f'{key}_deduit'] = deduit_widget
                
            else:
                # Carte simple
                value_widget = QLabel("0")
                color = "#10b981" if key == 'pttc' else "#0f172a"
                value_widget.setStyleSheet(f"""
                    color: {color};
                    font-size: 17px;
                    font-weight: 700;
                    padding: 2px 0;
                """)
                value_widget.setAlignment(Qt.AlignCenter)
                card_layout.addWidget(value_widget)
                self.recap_cards[key] = value_widget
            
            recap_layout.addWidget(card, row, col)
            
            col += 1
            if col >= max_cols:
                col = 0
                row += 1
        
        return recap_group

    def update_summary(self):
        """Met Ã  jour le rÃ©capitulatif avec les rÃ©ductions"""
        total = len(self.vehicles_data)
        selected = 0
        
        # Initialiser les totaux
        garanties_keys = ['rc', 'dr', 'vol', 'vb', 'incendie', 'bris_glace', 'ar', 'dta', 'ipt']
        selected_totals = {key: 0.0 for key in garanties_keys}
        reductions_amount = {key: 0.0 for key in garanties_keys}
        reduction_percentages = {key: [] for key in garanties_keys}
        
        # âœ… Totaux des primes (brute, nette, rÃ©ductions)
        total_prime_brute = 0
        total_prime_nette = 0
        total_reductions = 0
        
        # Totaux des frais
        total_accessoires = 0
        total_asac = 0
        total_carte_rose = 0
        total_vignette = 0
        total_tva = 0
        total_pttc = 0
        
        for row in range(self.vehicles_table.rowCount()):
            item = self.vehicles_table.item(row, 0)
            is_selected = item and item.checkState() == Qt.Checked
            
            if is_selected:
                selected += 1
            
            vehicle = self.vehicles_data[row] if row < len(self.vehicles_data) else None
            if vehicle and is_selected:
                garanties = vehicle.get('garanties', {})
                
                # Calculer pour chaque garantie
                for key in garanties_keys:
                    amount = garanties.get(key, 0)
                    selected_totals[key] += amount
                    
                    # RÃ©cupÃ©rer le pourcentage de rÃ©duction
                    reduction_key = f'reduction_{key}'
                    reduction_pct = garanties.get(reduction_key, 0)
                    if isinstance(reduction_pct, str):
                        try:
                            reduction_pct = float(reduction_pct.replace('%', '').strip())
                        except:
                            reduction_pct = 0
                    
                    if reduction_pct > 0:
                        reduction_percentages[key].append(reduction_pct)
                        deduit = amount * (reduction_pct / 100)
                        reductions_amount[key] += deduit
                
                # âœ… Additionner les primes brute et nette du vÃ©hicule
                total_prime_brute += vehicle.get('prime_brute', 0)
                total_prime_nette += vehicle.get('prime_nette', 0)
                total_reductions += vehicle.get('reductions', 0)
                
                # Frais
                total_accessoires += vehicle.get('accessoires', 0)
                total_asac += vehicle.get('asac', 0)
                total_carte_rose += vehicle.get('carte_rose', 0)
                total_vignette += vehicle.get('vignette', 0)
                total_tva += vehicle.get('tva', 0)
                total_pttc += vehicle.get('pttc', 0)
        
        # Mettre Ã  jour les cartes
        self.recap_cards['total'].setText(str(total))
        self.recap_cards['selected'].setText(str(selected))
        
        # Garanties avec leurs dÃ©tails
        for key in garanties_keys:
            # Montant total
            total_amount = selected_totals.get(key, 0)
            self.recap_cards[key].setText(f"{total_amount:,.0f}".replace(",", " "))
            
            # Pourcentage de rÃ©duction (moyenne ou dernier appliquÃ©)
            if reduction_percentages.get(key):
                avg_red = sum(reduction_percentages[key]) / len(reduction_percentages[key])
                self.recap_cards[f'{key}_red'].setText(f"{avg_red:.1f}%")
            else:
                self.recap_cards[f'{key}_red'].setText("0%")
            
            # Montant total dÃ©duit
            deduit = reductions_amount.get(key, 0)
            self.recap_cards[f'{key}_deduit'].setText(f"{deduit:,.0f}".replace(",", " "))
        
        # âœ… Totaux des primes (brute, nette, rÃ©ductions)
        self.recap_cards['prime_brute'].setText(f"{total_prime_brute:,.0f}".replace(",", " "))
        self.recap_cards['reductions_total'].setText(f"{total_reductions:,.0f}".replace(",", " "))
        self.recap_cards['prime_nette'].setText(f"{total_prime_nette:,.0f}".replace(",", " "))
        
        # Frais
        self.recap_cards['accessoires'].setText(f"{total_accessoires:,.0f}".replace(",", " "))
        self.recap_cards['asac'].setText(f"{total_asac:,.0f}".replace(",", " "))
        self.recap_cards['carte_rose'].setText(f"{total_carte_rose:,.0f}".replace(",", " "))
        self.recap_cards['vignette'].setText(f"{total_vignette:,.0f}".replace(",", " "))
        self.recap_cards['tva'].setText(f"{total_tva:,.0f}".replace(",", " "))
        self.recap_cards['pttc'].setText(f"{total_pttc:,.0f}".replace(",", " "))
        
        self.update_selection_count()

    def create_footer(self):
        """CrÃ©e le pied de page"""
        footer = QFrame()
        footer.setStyleSheet(f"border-top: 1px solid {AppColors.BORDER}; padding-top: 12px;")
        
        layout = QHBoxLayout(footer)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # Info sur les sÃ©lections
        self.selection_info = QLabel("0 vÃ©hicules sÃ©lectionnÃ©s")
        self.selection_info.setStyleSheet(f"color: {AppColors.GRAY}; font-size: 12px;")
        layout.addWidget(self.selection_info)
        
        layout.addStretch()
        
        self.cancel_btn = QPushButton("Annuler")
        self.cancel_btn.setProperty("class", "BtnSecondary")
        self.cancel_btn.setFixedSize(120, 35)
        self.cancel_btn.clicked.connect(self.reject)
        
        self.import_btn = QPushButton("âœ… Importer la flotte")
        self.import_btn.setProperty("class", "BtnSuccess")
        self.import_btn.setEnabled(False)
        self.import_btn.setFixedSize(160, 40)
        self.import_btn.clicked.connect(self.import_fleet)
        
        layout.addWidget(self.cancel_btn)
        layout.addSpacing(10)
        layout.addWidget(self.import_btn)
        
        return footer
    
    def on_table_item_clicked(self, item):
        """GÃ¨re le clic sur un Ã©lÃ©ment du tableau"""
        # Mettre Ã  jour le compteur de sÃ©lection
        self.update_selection_count()
    
    def update_selection_count(self):
        """Met Ã  jour le compteur de vÃ©hicules sÃ©lectionnÃ©s"""
        count = 0
        for row in range(self.vehicles_table.rowCount()):
            item = self.vehicles_table.item(row, 0)
            if item and item.checkState() == Qt.Checked:
                count += 1
        self.selection_info.setText(f"{count} vÃ©hicule(s) sÃ©lectionnÃ©(s)")
    
    def select_file(self):
        """SÃ©lectionne le fichier"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "SÃ©lectionner un fichier", "",
            "Excel (*.xlsx *.xls);;CSV (*.csv)"
        )
        
        if file_path:
            self.file_path = file_path
            self.file_label.setText(os.path.basename(file_path))
            self.load_file()

    # ============================================================================
    # MÃ‰THODE LOAD_FILE AMÃ‰LIORÃ‰E
    # ============================================================================

    def load_file(self):
        """
        Charge un fichier Excel/CSV avec dÃ©tection intelligente amÃ©liorÃ©e.
        Version avec mapping manuel, prÃ©visualisation, validation et gestion d'erreurs.
        """
        try:
            # 1. LECTURE DU FICHIER
            df_raw = self._read_file_with_formulas(self.file_path)
            if df_raw is None or df_raw.empty:
                self.file_info.setText("âŒ Fichier vide ou illisible")
                self.status_text.setText("Erreur: fichier vide")
                return
            
            # 2. DÃ‰TECTION AMÃ‰LIORÃ‰E DE L'EN-TÃŠTE
            header_row = self._detect_header_row(df_raw)
            if header_row == -1:
                self.file_info.setText("âŒ En-tÃªte non dÃ©tectÃ©")
                self.status_text.setText("Erreur: format de fichier non reconnu")
                return
            
            # 3. EXTRACTION DES DONNÃ‰ES
            df = self._extract_data(df_raw, header_row)
            if df is None or df.empty:
                self.file_info.setText("âŒ Aucune donnÃ©e valide")
                self.status_text.setText("Erreur: aucune donnÃ©e trouvÃ©e")
                return
            
            # 4. NETTOYAGE AMÃ‰LIORÃ‰
            df = self._clean_data(df)
            if df is None or df.empty:
                self.file_info.setText("âŒ DonnÃ©es vides aprÃ¨s nettoyage")
                self.status_text.setText("Erreur: aucune donnÃ©e valide")
                return
            
            # 5. MAPPING INTELLIGENT DES COLONNES
            df, mapping_warnings = self._map_columns_advanced(df)
            
            # 6. AFFICHAGE DE LA PRÃ‰VISUALISATION POUR CONFIRMATION
            if not self._show_preview(df, mapping_warnings):
                return  # L'utilisateur a annulÃ©
            
            # 7. EXTRACTION ET CONVERSION DES DONNÃ‰ES
            self.vehicles_data = self._extract_vehicles_data(df)
            
            if not self.vehicles_data:
                self.file_info.setText("âŒ Aucun vÃ©hicule trouvÃ©")
                self.status_text.setText("Erreur: aucun vÃ©hicule valide")
                return
            
            # 8. VALIDATION DES DONNÃ‰ES
            invalid_count = self._validate_vehicles_data()
            if invalid_count > 0:
                reply = QMessageBox.question(
                    self,
                    "âš ï¸ DonnÃ©es invalides",
                    f"{invalid_count} vÃ©hicule(s) ont des donnÃ©es invalides.\n"
                    "Voulez-vous continuer avec les donnÃ©es valides ?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.No:
                    return
            
            # 9. AFFICHAGE RÃ‰SULTATS
            self.file_info.setText(f"âœ… {len(self.vehicles_data)} vÃ©hicules trouvÃ©s")
            self.status_text.setText(f"{len(self.vehicles_data)} vÃ©hicules chargÃ©s")
            self.status_icon.setText("âœ…")
            self.calc_btn.setEnabled(True)
            self.display_vehicles()
            
        except Exception as e:
            self.file_info.setText(f"âŒ Erreur: {str(e)[:100]}")
            self.status_text.setText(f"Erreur: {str(e)}")
            self.status_icon.setText("âŒ")
            traceback.print_exc()

    # ============================================================================
    # 1. LECTURE AMÃ‰LIORÃ‰E DES FICHIERS
    # ============================================================================

    def _read_file_with_formulas(self, file_path):
        """Lit le fichier avec gestion des formules Excel et des grands fichiers"""
        try:
            if file_path.endswith(('.xlsx', '.xls')):
                import openpyxl
                from openpyxl.utils import get_column_letter
                
                # Chargement avec gestion de la mÃ©moire
                wb = openpyxl.load_workbook(
                    file_path, 
                    data_only=True,
                    read_only=True,  # Pour les gros fichiers
                    keep_vba=False
                )
                ws = wb.active
                
                # Lecture en mode itÃ©rateur pour Ã©conomiser la mÃ©moire
                data = []
                max_cols = 0
                for row in ws.iter_rows(values_only=True):
                    row_data = list(row)
                    # Nettoyer les valeurs None
                    row_data = [v if v is not None else '' for v in row_data]
                    data.append(row_data)
                    if len(row_data) > max_cols:
                        max_cols = len(row_data)
                
                wb.close()
                
                # Normaliser les lignes pour qu'elles aient toutes le mÃªme nombre de colonnes
                for i in range(len(data)):
                    if len(data[i]) < max_cols:
                        data[i].extend([''] * (max_cols - len(data[i])))
                
                return pd.DataFrame(data)
            else:
                # CSV avec dÃ©tection d'encodage
                encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
                for encoding in encodings:
                    try:
                        return pd.read_csv(
                            file_path, 
                            encoding=encoding, 
                            header=None,
                            dtype=str,  # Garder tout en string pour Ã©viter les problÃ¨mes de type
                            engine='python',  # Plus flexible pour les fichiers mal formÃ©s
                            on_bad_lines='skip'
                        )
                    except UnicodeDecodeError:
                        continue
                
                raise ValueError("Aucun encodage valide trouvÃ© pour le fichier CSV")
                
        except Exception as e:
            print(f"Erreur lecture fichier: {e}")
            return None

    # ============================================================================
    # 2. DÃ‰TECTION AMÃ‰LIORÃ‰E DE L'EN-TÃŠTE
    # ============================================================================

    def _detect_header_row(self, df):
        """DÃ©tecte la ligne d'en-tÃªte avec des mots-clÃ©s multiples"""
        # Mots-clÃ©s pour identifier l'en-tÃªte
        header_keywords = {
            'immatriculation': ['immatriculation', 'immat', 'plaque', 'matricule', 'immatricule'],
            'marque': ['marque', 'marque et type', 'type', 'modÃ¨le', 'model'],
            'puissance': ['puissance', 'cv', 'chevaux', 'fiscale'],
            'genre': ['genre', 'catÃ©gorie', 'categorie', 'type de vÃ©hicule'],
            'places': ['places', 'place', 'nb places', 'nombre de places'],
            'valeur': ['valeur', 'prix', 'montant', 'cout', 'coÃ»t'],
            'rc': ['rc', 'rti', 'responsabilitÃ©', 'responsabilite']
        }
        
        for idx, row in df.iterrows():
            row_values = [str(v).strip().lower() for v in row.values if pd.notna(v)]
            
            # Compter les correspondances
            matches = 0
            for keyword, variants in header_keywords.items():
                for variant in variants:
                    if any(variant in val for val in row_values):
                        matches += 1
                        break
            
            # Si plus de 5 mots-clÃ©s sont trouvÃ©s, c'est probablement l'en-tÃªte
            if matches >= 5:
                print(f"âœ… En-tÃªte dÃ©tectÃ© Ã  la ligne {idx} avec {matches} correspondances")
                return idx
            
            # VÃ©rification spÃ©cifique pour "Rang" et "Immatriculation"
            row_text = ' '.join(row_values)
            if 'rang' in row_text and ('immatriculation' in row_text or 'plaque' in row_text):
                print(f"âœ… En-tÃªte dÃ©tectÃ© Ã  la ligne {idx} (Rang + Immatriculation)")
                return idx
        
        return -1

    # ============================================================================
    # 3. EXTRACTION DES DONNÃ‰ES
    # ============================================================================

    def _extract_data(self, df, header_row):
        """Extrait les donnÃ©es Ã  partir de la ligne d'en-tÃªte"""
        try:
            # RÃ©cupÃ©rer les en-tÃªtes
            header_values = df.iloc[header_row].values
            
            # Trouver la derniÃ¨re colonne non vide
            max_cols = 0
            for i, val in enumerate(header_values):
                if pd.notna(val) and str(val).strip():
                    max_cols = i + 1
            
            if max_cols == 0:
                max_cols = len(header_values)
            
            # CrÃ©er les en-tÃªtes normalisÃ©s
            column_names = []
            for i in range(max_cols):
                if i < len(header_values) and pd.notna(header_values[i]):
                    name = str(header_values[i]).strip()
                    # Ã‰viter les doublons
                    if name in column_names:
                        name = f"{name}_{i}"
                    column_names.append(name)
                else:
                    column_names.append(f"Col_{i}")
            
            # Extraire les donnÃ©es
            data_values = df.values[header_row+1:, :max_cols]
            
            # CrÃ©er le DataFrame
            df_data = pd.DataFrame(data_values, columns=column_names)
            
            # Nettoyer les lignes totalement vides
            df_data = df_data.dropna(how='all')
            
            return df_data
            
        except Exception as e:
            print(f"Erreur extraction donnÃ©es: {e}")
            return None

    # ============================================================================
    # 4. NETTOYAGE AMÃ‰LIORÃ‰ DES DONNÃ‰ES
    # ============================================================================

    def _clean_data(self, df):
        """Nettoie les donnÃ©es avec prÃ©servation des informations"""
        try:
            # 1. Supprimer les lignes complÃ¨tement vides
            df = df.dropna(how='all')
            
            # 2. Identifier la colonne d'immatriculation
            immat_col = None
            for col in df.columns:
                col_lower = col.lower()
                if any(k in col_lower for k in ['immatriculation', 'immat', 'plaque', 'matricule']):
                    immat_col = col
                    break
            
            if immat_col is None:
                # Essayer de trouver une colonne avec des donnÃ©es qui ressemblent Ã  des plaques
                for col in df.columns:
                    sample = df[col].dropna().astype(str).head(10)
                    if sample.str.contains(r'[A-Z0-9\-]', regex=True).any():
                        immat_col = col
                        break
            
            if immat_col is None:
                return None
            
            # 3. Nettoyer la colonne d'immatriculation
            df[immat_col] = df[immat_col].astype(str).str.strip().str.upper()
            
            # 4. Filtrer les lignes sans immatriculation
            df = df[df[immat_col] != '']
            df = df[df[immat_col] != 'NAN']
            df = df[df[immat_col] != 'NONE']
            df = df[df[immat_col] != 'N/A']
            df = df[df[immat_col] != 'NAN']
            df = df[df[immat_col].notna()]
            
            # 5. Filtrer les lignes de total (avec plusieurs mots-clÃ©s)
            total_keywords = ['prime', 'total', 'sous-total', 'rÃ©duction', 'nette', 'brute', 
                            'carburant', 'kilomÃ©trage', 'km', 'essence', 'diesel']
            mask = ~df[immat_col].str.contains('|'.join(total_keywords), case=False, na=False)
            df = df[mask]
            
            # 6. Supprimer les colonnes inutiles (toutes les valeurs sont NaN ou vides)
            df = df.dropna(axis=1, how='all')
            
            # 7. Remplacer les NaN par des chaÃ®nes vides pour les colonnes string
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].fillna('')
            
            return df
            
        except Exception as e:
            print(f"Erreur nettoyage donnÃ©es: {e}")
            return None

    # ============================================================================
    # 5. MAPPING AVANCÃ‰ DES COLONNES
    # ============================================================================

    def _map_columns_advanced(self, df):
        """Mappe les colonnes avec dÃ©tection intelligente - Version avec les nouvelles colonnes"""
        
        # âœ… MAPPING MIS Ã€ JOUR AVEC LES NOUVELLES COLONNES
        column_mapping_config = {
            'Rang': ['rang', 'nÂ°', 'numero', '#'],
            'Immatriculation': ['immatriculation', 'immat', 'plaque', 'matricule', 'immatricule'],
            'Numero Chassi': ['numÃ©ro chassi', 'numero chassi', 'chassis', 'chÃ¢ssis', 'vin'],
            'Marque': ['marque', 'marque et type', 'marque_type', 'brand'],
            'Modele': ['modele', 'modÃ¨le', 'model', 'type', 'version'],
            'Genre': ['genre', 'catÃ©gorie', 'categorie', 'type vÃ©hicule'],
            'Puissance en CV': ['puissance', 'cv', 'chevaux', 'fiscale', 'puissance fiscale'],
            'Usage': ['usage', 'utilisation', 'type usage', 'destination'],
            'Nombre de Places': ['places', 'place', 'nb places', 'nombre de places', 'capacitÃ©'],
            'Nombre de jours assurÃ©s': ['jours', 'durÃ©e', 'pÃ©riode', 'pÃ©riode assurance', 'nbr jours'],
            'PMEC': ['pmec', 'poids', 'charge', 'ptac', 'poids max'],
            'Valeur Neuve': ['valeur neuf', 'valeur_neuf', 'prix neuf', 'achat'],
            'Valeur Venale': ['valeur venale', 'valeur_venale', 'prix actuel', 'prix venal', 'venale'],
            'Capital Assistance Ã  la rÃ©paration': ['capital assistance', 'capital ar', 'assistance rÃ©paration', 'assistance reparation'],
            'RC/RTI': ['rc', 'rti', 'responsabilitÃ© civile', 'responsabilite civile'],
            'DÃ©fense et Recours': ['dÃ©fense', 'defense', 'recours', 'defense recours'],
            'Vol/Vol partiel - Braquage': ['vol/vol partiel - braquage', 'vol/vol partiel', 'vol braquage', 'vol', 'vb'],
            'Incendie': ['incendie', 'feu'],
            'Bris de Glaces': ['bris', 'brise', 'bris de glace', 'vitres', 'pare-brise'],
            'Assistance Ã  la rÃ©paration': ['assistance rÃ©paration', 'assistance reparation', 'ar', 'assistance'],
            'Dommages Tous Accidents': ['dommages', 'dta', 'tous accidents', 'tous accid'],
            'Individuelle Personnes TransportÃ©es (I.P.T) + Conducteur': ['individuelle personnes transportÃ©es', 'ipt', 'individuelle', 'conducteur', 'personnes transportÃ©es'],
            'Prime brute': ['prime brute', 'prime_brute', 'brute'],
            'RÃ©ductions': ['rÃ©ductions', 'reductions', 'remise', 'rÃ©duction'],
            'Prime nette': ['prime nette', 'prime_nette', 'nette'],
            'Droit de Timbre Automobile': ['timbre', 'droit timbre', 'droit de timbre', 'vignette', 'attestation']
        }
        
        mapping = {}
        warnings = []
        
        # 1. DÃ©tection automatique
        for target, keywords in column_mapping_config.items():
            found = False
            for col in df.columns:
                col_lower = col.lower().strip()
                for keyword in keywords:
                    if keyword in col_lower:
                        mapping[target] = col
                        found = True
                        break
                if found:
                    break
            
            if not found:
                mapping[target] = None
                warnings.append(f"âŒ Colonne non trouvÃ©e: {target}")
        
        # âœ… VÃ‰RIFICATION SPÃ‰CIALE POUR "Assistance Ã  la rÃ©paration"
        if mapping.get('Assistance Ã  la rÃ©paration') is None:
            for col in df.columns:
                col_lower = col.lower().strip()
                if any(k in col_lower for k in ['assistance', 'ar', 'assistance reparation', 'assistance rÃ©paration']):
                    mapping['Assistance Ã  la rÃ©paration'] = col
                    warnings.append(f"âœ… Colonne 'Assistance Ã  la rÃ©paration' trouvÃ©e sous: {col}")
                    break
        
        # âœ… VÃ‰RIFICATION SPÃ‰CIALE POUR "Vol/Vol partiel - Braquage"
        if mapping.get('Vol/Vol partiel - Braquage') is None:
            for col in df.columns:
                col_lower = col.lower().strip()
                if any(k in col_lower for k in ['vol', 'braquage', 'vol partiel']):
                    mapping['Vol/Vol partiel - Braquage'] = col
                    warnings.append(f"âœ… Colonne 'Vol/Vol partiel - Braquage' trouvÃ©e sous: {col}")
                    break
        
        # 2. VÃ©rification des colonnes obligatoires
        required_cols = ['Immatriculation', 'Marque et Type']
        missing_required = [c for c in required_cols if mapping.get(c) is None]
        
        if missing_required:
            warnings.append(f"âš ï¸ Colonnes obligatoires manquantes: {', '.join(missing_required)}")
            
            # Demander Ã  l'utilisateur de sÃ©lectionner manuellement
            missing_cols = self._ask_for_missing_columns(df, missing_required)
            for target, col in missing_cols.items():
                if col:
                    mapping[target] = col
                    warnings.append(f"âœ… Colonne '{target}' mappÃ©e manuellement sur '{col}'")
        
        # 3. Appliquer le renommage
        rename_dict = {v: k for k, v in mapping.items() if v is not None}
        df = df.rename(columns=rename_dict)
        
        return df, warnings

    def _ask_for_missing_columns(self, df, missing_columns):
        """Demande Ã  l'utilisateur de mapper les colonnes manquantes"""
        result = {}
        
        for col in missing_columns:
            # CrÃ©er un petit dialogue de sÃ©lection
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Mapper la colonne: {col}")
            dialog.setModal(True)
            dialog.setMinimumWidth(400)
            
            layout = QVBoxLayout(dialog)
            layout.addWidget(QLabel(f"SÃ©lectionnez la colonne correspondant Ã  '{col}' :"))
            
            combo = QComboBox()
            combo.addItem("-- Aucune --", None)
            for c in df.columns:
                combo.addItem(str(c), c)
            
            layout.addWidget(combo)
            
            # Boutons
            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            buttons.accepted.connect(dialog.accept)
            buttons.rejected.connect(dialog.reject)
            layout.addWidget(buttons)
            
            if dialog.exec() == QDialog.Accepted:
                selected = combo.currentData()
                if selected:
                    result[col] = selected
        
        return result

    # ============================================================================
    # 6. PRÃ‰VISUALISATION
    # ============================================================================

    def _show_preview(self, df, warnings):
        """Affiche une prÃ©visualisation et demande confirmation"""
        
        preview_dialog = QDialog(self)
        preview_dialog.setWindowTitle("ðŸ“‹ PrÃ©visualisation des donnÃ©es")
        preview_dialog.setModal(True)
        preview_dialog.setMinimumSize(900, 600)
        
        layout = QVBoxLayout(preview_dialog)
        
        # Avertissements
        if warnings:
            warning_frame = QFrame()
            warning_frame.setStyleSheet("background-color: #fffbeb; border: 1px solid #fcd34d; border-radius: 8px; padding: 8px;")
            warning_layout = QVBoxLayout(warning_frame)
            for warn in warnings[:5]:
                warning_layout.addWidget(QLabel(warn))
            layout.addWidget(warning_frame)
        
        # Tableau de prÃ©visualisation
        preview_table = QTableWidget()
        preview_table.setColumnCount(min(len(df.columns), 10))
        
        # En-tÃªtes
        headers = df.columns[:10].tolist()
        preview_table.setHorizontalHeaderLabels(headers)
        
        # DonnÃ©es (5 premiÃ¨res lignes)
        n_rows = min(5, len(df))
        preview_table.setRowCount(n_rows)
        for i in range(n_rows):
            for j, col in enumerate(headers):
                value = str(df.iloc[i][col]) if col in df.columns else ''
                preview_table.setItem(i, j, QTableWidgetItem(value))
        
        preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        preview_table.setAlternatingRowColors(True)
        layout.addWidget(QLabel(f"ðŸ“Š {len(df)} lignes trouvÃ©es"))
        layout.addWidget(preview_table)
        
        # Boutons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("âŒ Annuler")
        cancel_btn.clicked.connect(preview_dialog.reject)
        
        confirm_btn = QPushButton("âœ… Confirmer l'import")
        confirm_btn.setStyleSheet("background-color: #10b981; color: white; font-weight: bold;")
        confirm_btn.clicked.connect(preview_dialog.accept)
        
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(confirm_btn)
        layout.addLayout(btn_layout)
        
        return preview_dialog.exec() == QDialog.Accepted

    # def _extract_vehicles_data(self, df):
    #     """Extrait les donnÃ©es des vÃ©hicules - Version avec traitement des textes spÃ©ciaux"""
    #     vehicles = []
        
    #     # âœ… Colonnes numÃ©riques avec les nouveaux noms
    #     numeric_cols = [
    #         'Puissance en CV', 'Nombre de Places', 'Nombre de jours assurÃ©s',
    #         'Valeur Neuve', 'Valeur Venale', 'Capital Assistance Ã  la rÃ©paration',
    #         'RC/RTI', 'DÃ©fense et Recours', 'Vol/Vol partiel - Braquage',
    #         'Incendie', 'Bris de Glaces', 'Assistance Ã  la rÃ©paration',
    #         'Dommages Tous Accidents', 'Individuelle Personnes TransportÃ©es (I.P.T) + Conducteur',
    #         'Prime brute', 'RÃ©ductions', 'Prime nette', 'Droit de Timbre Automobile'
    #     ]
        
    #     for idx, row in df.iterrows():
    #         try:
    #             # Extraire l'immatriculation
    #             immat = self._safe_str(row.get('Immatriculation', ''))
    #             if not immat:
    #                 continue
                
    #             # âœ… VÃ©rifier si c'est une ligne de total (exclure les lignes de rÃ©capitulatif)
    #             if any(keyword in immat.upper() for keyword in ['PRIME', 'REDUCTIONS', 'TOTAL', 'SOUS-TOTAL']):
    #                 print(f"â­ï¸ Ligne de total ignorÃ©e: {immat}")
    #                 continue
                
    #             # âœ… Extraire la marque et le modÃ¨le (sÃ©parÃ©s)
    #             marque = self._safe_str(row.get('Marque', ''))
    #             modele = self._safe_str(row.get('Modele', ''))
                
    #             # Si la colonne Marque est vide mais que Marque et Type existe
    #             if not marque:
    #                 marque_type = self._safe_str(row.get('Marque et Type', ''))
    #                 if marque_type:
    #                     # âœ… Ne pas traiter les lignes de rÃ©capitulatif
    #                     if not any(k in marque_type.upper() for k in ['PRIME', 'REDUCTIONS', 'TOTAL']):
    #                         parts = marque_type.split(' ', 1)
    #                         marque = parts[0] if parts else ''
    #                         modele = parts[1] if len(parts) > 1 else ''
                
    #             # Extraire les valeurs numÃ©riques
    #             values = {}
    #             for col in numeric_cols:
    #                 if col in row.index:
    #                     val = row[col]
    #                     if hasattr(val, 'iloc'):
    #                         val = val.iloc[0] if len(val) > 0 else 0
    #                     # âœ… Utiliser la nouvelle mÃ©thode safe_float qui gÃ¨re "INCLUS"
    #                     values[col] = self._safe_float(val)
    #                 else:
    #                     values[col] = 0
                
    #             # âœ… EXTRAIRE LA PUISSANCE ET L'Ã‰NERGIE DE LA COLONNE "Puissance en CV"
    #             puissance_raw = row.get('Puissance en CV', '')
    #             if hasattr(puissance_raw, 'iloc'):
    #                 puissance_raw = puissance_raw.iloc[0] if len(puissance_raw) > 0 else ''
                
    #             puissance, energie = self._extract_power_and_energy(puissance_raw)
                
    #             # âœ… RÃ‰CUPÃ‰RATION CORRECTE DE "Assistance Ã  la rÃ©paration"
    #             ar_value = 0
                
    #             #1. Essayer de rÃ©cupÃ©rer depuis la colonne "Assistance Ã  la rÃ©paration"
    #             if 'Assistance Ã  la rÃ©paration' in row.index:
    #                 val = row['Assistance Ã  la rÃ©paration']
    #                 if hasattr(val, 'iloc'):
    #                     val = val.iloc[0] if len(val) > 0 else 0
    #                 ar_value = self._safe_float(val)
    #                 if ar_value > 0:
    #                     print(f"âœ… AR trouvÃ©e dans colonne 'Assistance Ã  la rÃ©paration': {ar_value}")
                
    #             if 'Assistance Ã  la rÃ©paration' in row.index:
    #                 val = row['Prime brute']
    #                 if hasattr(val, 'iloc'):
    #                     val = val.iloc[0] if len(val) > 0 else 0
    #                 pb_value = self._safe_float(val)
    #                 if pb_value > 0:
    #                     print(f"âœ… PB trouvÃ©e dans colonne 'Prime brute': {pb_value}")

    #             # Extraire les dates
    #             pmec_val = row.get('PMEC')
    #             if hasattr(pmec_val, 'iloc'):
    #                 pmec_val = pmec_val.iloc[0] if len(pmec_val) > 0 else None
    #             pmec_date = self._extract_date_smart(pmec_val)
                
    #             # RÃ©cupÃ©rer le rang
    #             rang_val = row.get('Rang')
    #             if hasattr(rang_val, 'iloc'):
    #                 rang_val = rang_val.iloc[0] if len(rang_val) > 0 else idx + 1
                
    #             vehicle = {
    #                 'id': idx + 1,
    #                 'rang': self._safe_int(rang_val) or (idx + 1),
    #                 'immatriculation': immat,
    #                 'chassis': self._safe_str(row.get('Numero Chassi', '')),
    #                 'marque': marque,
    #                 'modele': modele,
    #                 'marque_type': f"{marque} {modele}".strip(),
    #                 'genre': self._safe_str(row.get('Genre', 'VP')),
    #                 'puissance': puissance,
    #                 'energie': energie,
    #                 'usage': self._safe_str(row.get('Usage', '')),
    #                 'places': int(values.get('Nombre de Places', 5)),
    #                 'nbr_jour': int(values.get('Nombre de jours assurÃ©s', 365)),
    #                 'pmec': pmec_date,
    #                 'valeur_neuf': values.get('Valeur Neuve', 0),
    #                 'valeur_venale': values.get('Valeur Venale', 0),
    #                 'capital_ar': values.get('Capital Assistance Ã  la rÃ©paration', 0),
    #                 'garanties': {
    #                     'rc': values.get('RC/RTI', 0),
    #                     'dr': values.get('DÃ©fense et Recours', 0),
    #                     'vol': values.get('Vol/Vol partiel - Braquage', 0),
    #                     'vb': values.get('Vol/Vol partiel - Braquage', 0),
    #                     'incendie': values.get('Incendie', 0),
    #                     'bris_glace': values.get('Bris de Glaces', 0),
    #                     'ar': ar_value,  # âœ… Utiliser la valeur corrigÃ©e
    #                     'dta': values.get('Dommages Tous Accidents', 0),
    #                     'ipt': values.get('Individuelle Personnes TransportÃ©es (I.P.T) + Conducteur', 0),
    #                 },
    #                 'prime_brute': row['Prime brute'],
    #                 'reductions': values.get('RÃ©ductions', 0),
    #                 'prime_nette': values.get('Prime nette', 0),
    #                 'droit_timbre': values.get('Droit de Timbre Automobile', 0),
    #                 'accessoires': 0,
    #                 'asac': 0,
    #                 'carte_rose': 0,
    #                 'vignette': 0,
    #                 'tva': 0,
    #                 'pttc': 0,
    #                 'status': 'loaded',
    #                 'from_fleet': False,
    #                 'categorie': self._safe_str(row.get('Genre', 'VP')),
    #                 'annee': self._safe_year(pmec_val),
    #             }
                
    #             # âœ… Calcul du total des garanties
    #             garanties_keys = ['rc', 'dr', 'vol', 'vb', 'incendie', 'bris_glace', 'ar', 'dta', 'ipt']
    #             total = sum(vehicle['garanties'].get(k, 0) for k in garanties_keys)
    #             vehicle['garanties']['total'] = total
                
    #             # âœ… Journalisation pour le dÃ©bogage
    #             print(f"âœ… {immat} - Marque: {marque}, ModÃ¨le: {modele}, AR: {ar_value} FCFA")
                
    #             vehicles.append(vehicle)
                
    #         except Exception as e:
    #             print(f"Erreur extraction ligne {idx}: {e}")
    #             import traceback
    #             traceback.print_exc()
    #             continue
        
    #     return vehicles

    def _extract_vehicles_data(self, df):
        """Extrait les donnÃ©es des vÃ©hicules - Version avec traitement des textes spÃ©ciaux"""
        vehicles = []
        
        # âœ… Colonnes numÃ©riques avec les nouveaux noms
        numeric_cols = [
            'Puissance en CV', 'Nombre de Places', 'Nombre de jours assurÃ©s',
            'Valeur Neuve', 'Valeur Venale', 'Capital Assistance Ã  la rÃ©paration',
            'RC/RTI', 'DÃ©fense et Recours', 'Vol/Vol partiel - Braquage',
            'Incendie', 'Bris de Glaces', 'Assistance Ã  la rÃ©paration',
            'Dommages Tous Accidents', 'Individuelle Personnes TransportÃ©es (I.P.T) + Conducteur',
            'Prime brute', 'RÃ©ductions', 'Prime nette', 'Droit de Timbre Automobile'
        ]
        
        for idx, row in df.iterrows():
            try:
                # Extraire l'immatriculation
                immat = self._safe_str(row.get('Immatriculation', ''))
                if not immat:
                    continue
                
                # âœ… VÃ©rifier si c'est une ligne de total (exclure les lignes de rÃ©capitulatif)
                if any(keyword in immat.upper() for keyword in ['PRIME', 'REDUCTIONS', 'TOTAL', 'SOUS-TOTAL']):
                    print(f"â­ï¸ Ligne de total ignorÃ©e: {immat}")
                    continue
                
                # âœ… Extraire la marque et le modÃ¨le (sÃ©parÃ©s)
                marque = self._safe_str(row.get('Marque', ''))
                modele = self._safe_str(row.get('Modele', ''))
                
                # Si la colonne Marque est vide mais que Marque et Type existe
                if not marque:
                    marque_type = self._safe_str(row.get('Marque et Type', ''))
                    if marque_type:
                        # âœ… Ne pas traiter les lignes de rÃ©capitulatif
                        if not any(k in marque_type.upper() for k in ['PRIME', 'REDUCTIONS', 'TOTAL']):
                            parts = marque_type.split(' ', 1)
                            marque = parts[0] if parts else ''
                            modele = parts[1] if len(parts) > 1 else ''
                
                # Extraire les valeurs numÃ©riques
                values = {}
                for col in numeric_cols:
                    if col in row.index:
                        val = row[col]
                        if hasattr(val, 'iloc'):
                            val = val.iloc[0] if len(val) > 0 else 0
                        values[col] = self._safe_float(val)
                    else:
                        values[col] = 0
                
                # âœ… EXTRAIRE LA PUISSANCE ET L'Ã‰NERGIE DE LA COLONNE "Puissance en CV"
                puissance_raw = row.get('Puissance en CV', '')
                if hasattr(puissance_raw, 'iloc'):
                    puissance_raw = puissance_raw.iloc[0] if len(puissance_raw) > 0 else ''
                
                puissance, energie = self._extract_power_and_energy(puissance_raw)
                
                # âœ… RÃ‰CUPÃ‰RATION CORRECTE DE "Assistance Ã  la rÃ©paration"
                ar_value = 0
                
                # 1. Essayer de rÃ©cupÃ©rer depuis la colonne "Assistance Ã  la rÃ©paration"
                if 'Assistance Ã  la rÃ©paration' in row.index:
                    val = row['Assistance Ã  la rÃ©paration']
                    if hasattr(val, 'iloc'):
                        val = val.iloc[0] if len(val) > 0 else 0
                    ar_value = self._safe_float(val)
                    if ar_value > 0:
                        print(f"âœ… AR trouvÃ©e dans colonne 'Assistance Ã  la rÃ©paration': {ar_value}")

                # Extraire les dates
                pmec_val = row.get('PMEC')
                if hasattr(pmec_val, 'iloc'):
                    pmec_val = pmec_val.iloc[0] if len(pmec_val) > 0 else None
                pmec_date = self._extract_date_smart(pmec_val)
                
                # RÃ©cupÃ©rer le rang
                rang_val = row.get('Rang')
                if hasattr(rang_val, 'iloc'):
                    rang_val = rang_val.iloc[0] if len(rang_val) > 0 else idx + 1
                
                # âœ… RÃ©cupÃ©rer les garanties
                garanties = {
                    'rc': values.get('RC/RTI', 0),
                    'dr': values.get('DÃ©fense et Recours', 0),
                    'vol': values.get('Vol/Vol partiel - Braquage', 0),
                    'vb': values.get('Vol/Vol partiel - Braquage', 0),
                    'incendie': values.get('Incendie', 0),
                    'bris_glace': values.get('Bris de Glaces', 0),
                    'ar': ar_value,
                    'dta': values.get('Dommages Tous Accidents', 0),
                    'ipt': values.get('Individuelle Personnes TransportÃ©es (I.P.T) + Conducteur', 0),
                }
                
                # âœ… Calcul du total des garanties (brut)
                garanties_keys = ['rc', 'dr', 'vol', 'vb', 'incendie', 'bris_glace', 'ar', 'dta', 'ipt']
                total_brut = sum(garanties.get(k, 0) for k in garanties_keys)
                garanties['total'] = total_brut
                
                # âœ… RÃ©cupÃ©rer les valeurs de prime depuis le fichier
                prime_brute_from_file = values.get('Prime brute', 0)
                prime_nette_from_file = values.get('Prime nette', 0)
                reductions_from_file = values.get('RÃ©ductions', 0)
                
                # âœ… Si les valeurs du fichier sont valides, les utiliser, sinon utiliser les calculs
                # Si prime_brute_from_file > 0, c'est qu'elle vient du fichier
                if prime_brute_from_file > 0:
                    prime_brute = prime_brute_from_file
                else:
                    prime_brute = total_brut
                
                if prime_nette_from_file > 0:
                    prime_nette = prime_nette_from_file
                else:
                    prime_nette = total_brut
                
                if reductions_from_file > 0:
                    reductions = reductions_from_file
                else:
                    reductions = 0
                
                vehicle = {
                    'id': idx + 1,
                    'rang': self._safe_int(rang_val) or (idx + 1),
                    'immatriculation': immat,
                    'chassis': self._safe_str(row.get('Numero Chassi', '')),
                    'marque': marque,
                    'modele': modele,
                    'marque_type': f"{marque} {modele}".strip(),
                    'genre': self._safe_str(row.get('Genre', 'VP')),
                    'puissance': puissance,
                    'energie': energie,
                    'usage': self._safe_str(row.get('Usage', '')),
                    'places': int(values.get('Nombre de Places', 5)),
                    'nbr_jour': int(values.get('Nombre de jours assurÃ©s', 365)),
                    'pmec': pmec_date,
                    'valeur_neuf': values.get('Valeur Neuve', 0),
                    'valeur_venale': values.get('Valeur Venale', 0),
                    'capital_ar': values.get('Capital Assistance Ã  la rÃ©paration', 0),
                    'garanties': garanties,
                    'prime_brute': prime_brute,
                    'reductions': reductions,
                    'prime_nette': prime_nette,
                    'droit_timbre': values.get('Droit de Timbre Automobile', 0),
                    'accessoires': 0,
                    'asac': 0,
                    'carte_rose': 0,
                    'vignette': 0,
                    'tva': 0,
                    'pttc': 0,
                    'status': 'loaded',
                    'from_fleet': False,
                    'categorie': self._safe_str(row.get('Genre', 'VP')),
                    'annee': self._safe_year(pmec_val),
                }
                
                # âœ… Journalisation pour le dÃ©bogage
                print(f"âœ… {immat} - Marque: {marque}, ModÃ¨le: {modele}, AR: {ar_value} FCFA")
                print(f"   Prime brute: {prime_brute}, Prime nette: {prime_nette}, RÃ©ductions: {reductions}")
                
                vehicles.append(vehicle)
                
            except Exception as e:
                print(f"Erreur extraction ligne {idx}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        return vehicles

    # ============================================================================
    # 8. VALIDATION DES DONNÃ‰ES
    # ============================================================================

    def _validate_vehicles_data(self):
        """Valide les donnÃ©es des vÃ©hicules"""
        invalid_count = 0
        
        for vehicle in self.vehicles_data:
            errors = []
            
            # VÃ©rifier l'immatriculation
            immat = vehicle.get('immatriculation', '')
            if not immat or len(immat) < 3:
                errors.append("Immatriculation invalide")
            
            # VÃ©rifier les valeurs numÃ©riques
            if vehicle.get('valeur_neuf', 0) < 0:
                errors.append("Valeur Ã  neuf nÃ©gative")
            if vehicle.get('valeur_venale', 0) < 0:
                errors.append("Valeur vÃ©nale nÃ©gative")
            if vehicle.get('puissance', 0) < 0:
                errors.append("Puissance nÃ©gative")
            if vehicle.get('places', 0) <= 0:
                errors.append("Nombre de places invalide")
            
            # VÃ©rifier les garanties
            garanties = vehicle.get('garanties', {})
            for key, value in garanties.items():
                if key != 'total' and value < 0:
                    errors.append(f"Garantie {key} nÃ©gative")
            
            if errors:
                vehicle['validation_errors'] = errors
                vehicle['valid'] = False
                invalid_count += 1
            else:
                vehicle['valid'] = True
        
        # Afficher les erreurs de validation
        if invalid_count > 0:
            error_msg = f"âš ï¸ {invalid_count} vÃ©hicule(s) ont des erreurs:\n\n"
            for v in self.vehicles_data:
                if not v.get('valid', True):
                    errors = v.get('validation_errors', [])
                    error_msg += f"â€¢ {v.get('immatriculation', 'N/A')}: {', '.join(errors)}\n"
                    if len(error_msg) > 1000:
                        error_msg += "...\n"
                        break
            
            QMessageBox.warning(self, "âš ï¸ Erreurs de validation", error_msg)
        
        return invalid_count

    # ============================================================================
    # 9. FONCTIONS UTILITAIRES
    # ============================================================================

    def _safe_str(self, value):
        """Convertit en chaÃ®ne de caractÃ¨res de maniÃ¨re sÃ©curisÃ©e"""
        if pd.isna(value) or value is None:
            return ''
        return str(value).strip()

    # def _safe_float(self, value):
    #     """Convertit en float de maniÃ¨re sÃ©curisÃ©e avec gestion des sÃ©parateurs"""
    #     try:
    #         if pd.isna(value) or value is None:
    #             return 0.0
    #         if isinstance(value, (int, float)):
    #             return float(value)
            
    #         # Nettoyer la chaÃ®ne
    #         value_str = str(value).strip()
            
    #         # GÃ©rer les sÃ©parateurs
    #         # 1. Supprimer les espaces
    #         value_str = value_str.replace(' ', '')
            
    #         # 2. GÃ©rer les virgules comme sÃ©parateur dÃ©cimal
    #         if ',' in value_str and '.' not in value_str:
    #             value_str = value_str.replace(',', '.')
    #         # 3. GÃ©rer les points comme sÃ©parateur de milliers
    #         elif '.' in value_str and value_str.count('.') > 1:
    #             value_str = value_str.replace('.', '')
    #         # 4. GÃ©rer les virgules comme sÃ©parateur de milliers
    #         elif ',' in value_str and '.' in value_str:
    #             value_str = value_str.replace(',', '')
            
    #         # Supprimer les caractÃ¨res non numÃ©riques sauf le point
    #         import re
    #         value_str = re.sub(r'[^\d.]+', '', value_str)
            
    #         return float(value_str) if value_str else 0.0
    #     except Exception as e:
    #         print(f"Erreur conversion float '{value}': {e}")
    #         return 0.0

    def _safe_float(self, value):
        """Convertit en float de maniÃ¨re sÃ©curisÃ©e avec gestion des sÃ©parateurs et des textes"""
        try:
            # âœ… Si c'est une Series, prendre la premiÃ¨re valeur
            if hasattr(value, 'iloc'):
                value = value.iloc[0] if len(value) > 0 else 0
            
            if pd.isna(value) or value is None:
                return 0.0
            
            # âœ… Si c'est dÃ©jÃ  un nombre
            if isinstance(value, (int, float)):
                return float(value)
            
            # âœ… Si c'est une chaÃ®ne
            value_str = str(value).strip()
            if not value_str:
                return 0.0
            
            # âœ… VÃ©rifier si c'est un texte spÃ©cial comme "INCLUS" ou "EXCLU"
            value_upper = value_str.upper()
            if value_upper in ['INCLUS', 'EXCLU', 'N/A', 'NON', 'OUI']:
                return 0.0
            
            # âœ… Supprimer les espaces insÃ©cables et autres caractÃ¨res spÃ©ciaux
            value_str = value_str.replace('\xa0', ' ').replace('â€¯', ' ')
            
            # Nettoyer la chaÃ®ne
            # GÃ©rer les sÃ©parateurs
            # 1. Supprimer les espaces
            value_str = value_str.replace(' ', '')
            
            # 2. GÃ©rer les virgules comme sÃ©parateur dÃ©cimal
            if ',' in value_str and '.' not in value_str:
                value_str = value_str.replace(',', '.')
            # 3. GÃ©rer les points comme sÃ©parateur de milliers
            elif '.' in value_str and value_str.count('.') > 1:
                value_str = value_str.replace('.', '')
            # 4. GÃ©rer les virgules comme sÃ©parateur de milliers
            elif ',' in value_str and '.' in value_str:
                value_str = value_str.replace(',', '')
            
            # Supprimer les caractÃ¨res non numÃ©riques sauf le point
            import re
            value_str = re.sub(r'[^\d.]+', '', value_str)
            
            return float(value_str) if value_str else 0.0
        except Exception as e:
            print(f"Erreur conversion float '{value}': {e}")
            return 0.0

    def _safe_int(self, value):
        """Convertit en entier de maniÃ¨re sÃ©curisÃ©e"""
        try:
            if pd.isna(value) or value is None:
                return 0
            if isinstance(value, (int, float)):
                return int(value)
            if isinstance(value, str):
                # Nettoyer la chaÃ®ne
                value_str = value.strip()
                if not value_str:
                    return 0
                # Supprimer les espaces et les caractÃ¨res non numÃ©riques
                import re
                value_str = re.sub(r'[^\d]', '', value_str)
                return int(value_str) if value_str else 0
            return int(value)
        except Exception as e:
            print(f"Erreur conversion int '{value}': {e}")
            return 0

    def _extract_date_smart(self, value):
        """Extrait une date de maniÃ¨re intelligente"""
        if pd.isna(value) or not value:
            return None
        
        if isinstance(value, (pd.Timestamp, datetime)):
            return value
        
        value_str = str(value).strip()
        
        # Essayer diffÃ©rents formats
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%m/%d/%Y',
            '%d-%m-%Y',
            '%m-%d-%Y',
            '%d.%m.%Y',
            '%Y.%m.%d',
            '%d/%m/%y',
            '%d-%m-%y',
            '%d %b %Y',  # 01 Jan 2024
            '%d %B %Y',  # 01 Janvier 2024
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(value_str, fmt)
            except ValueError:
                continue
        
        # Essayer les dates Excel
        try:
            num_val = float(value_str)
            if num_val > 1 and num_val < 50000:
                from datetime import timedelta
                base = datetime(1899, 12, 30)
                return base + timedelta(days=num_val)
        except:
            pass
        
        return None

    def _extract_energy(self, value):
        """Extrait le type d'Ã©nergie Ã  partir de la valeur de puissance"""
        if pd.isna(value) or not value:
            return 'SEE'  # Essence par dÃ©faut
        
        value_str = str(value).strip().upper()
        
        # DÃ©tection des lettres d'Ã©nergie
        energy_map = {
            'D': 'SED',  # Diesel
            'E': 'SEE',  # Essence
            'H': 'HYBRIDE',  # Hybride
            'G': 'GAZ',  # Gaz
            'EL': 'ELECTRIQUE'  # Ã‰lectrique
        }
        
        import re
        match = re.search(r'([DEHGL]+)', value_str)
        if match:
            energy_code = match.group(1)
            return energy_map.get(energy_code, 'SEE')
        
        return 'SEE'

    def _safe_year(self, value):
        """Extrait l'annÃ©e Ã  partir d'une date ou d'une chaÃ®ne"""
        try:
            if pd.isna(value) or not value:
                return None
            
            # Si c'est une date
            if isinstance(value, (pd.Timestamp, datetime)):
                return value.year
            
            # Si c'est une chaÃ®ne
            value_str = str(value).strip()
            import re
            match = re.search(r'\b(19|20)\d{2}\b', value_str)
            if match:
                return int(match.group(0))
            
            return None
        except:
            return None

    def setup_vehicle_table(self):
        """Configure le tableau avec les nouvelles colonnes"""
        headers = [
            "âœ“",  # 0
            "Rang",  # 1
            "Immatriculation",  # 2
            "Numero Chassi",  # 3
            "Marque",  # 4
            "ModÃ¨le",  # 5
            "Genre",  # 6
            "Puissance en CV",  # 7
            "Ã‰nergie",  # 8
            "Usage",  # 9
            "Nombre de Places",  # 10
            "Nombre de jours assurÃ©s",  # 11
            "PMEC",  # 12
            "Valeur Neuve",  # 13
            "Valeur Venale",  # 14
            "Capital Assistance Ã  la rÃ©paration",  # 15
            "RC/RTI",  # 16
            "DÃ©fense et Recours",  # 17
            "Vol/Vol partiel - Braquage",  # 18
            "Incendie",  # 19
            "Bris de Glaces",  # 20
            "Assistance Ã  la rÃ©paration",  # 21
            "Dommages Tous Accidents",  # 22
            "Individuelle Personnes TransportÃ©es (I.P.T) + Conducteur",  # 23
            "Prime brute",  # 24
            "RÃ©ductions",  # 25
            "Prime nette",  # 26
            "Droit de Timbre Automobile",  # 27
            "Actions"  # 28 
        ]
        
        self.vehicles_table = QTableWidget()
        self.vehicles_table.setColumnCount(len(headers))
        self.vehicles_table.setHorizontalHeaderLabels(headers)
        
        # Configuration des colonnes
        column_widths = {
            0: 30,   # âœ“
            1: 40,   # Rang
            2: 100,  # Immatriculation
            3: 100,  # Numero Chassi
            4: 120,  # Marque
            5: 120,  #ModÃ¨le
            6: 70,   # Genre
            7: 80,   # Puissance en CV
            8: 80,   # Ã‰nergie
            9: 80,   # Usage
            10: 60,  # Nombre de Places
            11: 80,   # Nombre de jours assurÃ©s
            12: 70,  # PMEC
            13: 90,  # Valeur Neuve
            14: 90,  # Valeur Venale
            15: 90,  # Capital Assistance
            16: 80,  # RC/RTI
            17: 80,  # DR
            18: 80,  # Vol/Vol partiel - Braquage
            19: 80,  # Incendie
            20: 80,  # Bris
            21: 80,  # AR
            22: 80,  # DTA
            23: 80,  # IPT
            24: 80,  # Prime brute
            25: 70,  # RÃ©ductions
            26: 80,  # Prime nette
            27: 80,  # Droit de Timbre
            28: 120  # Actions
        }
        
        for col, width in column_widths.items():
            self.vehicles_table.setColumnWidth(col, width)
        
        self.vehicles_table.setAlternatingRowColors(True)
        self.vehicles_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.vehicles_table.setMinimumHeight(350)
        self.vehicles_table.verticalHeader().setDefaultSectionSize(40)
        self.vehicles_table.verticalHeader().setMinimumSectionSize(35)
        
    def display_vehicles(self):
        """Affiche les vÃ©hicules dans le tableau avec les nouvelles colonnes"""
        self.vehicles_table.setRowCount(len(self.vehicles_data))
        
        for row, vehicle in enumerate(self.vehicles_data):
            # Case Ã  cocher
            check_item = QTableWidgetItem()
            check_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            check_item.setCheckState(Qt.Checked)
            self.vehicles_table.setItem(row, 0, check_item)
            
            # Mettre Ã  jour la ligne (utilise update_vehicle_row qui gÃ¨re toutes les colonnes)
            self.update_vehicle_row(row, vehicle)
            
            # Actions (colonne 27)
            vehicle_id = vehicle.get('id', row)
            actions_widget = VehicleActionsWidget(
                vehicle_id,
                self.edit_vehicle_garanties,
                self.edit_vehicle_dates,
                self.on_modify_vehicle
            )
            self.vehicles_table.setCellWidget(row, 28, actions_widget)
        
        self.adjust_row_heights()
        self.update_selection_count()
        self.update_summary()
    
    def apply_mass_reduction(self):
        """Applique une rÃ©duction en masse sur les garanties sÃ©lectionnÃ©es"""
        if not self.vehicles_data:
            QMessageBox.warning(self, "Erreur", "Aucun vÃ©hicule chargÃ©")
            return
        
        # VÃ©rifier qu'il y a des vÃ©hicules sÃ©lectionnÃ©s
        selected_rows = []
        for row in range(self.vehicles_table.rowCount()):
            item = self.vehicles_table.item(row, 0)
            if item and item.checkState() == Qt.Checked:
                selected_rows.append(row)
        
        if not selected_rows:
            QMessageBox.warning(self, "Erreur", "Veuillez sÃ©lectionner au moins un vÃ©hicule")
            return
        
        # Ouvrir le dialogue de rÃ©duction
        dialog = MassReductionDialog(self)
        if dialog.exec() != QDialog.Accepted:
            return
        
        selected_garanties = dialog.get_selected_garanties()
        if not selected_garanties:
            QMessageBox.warning(self, "Erreur", "Veuillez sÃ©lectionner au moins une garantie")
            return
        
        reduction_percent = dialog.get_reduction_percentage()
        if reduction_percent <= 0:
            QMessageBox.warning(self, "Erreur", "Le pourcentage de rÃ©duction doit Ãªtre supÃ©rieur Ã  0")
            return
        
        # âœ… Appliquer la rÃ©duction (UNIQUEMENT en pourcentage)
        applied_count = 0
        modifications = []
        
        for row in selected_rows:
            vehicle = self.vehicles_data[row]
            if 'garanties' not in vehicle:
                continue
            
            garanties = vehicle['garanties']
            vehicle_modifications = {}
            vehicle_modified = False
            
            for key in selected_garanties:
                if key not in garanties:
                    continue
                
                current_value = garanties.get(key, 0)
                
                # âœ… Appliquer le pourcentage de rÃ©duction
                reduction_factor = 1 - (reduction_percent / 100)
                new_value = round(current_value * reduction_factor, 2)
                
                # Arrondir Ã  l'entier le plus proche pour les montants en FCFA
                new_value = round(new_value)
                
                if new_value != current_value:
                    vehicle_modifications[key] = new_value
                    vehicle_modified = True
            
            if vehicle_modified:
                # Appliquer les modifications
                for key, new_value in vehicle_modifications.items():
                    garanties[key] = new_value
                
                # Recalculer le total
                total_keys = ['rc', 'dr', 'vol', 'vb', 'incendie', 'bris_glace', 'ar', 'dta', 'ipt']
                garanties['total'] = sum(garanties.get(k, 0) for k in total_keys)
                
                # Mettre Ã  jour les colonnes du tableau
                garanties_cols = {
                    'rc': 16, 'dr': 17, 'vol': 18, 'vb': 18,
                    'incendie': 19, 'bris_glace': 20, 'ar': 21,
                    'dta': 22, 'ipt': 23
                }
                
                for key, col in garanties_cols.items():
                    if key in vehicle_modifications:
                        amount = garanties.get(key, 0)
                        item = self.vehicles_table.item(row, col)
                        if item:
                            item.setText(f"{amount:,.0f}".replace(",", " "))
                
                # Mettre Ã  jour la prime brute
                prime_brute = garanties.get('total', 0)
                pb_item = self.vehicles_table.item(row, 24)
                if pb_item:
                    pb_item.setText(f"{prime_brute:,.0f}".replace(",", " "))
                
                # Mettre Ã  jour la prime nette
                pn_item = self.vehicles_table.item(row, 26)
                if pn_item:
                    pn_item.setText(f"{prime_brute:,.0f}".replace(",", " "))
                
                # Mettre Ã  jour les rÃ©ductions
                reductions_item = self.vehicles_table.item(row, 25)
                if reductions_item:
                    reductions_item.setText(f"{reduction_percent:.0f}%")
                
                applied_count += 1
                modifications.append(vehicle['immatriculation'])
        
        # Mettre Ã  jour les rÃ©capitulatifs
        self.update_summary()
        
        if applied_count > 0:
            mod_list = "\n".join(modifications[:5]) + ("\n..." if len(modifications) > 5 else "")
            QMessageBox.information(
                self, 
                "âœ… SuccÃ¨s", 
                f"RÃ©duction appliquÃ©e Ã  {applied_count} vÃ©hicule(s)\n\n"
                f"Garanties concernÃ©es: {len(selected_garanties)}\n"
                f"RÃ©duction: {reduction_percent:.1f}%\n\n"
                f"VÃ©hicules modifiÃ©s:\n{mod_list}"
            )
        else:
            QMessageBox.information(
                self, 
                "â„¹ï¸ Information", 
                "Aucune modification n'a Ã©tÃ© appliquÃ©e.\n"
                "Les garanties sÃ©lectionnÃ©es Ã©taient peut-Ãªtre dÃ©jÃ  Ã  0."
            )

    def start_calculation(self):
        """DÃ©marre le calcul des garanties"""
        if not self.vehicles_data:
            return
        
        # Calculer la durÃ©e
        debut = self.date_debut.date().toPython()
        fin = self.date_fin.date().toPython()
        jours = max(1, (fin - debut).days)
        
        params = {
            'compagny_id': 1,  # ID par dÃ©faut
            'zone': 'A',
            'code_tarif': '',
            'avec_remorque': False,
            'duree_jours': jours
        }
        
        # Mettre Ã  jour l'UI
        self.calc_btn.setEnabled(False)
        self.calc_btn.setText("Calcul en cours...")
        self.progress_widget.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_icon.setText("ðŸ”„")
        self.status_text.setText("Calcul des garanties en cours...")
        
        # DÃ©marrer le thread
        self.calculation_thread = CalculationThread(self.controller, self.vehicles_data, params)
        self.calculation_thread.progress.connect(self.on_calculation_progress)
        self.calculation_thread.finished_signal.connect(self.on_calculation_finished)
        self.calculation_thread.start()
      
    def get_dataon_calculation_finished(self, results):
        """Termine le calcul"""
        self.vehicles_data = results
        self.progress_widget.setVisible(False)
        self.calc_btn.setEnabled(False)
        self.calc_btn.setText("Calcul terminÃ©")
        self.import_btn.setEnabled(True)
        self.mass_reduction_btn.setEnabled(True)
        self.refresh_btn.setEnabled(True)
        
        self.status_icon.setText("âœ…")
        self.status_text.setText("Calcul terminÃ©, vous pouvez importer les vÃ©hicules sÃ©lectionnÃ©s")
        
        self.update_summary()

    def on_calculation_finished(self, results):
        """Termine le calcul"""
        self.vehicles_data = results
        self.progress_widget.setVisible(False)
        self.calc_btn.setEnabled(False)
        self.calc_btn.setText("Calcul terminÃ©")
        self.import_btn.setEnabled(True)
        self.mass_reduction_btn.setEnabled(True)
        self.refresh_btn.setEnabled(True)
        
        self.status_icon.setText("âœ…")
        self.status_text.setText("Calcul terminÃ©, vous pouvez importer les vÃ©hicules sÃ©lectionnÃ©s")
        
        self.update_summary()

    def on_calculation_progress(self, current, total, immat, garanties):
        """Met Ã  jour la progression"""
        progress = int(current / total * 100)
        self.progress_bar.setValue(progress)
        self.progress_label.setText(f"Traitement de {immat}... ({current}/{total})")
        
        row = current - 1
        if row >= len(self.vehicles_data):
            return
        
        vehicle = self.vehicles_data[row]
        vehicle['garanties'] = garanties
        
        # Mettre Ã  jour la ligne
        self.update_vehicle_row(row, vehicle)
        
        # Actions
        if not self.vehicles_table.cellWidget(row, 27):
            vehicle_id = vehicle.get('id', row)
            actions_widget = VehicleActionsWidget(
                vehicle_id,
                self.edit_vehicle_garanties,
                self.edit_vehicle_dates,
                self.on_modify_vehicle
            )
            self.vehicles_table.setCellWidget(row, 27, actions_widget)
        
        if current == total:
            self.adjust_row_heights()
            self.update_summary()
            self.mass_reduction_btn.setEnabled(True)
    
    def download_template(self):
        """TÃ©lÃ©charge le modÃ¨le Excel avec les nouvelles colonnes"""
        template_path, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer le modÃ¨le", "modele_flotte_import.xlsx", "Excel (*.xlsx)"
        )
        
        if template_path:
            # CrÃ©er un DataFrame avec les nouvelles colonnes
            data = {
                'Rang': [1, 2, 3, 4, 5],
                'Immatriculation': ['LT-001-AB', 'LT-002-BC', 'LT-003-CD', 'LT-004-DE', 'LT-005-EF'],
                'Marque et Type': ['Toyota Hilux', 'Renault Kangoo', 'Mitsubishi Outlander', 'Peugeot Partner', 'Mercedes Sprinter'],
                'Genre': ['VP', 'VU', 'VP', 'VU', 'PL'],
                'Puissance en CV': [7, 5, 6, 4, 8],
                'Usage': ['Transport de personnes', 'Transport de marchandises', 'Transport de personnes', 'Transport de marchandises', 'Transport de personnes'],
                'Nombre de Places': [5, 3, 7, 5, 9],
                'Nombre de jours assurÃ©s': [365, 365, 365, 365, 365],
                'PMEC': [2000, 1200, 1800, 1500, 3500],
                'Valeur Neuve': [35000000, 18000000, 28000000, 22000000, 45000000],
                'Valeur Venale': [32000000, 15000000, 26000000, 20000000, 42000000],
                'Capital Assistance Ã  la rÃ©paration': [5000000, 3000000, 4000000, 3500000, 6000000],
                'RC/RTI': [0, 0, 0, 0, 0],
                'DÃ©fense et Recours': [0, 0, 0, 0, 0],
                'Vol/Vol partie': [0, 0, 0, 0, 0],
                'Vol Braquage': [0, 0, 0, 0, 0],
                'Incendie': [0, 0, 0, 0, 0],
                'Bris de Glaces': [0, 0, 0, 0, 0],
                'Assistance Ã  la rÃ©paration': [0, 0, 0, 0, 0],
                'Dommages Tous Accidents': [0, 0, 0, 0, 0],
                'IPT + Conducteur': [0, 0, 0, 0, 0],
                'Prime brute': [0, 0, 0, 0, 0],
                'RÃ©ductions': [0, 0, 0, 0, 0],
                'Prime nette': [0, 0, 0, 0, 0],
                'Droit de Timbre Automobile': [5000, 5000, 5000, 5000, 5000]
            }
            
            df = pd.DataFrame(data)
            
            with pd.ExcelWriter(template_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='VÃ©hicules', index=False)
                
                # Feuille d'instructions
                instructions_data = {
                    'Colonne': list(data.keys()),
                    'Description': [
                        'NumÃ©ro de rang (optionnel)',
                        'ðŸ”´ OBLIGATOIRE - Plaque d\'immatriculation',
                        'ðŸ”´ OBLIGATOIRE - Marque et modÃ¨le du vÃ©hicule',
                        'Genre du vÃ©hicule (VP, VU, PL, etc.)',
                        'Puissance fiscale en CV',
                        'Usage du vÃ©hicule',
                        'Nombre de places assises',
                        'Nombre de jours d\'assurance (365 par dÃ©faut)',
                        'PMEC (Poids Maximum AutorisÃ©)',
                        'Valeur Ã  neuf du vÃ©hicule (FCFA)',
                        'Valeur vÃ©nale actuelle (FCFA)',
                        'Capital pour l\'assistance Ã  la rÃ©paration',
                        'RC/RTI - Laisser 0 pour calcul automatique',
                        'DÃ©fense et Recours - Laisser 0 pour calcul automatique',
                        'Vol/Vol partie - Laisser 0 pour calcul automatique',
                        'Vol Braquage - Laisser 0 pour calcul automatique',
                        'Incendie - Laisser 0 pour calcul automatique',
                        'Bris de Glaces - Laisser 0 pour calcul automatique',
                        'Assistance Ã  la rÃ©paration - Laisser 0 pour calcul automatique',
                        'Dommages Tous Accidents - Laisser 0 pour calcul automatique',
                        'IPT + Conducteur - Laisser 0 pour calcul automatique',
                        'Prime brute - Laisser 0 pour calcul automatique',
                        'RÃ©ductions - Appliquer manuellement si besoin',
                        'Prime nette - Laisser 0 pour calcul automatique',
                        'Droit de Timbre Automobile (5000 FCFA par dÃ©faut)'
                    ],
                    'Statut': [
                        'Optionnel', 'ðŸ”´ Requis', 'ðŸ”´ Requis',
                        'RecommandÃ©', 'RecommandÃ©', 'Optionnel',
                        'RecommandÃ©', 'Optionnel', 'Optionnel',
                        'RecommandÃ©', 'RecommandÃ©', 'Optionnel',
                        'Optionnel', 'Optionnel', 'Optionnel',
                        'Optionnel', 'Optionnel', 'Optionnel',
                        'Optionnel', 'Optionnel', 'Optionnel',
                        'Optionnel', 'Optionnel', 'Optionnel',
                        'Optionnel'
                    ]
                }
                
                instructions_df = pd.DataFrame(instructions_data)
                instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
                
                # Ajuster la largeur des colonnes
                from openpyxl.styles import PatternFill, Font, Alignment
                
                workbook = writer.book
                
                # Colorer les en-tÃªtes
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for sheet_name in ['VÃ©hicules', 'Instructions']:
                    ws = workbook[sheet_name]
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Ajuster la largeur des colonnes
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 40)
                        ws.column_dimensions[column_letter].width = adjusted_width
            
            QMessageBox.information(
                self, 
                "SuccÃ¨s", 
                f"âœ… ModÃ¨le crÃ©Ã© avec succÃ¨s !\n\n"
                f"ðŸ“ Emplacement: {template_path}\n\n"
                f"ðŸ“‹ Le fichier contient:\n"
                f"  â€¢ Une feuille 'VÃ©hicules' avec les donnÃ©es Ã  remplir\n"
                f"  â€¢ Une feuille 'Instructions' avec la description des colonnes\n\n"
                f"ðŸ”´ Les colonnes obligatoires sont:\n"
                f"  â€¢ Immatriculation\n"
                f"  â€¢ Marque et Type"
            )
    
    def adjust_row_heights(self):
        """Ajuste la hauteur des lignes"""
        for row in range(self.vehicles_table.rowCount()):
            self.vehicles_table.setRowHeight(row, 40)
    
    def _extract_power_and_energy(self, value):
        """
        Extrait la puissance et le type d'Ã©nergie d'une chaÃ®ne comme '9D', '6E', '4H'
        Retourne (puissance, energie)
        D = Diesel (SED)
        E = Essence (SEE)
        H = Hybride
        """
        try:
            # âœ… Si c'est une Series, prendre la premiÃ¨re valeur
            if hasattr(value, 'iloc'):
                value = value.iloc[0] if len(value) > 0 else ''
            
            if pd.isna(value) or not value:
                return (0, 'SEE')  # Essence par dÃ©faut
            
            val_str = str(value).strip().upper()
            
            # Si c'est dÃ©jÃ  un nombre pur
            if val_str.isdigit():
                return (int(val_str), 'SEE')
            
            # Extraire le nombre et la lettre
            import re
            match = re.match(r'^(\d+)\s*([DEH]?)$', val_str)
            if match:
                puissance = int(match.group(1))
                energie_letter = match.group(2)
                
                # Convertir la lettre en nom complet
                if energie_letter == 'D':
                    energie = 'SED'  # Diesel
                elif energie_letter == 'E':
                    energie = 'SEE'  # Essence
                elif energie_letter == 'H':
                    energie = 'HYBRIDE'  # Hybride
                else:
                    energie = 'SEE'  # Essence par dÃ©faut
                
                return (puissance, energie)
            
            # Fallback: essayer d'extraire juste le nombre
            numbers = re.findall(r'\d+', val_str)
            if numbers:
                return (int(numbers[0]), 'SEE')
            
            return (0, 'SEE')
            
        except Exception as e:
            print(f"Erreur extraction puissance/Ã©nergie: {e}")
            return (0, 'SEE')

    # MÃ©thodes de gestion des flottes existantes (simplifiÃ©es)
    def on_mode_changed(self):
        """Change le mode d'importation"""
        is_new = self.mode_new.isChecked()
        self.new_fleet_widget.setVisible(is_new)
        self.existing_fleet_widget.setVisible(not is_new)
        
        if not is_new:
            self.load_existing_fleets()
            # Connecter le signal de changement de sÃ©lection
            try:
                self.existing_fleet_combo.currentIndexChanged.disconnect()
            except:
                pass
            self.existing_fleet_combo.currentIndexChanged.connect(self.on_existing_fleet_selected)
        else:
            # DÃ©connecter pour Ã©viter les appels inutiles
            try:
                self.existing_fleet_combo.currentIndexChanged.disconnect()
            except:
                pass
      
    def load_existing_fleets(self):
        """Charge les flottes existantes depuis la base de donnÃ©es"""
        try:
            self.existing_fleet_combo.clear()
            self.existing_fleet_combo.addItem("SÃ©lectionner une flotte", None)
            
            # RÃ©cupÃ©rer l'ID du contact propriÃ©taire
            contact_id = None
            if hasattr(self.parent(), 'contact'):
                contact_id = self.parent().contact.id
            elif hasattr(self.controller, 'current_user_id'):
                # RÃ©cupÃ©rer le contact associÃ© Ã  l'utilisateur courant
                try:
                    current_user = self.controller.users.get_by_id(self.controller.current_user_id)
                    if current_user and hasattr(current_user, 'contact_id'):
                        contact_id = current_user.contact_id
                except Exception as e:
                    print(f"Erreur rÃ©cupÃ©ration utilisateur: {e}")
            
            # Charger les flottes
            fleets = []
            if contact_id:
                try:
                    fleets = self.controller.fleets.get_fleets_by_owner(contact_id)
                except AttributeError:
                    # Si la mÃ©thode n'existe pas, essayer d'autres mÃ©thodes
                    try:
                        fleets = self.controller.fleets.get_all_fleets()
                    except:
                        pass
            
            # Si pas de flottes trouvÃ©es, essayer de charger toutes les flottes
            if not fleets:
                try:
                    fleets = self.controller.fleets.get_all_fleets()
                except:
                    pass
            
            if fleets:
                for fleet in fleets:
                    # RÃ©cupÃ©rer le nom de la flotte
                    if hasattr(fleet, 'nom_flotte'):
                        name = fleet.nom_flotte
                    elif hasattr(fleet, 'nom'):
                        name = fleet.nom
                    elif hasattr(fleet, 'name'):
                        name = fleet.name
                    else:
                        name = f"Flotte {fleet.id if hasattr(fleet, 'id') else ''}"
                    
                    # Compter les vÃ©hicules
                    vehicle_count = 0
                    if hasattr(fleet, 'vehicles'):
                        vehicle_count = len(fleet.vehicles)
                    elif hasattr(fleet, 'vehicle_count'):
                        vehicle_count = fleet.vehicle_count
                    
                    # RÃ©cupÃ©rer l'ID
                    fleet_id = fleet.id if hasattr(fleet, 'id') else None
                    
                    if fleet_id:
                        display_name = f"{name} ({vehicle_count} vÃ©hicules)" if vehicle_count > 0 else name
                        self.existing_fleet_combo.addItem(display_name, fleet_id)
                
                # SÃ©lectionner le premier Ã©lÃ©ment si disponible
                if self.existing_fleet_combo.count() > 1:
                    self.existing_fleet_combo.setCurrentIndex(1)
                    self.on_existing_fleet_selected(1)
            else:
                self.existing_fleet_combo.addItem("Aucune flotte disponible", None)
                
        except Exception as e:
            print(f"Erreur chargement flottes: {e}")
            traceback.print_exc()
            self.existing_fleet_combo.clear()
            self.existing_fleet_combo.addItem("Erreur de chargement", None)

    def on_existing_fleet_selected(self, index):
        """Charge les vÃ©hicules de la flotte sÃ©lectionnÃ©e"""
        fleet_id = self.existing_fleet_combo.currentData()
        if not fleet_id:
            return
        
        try:
            self.status_icon.setText("ðŸ”„")
            self.status_text.setText("Chargement des vÃ©hicules de la flotte...")
            
            # RÃ©cupÃ©rer les vÃ©hicules de la flotte
            vehicles = []
            try:
                vehicles = self.controller.vehicles.get_vehicles_by_fleet(fleet_id)
            except AttributeError:
                # Essayer d'autres mÃ©thodes
                try:
                    vehicles = self.controller.fleets.get_fleet_vehicles(fleet_id)
                except:
                    pass
            
            if not vehicles:
                self.status_text.setText("Aucun vÃ©hicule dans cette flotte")
                self.status_icon.setText("ðŸ“­")
                self.vehicles_data = []
                self.vehicles_table.setRowCount(0)
                self.calc_btn.setEnabled(False)
                self.import_btn.setEnabled(False)
                return
            
            # Transformer les donnÃ©es des vÃ©hicules
            self.vehicles_data = []
            for idx, vehicle in enumerate(vehicles):
                # Si c'est un objet SQLAlchemy, le convertir en dict
                if hasattr(vehicle, '__dict__') and not isinstance(vehicle, dict):
                    vehicle_dict = {
                        'id': getattr(vehicle, 'id', idx + 1),
                        'immatriculation': getattr(vehicle, 'immatriculation', ''),
                        'chassis': getattr(vehicle, 'chassis', ''),
                        'marque': getattr(vehicle, 'marque', ''),
                        'modele': getattr(vehicle, 'modele', ''),
                        'categorie': getattr(vehicle, 'categorie', 'VP'),
                        'genre': getattr(vehicle, 'genre', 'VP'),
                        'annee': getattr(vehicle, 'annee', None),
                        'energie': getattr(vehicle, 'energie', 'SEE'),
                        'puissance': getattr(vehicle, 'usage', 0) or getattr(vehicle, 'puissance', 0),
                        'places': getattr(vehicle, 'places', 5),
                        'valeur_neuf': getattr(vehicle, 'valeur_neuf', 0),
                        'valeur_venale': getattr(vehicle, 'valeur_venale', 0),
                        'zone': getattr(vehicle, 'zone', 'A'),
                        'date_debut': getattr(vehicle, 'date_debut', None),
                        'date_fin': getattr(vehicle, 'date_fin', None),
                        'nbr_jour': getattr(vehicle, 'nbr_jour', 365),
                        'statut': getattr(vehicle, 'statut', 'En Circulation'),
                        'prime_brute': getattr(vehicle, 'prime_brute', 0),
                        'prime_nette': getattr(vehicle, 'prime_nette', 0),
                        'reductions': getattr(vehicle, 'reductions', 0),
                        'droit_timbre': getattr(vehicle, 'droit_timbre', 0),
                        'accessoires': getattr(vehicle, 'accessoires', 0),
                        'asac': getattr(vehicle, 'asac', 0),
                        'carte_rose': getattr(vehicle, 'carte_rose', 0),
                        'vignette': getattr(vehicle, 'vignette', 0),
                        'tva': getattr(vehicle, 'tva', 0),
                        'pttc': getattr(vehicle, 'pttc', 0),
                        'marque_type': f"{getattr(vehicle, 'marque', '')} {getattr(vehicle, 'modele', '')}".strip(),
                        'usage': getattr(vehicle, 'usage', ''),
                        'pmec': getattr(vehicle, 'pmec', 0),
                        'capital_ar': getattr(vehicle, 'capital_ar', 0),
                        'status': 'loaded',
                        'from_fleet': True
                    }
                    
                    # Charger les garanties
                    garanties = {}
                    guarantee_keys = ['rc', 'dr', 'vol', 'vb', 'incendie', 'bris_glace', 'ar', 'dta', 'ipt']
                    for key in guarantee_keys:
                        # Essayer de rÃ©cupÃ©rer depuis l'objet
                        if hasattr(vehicle, key):
                            garanties[key] = getattr(vehicle, key, 0)
                        elif hasattr(vehicle, 'guarantees') and vehicle.guarantees:
                            garanties[key] = getattr(vehicle.guarantees, key, 0)
                        else:
                            garanties[key] = 0
                    
                    garanties['total'] = sum(garanties.get(k, 0) for k in guarantee_keys)
                    vehicle_dict['garanties'] = garanties
                    
                else:
                    # Si c'est dÃ©jÃ  un dict
                    vehicle_dict = vehicle.copy() if isinstance(vehicle, dict) else {'immatriculation': str(vehicle)}
                    if 'garanties' not in vehicle_dict:
                        vehicle_dict['garanties'] = {}
                    vehicle_dict['from_fleet'] = True
                    if 'id' not in vehicle_dict:
                        vehicle_dict['id'] = idx + 1
                
                self.vehicles_data.append(vehicle_dict)
            
            # Afficher les vÃ©hicules
            self.display_vehicles()
            
            self.file_info.setText(f"âœ… {len(self.vehicles_data)} vÃ©hicules chargÃ©s depuis la flotte")
            self.status_text.setText(f"{len(self.vehicles_data)} vÃ©hicules chargÃ©s")
            self.status_icon.setText("âœ…")
            
            self.calc_btn.setEnabled(False)
            self.calc_btn.setText("Calcul dÃ©jÃ  effectuÃ©")
            self.import_btn.setEnabled(True)
            self.refresh_btn.setEnabled(True)
            self.mass_reduction_btn.setEnabled(True)
            
        except Exception as e:
            self.status_text.setText(f"Erreur: {str(e)}")
            self.status_icon.setText("âŒ")
            QMessageBox.critical(self, "Erreur", f"Impossible de charger les vÃ©hicules: {str(e)}")
            traceback.print_exc()

    def refresh_fleet_guarantees(self):
        """RafraÃ®chit les garanties de la flotte"""
        if not self.vehicles_data:
            QMessageBox.warning(self, "Erreur", "Aucun vÃ©hicule Ã  rafraÃ®chir")
            return
        
        # VÃ©rifier que les donnÃ©es sont valides
        for vehicle in self.vehicles_data:
            if 'immatriculation' not in vehicle:
                QMessageBox.warning(self, "Erreur", "DonnÃ©es de vÃ©hicule invalides")
                return
        
        self.start_calculation()
       
    def apply_global_frais_to_selected(self):
        """Applique les frais aux vÃ©hicules sÃ©lectionnÃ©s avec gestion du mode global/unitaire"""
        selected_rows = []
        for row in range(self.vehicles_table.rowCount()):
            item = self.vehicles_table.item(row, 0)
            if item and item.checkState() == Qt.Checked:
                selected_rows.append(row)
        
        if not selected_rows:
            QMessageBox.warning(self, "Aucune sÃ©lection", "Veuillez sÃ©lectionner au moins un vÃ©hicule")
            return
        
        try:
            accessoires = float(self.global_accessoires.text().replace(" ", "").replace(",", "") or 0)
            asac = float(self.global_asac.text().replace(" ", "").replace(",", "") or 0)
            carte_rose = float(self.global_carte_rose.text().replace(" ", "").replace(",", "") or 0)
            vignette = float(self.global_vignette.text().replace(" ", "").replace(",", "") or 0)
            
            is_global = self.frais_mode_combo.currentIndex() == 1
            count = len(selected_rows)
            
            if is_global and count > 0:
                # Mode global: rÃ©partir les frais entre tous les vÃ©hicules
                accessoires_par_vehicule = accessoires / count
                asac_par_vehicule = asac / count
                carte_rose_par_vehicule = carte_rose / count
                vignette_par_vehicule = vignette / count
            else:
                # Mode unitaire: chaque vÃ©hicule paie le montant entier
                accessoires_par_vehicule = accessoires
                asac_par_vehicule = asac
                carte_rose_par_vehicule = carte_rose
                vignette_par_vehicule = vignette
            
            for row in selected_rows:
                vehicle = self.vehicles_data[row]
                vehicle['accessoires'] = accessoires_par_vehicule
                vehicle['asac'] = asac_par_vehicule
                vehicle['carte_rose'] = carte_rose_par_vehicule
                vehicle['vignette'] = vignette_par_vehicule
                
                # Recalculer la TVA et le PTTC pour chaque vÃ©hicule
                garanties = vehicle.get('garanties', {})
                total_garanties = garanties.get('total', 0)
                print(total_garanties)
                tva_rate = 0.1925
                tva = (total_garanties + accessoires_par_vehicule + asac_par_vehicule) * tva_rate
                vehicle['tva'] = tva
                vehicle['pttc'] = total_garanties + accessoires_par_vehicule + asac_par_vehicule + tva + vignette_par_vehicule + carte_rose_par_vehicule
            
            self.update_summary()
            self.update_vehicles_table()
            
            mode_text = "rÃ©partis" if is_global else "appliquÃ©s individuellement"
            QMessageBox.information(
                self, 
                "SuccÃ¨s", 
                f"Frais {mode_text} Ã  {len(selected_rows)} vÃ©hicule(s)\n\n"
                f"Accessoires: {accessoires_par_vehicule:,.0f}".replace(",", " ") + " FCFA/vÃ©hicule\n"
                f"ASAC: {asac_par_vehicule:,.0f}".replace(",", " ") + " FCFA/vÃ©hicule\n"
                f"Carte Rose: {carte_rose_par_vehicule:,.0f}".replace(",", " ") + " FCFA/vÃ©hicule\n"
                f"Vignette: {vignette_par_vehicule:,.0f}".replace(",", " ") + " FCFA/vÃ©hicule"
            )
            
        except ValueError as e:
            QMessageBox.warning(self, "Erreur", f"Valeur invalide: {str(e)}")

    def update_vehicles_table(self):
        """Met Ã  jour l'affichage du tableau des vÃ©hicules"""
        for row, vehicle in enumerate(self.vehicles_data):
            self.update_vehicle_row(row, vehicle)
        self.update_summary()

    def edit_vehicle_garanties(self, vehicle_id):
        """Modifie les garanties d'un vÃ©hicule"""
        row = self.find_row_by_vehicle_id(vehicle_id)
        if row == -1:
            QMessageBox.warning(self, "Erreur", "VÃ©hicule non trouvÃ©")
            return
        
        vehicle = self.vehicles_data[row]
        garanties = vehicle.get('garanties', {})
        
        dialog = VehicleGarantieDialog(vehicle, garanties, self)
        if dialog.exec() == QDialog.Accepted:
            new_data = dialog.get_data()
            
            # Mettre Ã  jour les garanties
            for key in ['rc', 'dr', 'vol', 'vb', 'incendie', 'bris_glace', 'ar', 'dta', 'ipt']:
                if key in new_data:
                    vehicle['garanties'][key] = new_data[key]
                    reduction_key = f'reduction_{key}'
                    if reduction_key in new_data:
                        vehicle['garanties'][reduction_key] = new_data[reduction_key]
            
            # Recalculer le total
            total_keys = ['rc', 'dr', 'vol', 'vb', 'incendie', 'bris_glace', 'ar', 'dta', 'ipt']
            vehicle['garanties']['total'] = sum(vehicle['garanties'].get(k, 0) for k in total_keys)
            
            # Mettre Ã  jour les informations du vÃ©hicule
            for key in ['immatriculation', 'Numero Chassis', 'marque', 'modele', 'categorie', 'annee', 'energie', 'puissance', 'places', 'valeur_neuf', 'valeur_venale']:
                if key in new_data:
                    vehicle[key] = new_data[key]
            
            # RafraÃ®chir l'affichage
            self.update_vehicle_row(row, vehicle)
            self.update_summary()
            
            QMessageBox.information(self, "SuccÃ¨s", f"VÃ©hicule {vehicle['immatriculation']} mis Ã  jour")
    
    def edit_vehicle_dates(self, vehicle_id):
        """Modifie les dates d'un vÃ©hicule"""
        row = self.find_row_by_vehicle_id(vehicle_id)
        if row == -1:
            if isinstance(vehicle_id, int) and vehicle_id < len(self.vehicles_data):
                row = vehicle_id
            else:
                QMessageBox.warning(self, "Erreur", "VÃ©hicule non trouvÃ©")
                return
        
        vehicle = self.vehicles_data[row]
        
        dialog = VehicleDatesDialog(vehicle, self)
        if dialog.exec() == QDialog.Accepted:
            new_dates = dialog.get_dates()
            
            # Sauvegarder les montants annuels
            if 'garanties_annuelles' not in vehicle:
                vehicle['garanties_annuelles'] = vehicle.get('garanties', {}).copy()
            
            # Mettre Ã  jour les dates
            vehicle['date_debut'] = new_dates['date_debut']
            vehicle['date_fin'] = new_dates['date_fin']
            vehicle['nbr_jour'] = new_dates['nbr_jour']
            vehicle['statut'] = new_dates['statut']
            
            # Recalculer les garanties au prorata
            prorata = new_dates['prorata']
            garanties_annuelles = vehicle['garanties_annuelles']
            new_garanties = {}
            for key, amount in garanties_annuelles.items():
                if key != 'total' and not key.startswith('reduction_'):
                    new_garanties[key] = amount * prorata
            
            new_garanties['total'] = sum(new_garanties.values())
            vehicle['garanties'] = new_garanties
            
            # Mettre Ã  jour l'affichage
            self.update_vehicle_row(row, vehicle)
            self.update_summary()
            
            QMessageBox.information(self, "SuccÃ¨s", f"Dates mises Ã  jour pour {vehicle['immatriculation']}")
    
    def on_modify_vehicle(self, vehicle_id):
        """Modifie complÃ¨tement un vÃ©hicule"""
        row = self.find_row_by_vehicle_id(vehicle_id)
        if row == -1:
            if isinstance(vehicle_id, int) and vehicle_id < len(self.vehicles_data):
                row = vehicle_id
            else:
                QMessageBox.warning(self, "Erreur", "VÃ©hicule non trouvÃ©")
                return
        
        vehicle = self.vehicles_data[row]
        garanties = vehicle.get('garanties', {})
        
        dialog = VehicleGarantieDialog(vehicle, garanties, self)
        if dialog.exec() == QDialog.Accepted:
            new_data = dialog.get_data()
            
            # Mettre Ã  jour toutes les donnÃ©es
            for key in ['rc', 'dr', 'vol', 'vb', 'incendie', 'bris_glace', 'ar', 'dta', 'ipt']:
                if key in new_data:
                    vehicle['garanties'][key] = new_data[key]
                    reduction_key = f'reduction_{key}'
                    if reduction_key in new_data:
                        vehicle['garanties'][reduction_key] = new_data[reduction_key]
            
            # Recalculer le total
            total_keys = ['rc', 'dr', 'vol', 'vb', 'incendie', 'bris_glace', 'ar', 'dta', 'ipt']
            vehicle['garanties']['total'] = sum(vehicle['garanties'].get(k, 0) for k in total_keys)
            
            # Mettre Ã  jour les informations
            for key in ['immatriculation', 'Numero Chassis', 'marque', 'modele', 'categorie', 'annee', 'energie', 'puissance', 'places', 'valeur_neuf', 'valeur_venale']:
                if key in new_data:
                    vehicle[key] = new_data[key]
            
            # RafraÃ®chir l'affichage
            self.update_vehicle_row(row, vehicle)
            self.update_summary()
            
            QMessageBox.information(self, "SuccÃ¨s", f"VÃ©hicule {vehicle['immatriculation']} mis Ã  jour")
    
    def find_row_by_vehicle_id(self, vehicle_id):
        """Trouve l'index de la ligne Ã  partir de l'ID du vÃ©hicule"""
        for i, vehicle in enumerate(self.vehicles_data):
            if vehicle.get('id') == vehicle_id:
                return i
            # Fallback par immatriculation
            if vehicle.get('immatriculation') == str(vehicle_id):
                return i
        return -1
    
    def update_vehicle_row(self, row, vehicle):
        """Met Ã  jour l'affichage d'une ligne du tableau - Version avec Marque/ModÃ¨le sÃ©parÃ©s"""
        if row >= self.vehicles_table.rowCount():
            return
        
        # Mettre Ã  jour les colonnes
        self.vehicles_table.setItem(row, 1, QTableWidgetItem(str(vehicle.get('rang', row + 1))))
        self.vehicles_table.setItem(row, 2, QTableWidgetItem(vehicle.get('immatriculation', '')))
        self.vehicles_table.setItem(row, 3, QTableWidgetItem(vehicle.get('chassis', '')))
        
        # Marque et ModÃ¨le sÃ©parÃ©s
        self.vehicles_table.setItem(row, 4, QTableWidgetItem(vehicle.get('marque', '')))
        self.vehicles_table.setItem(row, 5, QTableWidgetItem(vehicle.get('modele', '')))
        
        self.vehicles_table.setItem(row, 6, QTableWidgetItem(vehicle.get('genre', 'VP')))
        
        puissance_item = QTableWidgetItem(str(vehicle.get('puissance', 0)))
        puissance_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.vehicles_table.setItem(row, 7, puissance_item)
        
        energie = vehicle.get('energie', 'SEE')
        energie_labels = {
            'SEE': 'Essence',
            'SED': 'Diesel',
            'HYBRIDE': 'Hybride',
            'GAZ': 'Gaz',
            'ELECTRIQUE': 'Ã‰lectrique'
        }
        energie_label = energie_labels.get(energie, energie)
        energie_item = QTableWidgetItem(energie_label)
        energie_item.setTextAlignment(Qt.AlignCenter)
        self.vehicles_table.setItem(row, 8, energie_item)
        
        self.vehicles_table.setItem(row, 9, QTableWidgetItem(vehicle.get('usage', '')))
        
        places_item = QTableWidgetItem(str(vehicle.get('places', 5)))
        places_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.vehicles_table.setItem(row, 10, places_item)
        
        jours_item = QTableWidgetItem(str(vehicle.get('nbr_jour', 365)))
        jours_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.vehicles_table.setItem(row, 11, jours_item)
        
        # Valeurs
        pmec_value = vehicle.get('pmec', '')
        pmec_item = QTableWidgetItem(str(pmec_value) if pmec_value else '')
        pmec_item.setTextAlignment(Qt.AlignCenter)
        self.vehicles_table.setItem(row, 12, pmec_item)
        
        vn_item = QTableWidgetItem(f"{vehicle.get('valeur_neuf', 0):,.0f}".replace(",", " "))
        vn_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.vehicles_table.setItem(row, 13, vn_item)
        
        vv_item = QTableWidgetItem(f"{vehicle.get('valeur_venale', 0):,.0f}".replace(",", " "))
        vv_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.vehicles_table.setItem(row, 14, vv_item)
        
        capital_item = QTableWidgetItem(f"{vehicle.get('capital_ar', 0):,.0f}".replace(",", " "))
        capital_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.vehicles_table.setItem(row, 15, capital_item)
        
        # Garanties avec les nouvelles positions
        garanties = vehicle.get('garanties', {})
        garanties_cols = {
            'rc': 16,
            'dr': 17,
            'vol': 18,
            'vb': 18,  # MÃªme colonne que Vol/Vol partiel
            'incendie': 19,
            'bris_glace': 20,
            'ar': 21,
            'dta': 22,
            'ipt': 23
        }
        
        for key, col in garanties_cols.items():
            amount = garanties.get(key, 0)
            item = QTableWidgetItem(f"{amount:,.0f}".replace(",", " "))
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.vehicles_table.setItem(row, col, item)
        
        # âœ… CORRECTION : RÃ©cupÃ©rer Prime brute en toute sÃ©curitÃ©
        prime_brute = 0
        
        # Essayer diffÃ©rentes clÃ©s possibles
        if 'Prime brute' in vehicle:
            prime_brute = vehicle.get('Prime brute', 0)
        elif 'Prime_brute' in vehicle:
            prime_brute = vehicle.get('Prime_brute', 0)
        elif 'prime_brute' in vehicle:
            prime_brute = vehicle.get('prime_brute', 0)
        elif 'garanties' in vehicle and 'total' in vehicle['garanties']:
            prime_brute = vehicle['garanties'].get('total', 0)
        
        # Si prime_brute est une chaÃ®ne, la nettoyer et la convertir
        if isinstance(prime_brute, str):
            try:
                # Supprimer les espaces insÃ©cables et les sÃ©parateurs de milliers
                import re
                cleaned = re.sub(r'[^\d.,]', '', prime_brute)
                cleaned = cleaned.replace(',', '.').replace(' ', '')
                prime_brute = float(cleaned) if cleaned else 0
            except (ValueError, TypeError):
                prime_brute = 0
        
        # S'assurer que prime_brute est un nombre
        try:
            prime_brute = float(prime_brute) if prime_brute else 0
        except (ValueError, TypeError):
            prime_brute = 0
        
        pb_item = QTableWidgetItem(f"{prime_brute:,.0f}".replace(",", " "))
        pb_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.vehicles_table.setItem(row, 24, pb_item)
        
        reductions_item = QTableWidgetItem(str(vehicle.get('reductions', 0)))
        reductions_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.vehicles_table.setItem(row, 25, reductions_item)
        
        prime_nette = vehicle.get('prime_nette', prime_brute)
        pn_item = QTableWidgetItem(f"{prime_nette:,.0f}".replace(",", " "))
        pn_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.vehicles_table.setItem(row, 26, pn_item)
        
        timbre_item = QTableWidgetItem(f"{vehicle.get('droit_timbre', 0):,.0f}".replace(",", " "))
        timbre_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.vehicles_table.setItem(row, 27, timbre_item)

    def import_fleet(self):
        """Importe la flotte dans la base de donnÃ©es"""
        # RÃ©cupÃ©rer les vÃ©hicules sÃ©lectionnÃ©s
        selected = []
        for row in range(self.vehicles_table.rowCount()):
            item = self.vehicles_table.item(row, 0)
            if item and item.checkState() == Qt.Checked:
                if row < len(self.vehicles_data):
                    selected.append(self.vehicles_data[row])
                    # self.apply_global_frais_to_selected()
        
        if not selected:
            QMessageBox.warning(self, "Erreur", "Aucun vÃ©hicule sÃ©lectionnÃ©")
            return
        
        # âœ… VÃ‰RIFIER QU'UNE COMPAGNIE EST SÃ‰LECTIONNÃ‰E
        # RÃ©cupÃ©rer l'ID de la compagnie depuis la combo box
        compagny_id = self.compagny_combo.currentData() if hasattr(self, 'compagny_combo') else None
        
        # Si la combo box n'existe pas ou n'est pas dans l'UI, on la crÃ©e
        if not hasattr(self, 'compagny_combo'):
            # Demander Ã  l'utilisateur de sÃ©lectionner une compagnie
            compagny_id = self._ask_for_compagny()
            if not compagny_id:
                return
        elif not compagny_id:
            QMessageBox.warning(self, "Erreur", "Veuillez sÃ©lectionner une compagnie d'assurance")
            return
        
        try:
            # RÃ©cupÃ©rer l'ID du propriÃ©taire
            owner_id = None
            if hasattr(self.parent(), 'contact'):
                owner_id = self.parent().contact.id
            
            if hasattr(self.controller, 'current_user_id'):
                current_user_id = self.controller.current_user_id
            else:
                current_user_id = 1
            
            # CrÃ©er ou rÃ©cupÃ©rer la flotte
            fleet_id = None
            
            if self.mode_new.isChecked():
                fleet_name = self.fleet_name.text().strip()
                if not fleet_name:
                    QMessageBox.warning(self, "Erreur", "Veuillez entrer un nom de flotte")
                    return
                
                fleet_data = {
                    'nom_flotte': fleet_name,
                    'code_flotte': self.fleet_code.text().strip(),
                    'owner_id': owner_id,
                    'statut': 'Actif',
                    'assureur': compagny_id,  # âœ… Utiliser l'ID de la compagnie
                    'date_debut': self.date_debut.date().toPython(),
                    'date_fin': self.date_fin.date().toPython(),
                }
                
                success, result = self.controller.fleets.create_fleet(fleet_data, current_user_id)
                if not success:
                    QMessageBox.critical(self, "Erreur", f"Erreur crÃ©ation flotte: {result}")
                    return
                fleet_id = result.id if hasattr(result, 'id') else result
            else:
                fleet_id = self.existing_fleet_combo.currentData()
                if not fleet_id:
                    QMessageBox.warning(self, "Erreur", "Veuillez sÃ©lectionner une flotte")
                    return
            
            # âœ… IMPORTER LES VÃ‰HICULES
            imported = 0
            errors = []
            
            for vehicle in selected:
                try:
                    # PrÃ©parer les donnÃ©es du vÃ©hicule
                    garanties = vehicle.get('garanties', {})
                    
                    chassis_value = vehicle.get('chassis', '')
                    if not chassis_value or chassis_value == '':
                        chassis_value = f"CH-{vehicle['immatriculation']}"
                    
                    # Dates
                    if vehicle.get('date_debut'):
                        debut = vehicle['date_debut']
                        if isinstance(debut, datetime):
                            debut = debut.date()
                    else:
                        debut = self.date_debut.date().toPython()
                    
                    if vehicle.get('date_fin'):
                        fin = vehicle['date_fin']
                        if isinstance(fin, datetime):
                            fin = fin.date()
                    else:
                        fin = self.date_fin.date().toPython()
                    
                    jours = max(1, (fin - debut).days) if fin and debut else 365
                    
                    # CatÃ©gorie
                    categorie_value = vehicle.get('categorie', 'VP')
                    if not categorie_value:
                        categorie_value = 'VP'
                    
                    # Calcul TVA
                    tva_rate = 0.1925
                    total_garanties = garanties.get('total', 0)
                    tva_amount = total_garanties * tva_rate
                    
                    # RÃ©cupÃ©rer les frais
                    accessoires = vehicle.get('accessoires', 0)
                    asac = vehicle.get('asac', 0)
                    carte_rose = vehicle.get('carte_rose', 0)
                    vignette = vehicle.get('vignette', 0)
                    
                    # âœ… PRÃ‰PARER LES DONNÃ‰ES DU VÃ‰HICULE
                    vehicle_data = build_vehicle_import_payload(
                        vehicle=vehicle,
                        owner_id=owner_id,
                        compagny_id=compagny_id,
                        fleet_id=fleet_id,
                        current_user_id=current_user_id,
                        debut=debut,
                        fin=fin,
                        jours=jours,
                        total_garanties=total_garanties,
                        tva_amount=tva_amount,
                        accessoires=accessoires,
                        asac=asac,
                        carte_rose=carte_rose,
                        vignette=vignette,
                    )
                    
                    print(vehicle_data)
                    # âœ… APPEL AU CONTROLLER
                    result = self.controller.vehicles.create_vehicle(vehicle_data, current_user_id)
                    
                    if isinstance(result, tuple):
                        if len(result) == 2:
                            success, message = result
                        elif len(result) == 3:
                            success, _, message = result
                        else:
                            success = False
                            message = "Format de retour inattendu"
                    else:
                        success = bool(result)
                        message = "SuccÃ¨s" if success else "Erreur"
                    
                    if success:
                        imported += 1
                        print(f"âœ… VÃ©hicule {vehicle['immatriculation']} importÃ© avec succÃ¨s")
                    else:
                        errors.append(f"{vehicle['immatriculation']}: {message}")
                        
                except Exception as e:
                    errors.append(f"{vehicle['immatriculation']}: {str(e)}")
                    traceback.print_exc()
            
            # âœ… AFFICHER LE RÃ‰SULTAT
            if imported > 0:
                msg = f"âœ… {imported} vÃ©hicule(s) importÃ©s avec succÃ¨s"
                if errors:
                    msg += f"\n\nâš ï¸ {len(errors)} erreur(s):\n" + "\n".join(errors[:5])
                QMessageBox.information(self, "Importation terminÃ©e", msg)
                
                # âœ… Ã‰mettre le signal pour rafraÃ®chir la vue parente
                self.data_changed.emit()
                self.accept()
            else:
                QMessageBox.critical(self, "Erreur", "\n".join(errors[:5]))
                
        except Exception as e:
            QMessageBox.critical(self, "Erreur", str(e))
            traceback.print_exc()

    def _ask_for_compagny(self):
        """Demande Ã  l'utilisateur de sÃ©lectionner une compagnie"""
        from PySide6.QtWidgets import QInputDialog, QComboBox, QDialogButtonBox, QVBoxLayout
        
        # CrÃ©er un dialogue simple
        dialog = QDialog(self)
        dialog.setWindowTitle("SÃ©lectionner une compagnie")
        dialog.setModal(True)
        dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("Veuillez sÃ©lectionner une compagnie d'assurance :"))
        
        combo = QComboBox()
        # Charger les compagnies
        try:
            if hasattr(self.controller, 'compagnies'):
                compagnies = self.controller.compagnies.get_all_active_compagnies()
                for cie in compagnies:
                    if hasattr(cie, 'nom'):
                        combo.addItem(cie.nom, cie.id)
                    elif isinstance(cie, dict):
                        combo.addItem(cie.get('nom', ''), cie.get('id'))
        except Exception as e:
            print(f"Erreur chargement compagnies: {e}")
        
        layout.addWidget(combo)
        
        # Boutons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        if dialog.exec() == QDialog.Accepted:
            return combo.currentData()
        return None
    
    def _load_compagnies(self):
        """Charge les compagnies (simplifiÃ©)"""
        pass

    def create_compagny_section(self):
        """Section de sÃ©lection de la compagnie"""
        group = QGroupBox("ðŸ¢ 2.5. Compagnie d'assurance")
        layout = QVBoxLayout(group)
        
        self.compagny_combo = QComboBox()
        self.compagny_combo.addItem("SÃ©lectionner une compagnie", None)
        
        # Charger les compagnies
        try:
            if hasattr(self.controller, 'compagnies'):
                compagnies = self.controller.compagnies.get_all()
                for cie in compagnies:
                    if hasattr(cie, 'nom'):
                        self.compagny_combo.addItem(cie.nom, cie.id)
                    elif isinstance(cie, dict):
                        self.compagny_combo.addItem(cie.get('nom', ''), cie.get('id'))
        except Exception as e:
            print(f"Erreur chargement compagnies: {e}")
        
        layout.addWidget(self.compagny_combo)
        
        # Label d'information
        info = QLabel("âš ï¸ La compagnie est obligatoire pour l'importation")
        info.setStyleSheet("color: #ef4444; font-size: 11px;")
        layout.addWidget(info)
        
        return group
    
# ============================================================================
# FONCTION PRINCIPALE D'EXPORT
# ============================================================================

def create_fleet_import_dialog(controller, parent=None):
    """CrÃ©e et retourne le dialogue d'importation de flotte"""
    return FleetImportAdvancedDialog(controller, parent)