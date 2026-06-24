# # addons/Automobiles/views/reports_view.py
# """
# Vue des rapports et statistiques - Style Power BI
# Design professionnel avec graphiques interactifs et exports avancés
# """

# from PySide6.QtWidgets import (
#     QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
#     QFrame, QComboBox, QDateEdit, QGroupBox, QGridLayout,
#     QScrollArea, QProgressBar, QTableWidget, QTableWidgetItem,
#     QHeaderView, QMessageBox, QFileDialog, QSplitter,
#     QCheckBox, QTabWidget, QLineEdit, QTextEdit, QMenu,
#     QToolButton, QSizePolicy, QSpacerItem, QDialog,
#     QDialogButtonBox, QFormLayout
# )
# from PySide6.QtCore import Qt, QDate, QDateTime, Signal, QTimer, QSettings
# from PySide6.QtGui import QColor, QFont, QPainter, QPen, QAction, QIcon
# from PySide6.QtCharts import (
#     QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis,
#     QValueAxis, QLineSeries, QPieSeries, QPieSlice, QHorizontalBarSeries,
#     QScatterSeries, QAreaSeries, QSplineSeries
# )

# from datetime import datetime, timedelta
# from typing import Dict, Any, List
# import json
# import csv
# import os

# from addons.Automobiles.views.style import Colors, Fonts, Spacing
# from addons.Automobiles.views.widgets.modern_card import ModernCard


# class ExportDialog(QDialog):
#     """Dialogue d'export avancé"""
    
#     def __init__(self, parent=None):
#         super().__init__(parent)
#         self.setWindowTitle("📤 Exporter le rapport")
#         self.setMinimumSize(500, 400)
#         self.setup_ui()
    
#     def setup_ui(self):
#         layout = QVBoxLayout(self)
#         layout.setSpacing(Spacing.LG)
        
#         # Titre
#         title = QLabel("Options d'export")
#         title.setStyleSheet(f"font-size: {Fonts.H3}px; font-weight: {Fonts.BOLD};")
#         layout.addWidget(title)
        
#         # Formats
#         group = QGroupBox("Format")
#         group_layout = QVBoxLayout(group)
        
#         self.format_combo = QComboBox()
#         self.format_combo.addItems([
#             "PDF - Document imprimable",
#             "CSV - Données tabulaires",
#             "JSON - Données structurées",
#             "Excel - Tableur (XLSX)",
#             "HTML - Page web",
#             "Image PNG - Graphique",
#             "Image SVG - Vectoriel"
#         ])
#         group_layout.addWidget(self.format_combo)
#         layout.addWidget(group)
        
#         # Options
#         options_group = QGroupBox("Options")
#         options_layout = QGridLayout(options_group)
        
#         self.include_charts = QCheckBox("Inclure les graphiques")
#         self.include_charts.setChecked(True)
#         options_layout.addWidget(self.include_charts, 0, 0)
        
#         self.include_tables = QCheckBox("Inclure les tableaux")
#         self.include_tables.setChecked(True)
#         options_layout.addWidget(self.include_tables, 0, 1)
        
#         self.include_summary = QCheckBox("Inclure le résumé")
#         self.include_summary.setChecked(True)
#         options_layout.addWidget(self.include_summary, 1, 0)
        
#         self.include_raw_data = QCheckBox("Inclure les données brutes")
#         options_layout.addWidget(self.include_raw_data, 1, 1)
        
#         layout.addWidget(options_group)
        
#         # Période
#         period_group = QGroupBox("Période")
#         period_layout = QGridLayout(period_group)
        
#         period_layout.addWidget(QLabel("De:"), 0, 0)
#         self.period_start = QDateEdit()
#         self.period_start.setCalendarPopup(True)
#         self.period_start.setDate(QDate.currentDate().addMonths(-1))
#         period_layout.addWidget(self.period_start, 0, 1)
        
#         period_layout.addWidget(QLabel("À:"), 1, 0)
#         self.period_end = QDateEdit()
#         self.period_end.setCalendarPopup(True)
#         self.period_end.setDate(QDate.currentDate())
#         period_layout.addWidget(self.period_end, 1, 1)
        
#         layout.addWidget(period_group)
        
#         # Boutons
#         buttons = QDialogButtonBox()
#         export_btn = buttons.addButton("📤 Exporter", QDialogButtonBox.AcceptRole)
#         export_btn.setStyleSheet(f"""
#             QPushButton {{
#                 background-color: {Colors.PRIMARY};
#                 color: white;
#                 padding: 10px 20px;
#                 border-radius: 8px;
#                 font-weight: {Fonts.SEMIBOLD};
#             }}
#         """)
#         cancel_btn = buttons.addButton("Annuler", QDialogButtonBox.RejectRole)
#         buttons.accepted.connect(self.accept)
#         buttons.rejected.connect(self.reject)
#         layout.addWidget(buttons)


# class ReportView(QWidget):
#     """Vue des rapports et statistiques - Style Power BI"""
    
#     def __init__(self, controller, user=None):
#         super().__init__()
#         self.controller = controller
#         self.user = user
#         self.current_data = {}
#         self.charts = {}
#         self.settings = QSettings("LOMETA", "Reports")
        
#         self.setup_ui()
#         self.load_dashboard()
    
#     def setup_ui(self):
#         layout = QVBoxLayout(self)
#         layout.setContentsMargins(0, 0, 0, 0)
#         layout.setSpacing(0)
        
#         # En-tête
#         header = self._create_header()
#         layout.addWidget(header)
        
#         # Barre d'outils
#         toolbar = self._create_toolbar()
#         layout.addWidget(toolbar)
        
#         # Scroll area
#         scroll = QScrollArea()
#         scroll.setWidgetResizable(True)
#         scroll.setFrameShape(QScrollArea.NoFrame)
#         scroll.setStyleSheet("background: #f1f5f9; border: none;")
        
#         content = QWidget()
#         content_layout = QVBoxLayout(content)
#         content_layout.setContentsMargins(Spacing.LG, Spacing.LG, Spacing.LG, Spacing.LG)
#         content_layout.setSpacing(Spacing.XL)
        
#         # Ligne 1: KPI Cards
#         self._create_kpi_cards()
#         content_layout.addWidget(self.kpi_cards_container)
        
#         # Ligne 2: Graphiques principaux (2 colonnes)
#         row2 = QHBoxLayout()
#         row2.setSpacing(Spacing.XL)
        
#         self._create_contracts_chart()
#         self._create_revenue_chart()
        
#         row2.addWidget(self.contracts_chart_card, 1)
#         row2.addWidget(self.revenue_chart_card, 1)
#         content_layout.addLayout(row2)
        
#         # Ligne 3: Graphiques secondaires (3 colonnes)
#         row3 = QHBoxLayout()
#         row3.setSpacing(Spacing.XL)
        
#         self._create_category_pie()
#         self._create_zone_chart()
#         self._create_energy_chart()
        
#         row3.addWidget(self.category_chart_card, 1)
#         row3.addWidget(self.zone_chart_card, 1)
#         row3.addWidget(self.energy_chart_card, 1)
#         content_layout.addLayout(row3)
        
#         # Ligne 4: Tableaux
#         row4 = QHBoxLayout()
#         row4.setSpacing(Spacing.XL)
        
#         self._create_top_vehicles()
#         self._create_top_clients()
        
#         row4.addWidget(self.top_vehicles_card, 1)
#         row4.addWidget(self.top_clients_card, 1)
#         content_layout.addLayout(row4)
        
#         # Ligne 5: Rapport détaillé
#         self._create_detailed_report()
#         content_layout.addWidget(self.detailed_report_card)
        
#         scroll.setWidget(content)
#         layout.addWidget(scroll)
        
#         # Barre de statut
#         self._create_status_bar()
#         layout.addWidget(self.status_bar)
    
#     def _create_header(self):
#         """Crée l'en-tête"""
#         header = QFrame()
#         header.setStyleSheet(f"""
#             QFrame {{
#                 background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
#                     stop:0 #1e293b, stop:1 #0f172a);
#                 padding: 16px 24px;
#             }}
#         """)
        
#         layout = QHBoxLayout(header)
        
#         title = QLabel("📊 Centre de Rapports & Analytics")
#         title.setStyleSheet("color: white; font-size: 22px; font-weight: 800;")
        
#         # Date actuelle
#         date_label = QLabel(datetime.now().strftime("%d %B %Y"))
#         date_label.setStyleSheet("color: #94a3b8; font-size: 13px;")
        
#         layout.addWidget(title)
#         layout.addStretch()
#         layout.addWidget(date_label)
        
#         return header
    
#     def _create_toolbar(self):
#         """Crée la barre d'outils"""
#         toolbar = QFrame()
#         toolbar.setStyleSheet(f"""
#             QFrame {{
#                 background: white;
#                 border-bottom: 1px solid #e2e8f0;
#                 padding: 8px 16px;
#             }}
#         """)
        
#         layout = QHBoxLayout(toolbar)
#         layout.setSpacing(Spacing.MD)
        
#         # Période
#         layout.addWidget(QLabel("Période:"))
        
#         self.period_combo = QComboBox()
#         self.period_combo.addItems(["Aujourd'hui", "7 jours", "30 jours", "Ce mois", "Ce trimestre", "Cette année", "Personnalisé"])
#         self.period_combo.setFixedWidth(140)
#         self.period_combo.currentTextChanged.connect(self.on_period_change)
#         layout.addWidget(self.period_combo)
        
#         # Date début
#         self.start_date = QDateEdit()
#         self.start_date.setCalendarPopup(True)
#         self.start_date.setDate(QDate.currentDate().addMonths(-1))
#         self.start_date.setFixedWidth(120)
#         self.start_date.dateChanged.connect(self.on_date_change)
#         layout.addWidget(self.start_date)
        
#         # Date fin
#         self.end_date = QDateEdit()
#         self.end_date.setCalendarPopup(True)
#         self.end_date.setDate(QDate.currentDate())
#         self.end_date.setFixedWidth(120)
#         self.end_date.dateChanged.connect(self.on_date_change)
#         layout.addWidget(self.end_date)
        
#         # Bouton générer
#         self.generate_btn = QPushButton("🔄 Générer")
#         self.generate_btn.setStyleSheet(f"""
#             QPushButton {{
#                 background-color: {Colors.PRIMARY};
#                 color: white;
#                 border-radius: 8px;
#                 padding: 8px 20px;
#                 font-weight: {Fonts.SEMIBOLD};
#             }}
#             QPushButton:hover {{
#                 background-color: {Colors.PRIMARY_DARK};
#             }}
#         """)
#         self.generate_btn.clicked.connect(self.generate_report)
#         layout.addWidget(self.generate_btn)
        
#         layout.addStretch()
        
#         # Boutons d'export
#         self.export_btn = QPushButton("📤 Exporter")
#         self.export_btn.setStyleSheet(f"""
#             QPushButton {{
#                 background-color: #10b981;
#                 color: white;
#                 border-radius: 8px;
#                 padding: 8px 20px;
#                 font-weight: {Fonts.SEMIBOLD};
#             }}
#             QPushButton:hover {{
#                 background-color: #059669;
#             }}
#         """)
#         self.export_btn.clicked.connect(self.show_export_dialog)
#         layout.addWidget(self.export_btn)
        
#         # Bouton partager
#         share_btn = QPushButton("📧 Partager")
#         share_btn.setStyleSheet(f"""
#             QPushButton {{
#                 background-color: #8b5cf6;
#                 color: white;
#                 border-radius: 8px;
#                 padding: 8px 20px;
#                 font-weight: {Fonts.SEMIBOLD};
#             }}
#             QPushButton:hover {{
#                 background-color: #7c3aed;
#             }}
#         """)
#         share_btn.clicked.connect(self.share_report)
#         layout.addWidget(share_btn)
        
#         return toolbar
    
#     def _create_kpi_cards(self):
#         """Crée les cartes KPI"""
#         self.kpi_cards_container = QFrame()
#         self.kpi_cards_container.setStyleSheet("""
#             QFrame {
#                 background: transparent;
#             }
#         """)
        
#         layout = QGridLayout(self.kpi_cards_container)
#         layout.setSpacing(Spacing.MD)
        
#         kpis = [
#             ("💰", "Chiffre d'affaires", "0 XAF", "+12.5%", Colors.PRIMARY),
#             ("📄", "Contrats actifs", "0", "+8.3%", Colors.SUCCESS),
#             ("🚗", "Véhicules", "0", "+5.7%", Colors.INFO),
#             ("👥", "Clients", "0", "+15.2%", Colors.WARNING),
#             ("⭐", "Taux de satisfaction", "0%", "+2.1%", "#8b5cf6"),
#             ("📊", "Taux de conversion", "0%", "-1.2%", Colors.DANGER)
#         ]
        
#         self.kpi_cards = {}
#         for i, (icon, title, value, trend, color) in enumerate(kpis):
#             card = self._create_kpi_card(icon, title, value, trend, color)
#             layout.addWidget(card, i // 3, i % 3)
#             self.kpi_cards[title] = card
    
#     def _create_kpi_card(self, icon, title, value, trend, color):
#         """Crée une carte KPI individuelle"""
#         card = QFrame()
#         card.setStyleSheet(f"""
#             QFrame {{
#                 background: white;
#                 border-radius: 16px;
#                 border: 1px solid #e2e8f0;
#                 padding: 16px;
#             }}
#             QFrame:hover {{
#                 border-color: {color};
#                 box-shadow: 0 4px 12px rgba(0,0,0,0.05);
#             }}
#         """)
        
#         layout = QVBoxLayout(card)
#         layout.setSpacing(Spacing.XS)
        
#         # Ligne 1: Icône + Trend
#         header_layout = QHBoxLayout()
        
#         icon_label = QLabel(icon)
#         icon_label.setStyleSheet(f"font-size: 24px;")
#         header_layout.addWidget(icon_label)
        
#         trend_label = QLabel(trend)
#         if trend.startswith('+'):
#             trend_label.setStyleSheet(f"color: {Colors.SUCCESS}; font-weight: {Fonts.BOLD}; font-size: 11px;")
#         elif trend.startswith('-'):
#             trend_label.setStyleSheet(f"color: {Colors.DANGER}; font-weight: {Fonts.BOLD}; font-size: 11px;")
#         else:
#             trend_label.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; font-size: 11px;")
#         header_layout.addStretch()
#         header_layout.addWidget(trend_label)
        
#         layout.addLayout(header_layout)
        
#         # Valeur
#         value_label = QLabel(value)
#         value_label.setStyleSheet(f"""
#             font-size: 28px;
#             font-weight: 800;
#             color: {color};
#         """)
#         layout.addWidget(value_label)
        
#         # Titre
#         title_label = QLabel(title)
#         title_label.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; font-size: 12px;")
#         layout.addWidget(title_label)
        
#         # Stocker la référence
#         card.value_label = value_label
        
#         return card
    
#     def _create_contracts_chart(self):
#         """Crée le graphique des contrats"""
#         self.contracts_chart_card = ModernCard(title="📈 Évolution des contrats", icon="📊")
        
#         self.contracts_chart = QChart()
#         self.contracts_chart.setBackgroundVisible(False)
#         self.contracts_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
#         view = QChartView(self.contracts_chart)
#         view.setRenderHint(QPainter.Antialiasing)
#         view.setMinimumHeight(300)
        
#         self.contracts_chart_card.add_widget(view)
    
#     def _create_revenue_chart(self):
#         """Crée le graphique des revenus"""
#         self.revenue_chart_card = ModernCard(title="💰 Revenus mensuels", icon="💰")
        
#         self.revenue_chart = QChart()
#         self.revenue_chart.setBackgroundVisible(False)
#         self.revenue_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
#         view = QChartView(self.revenue_chart)
#         view.setRenderHint(QPainter.Antialiasing)
#         view.setMinimumHeight(300)
        
#         self.revenue_chart_card.add_widget(view)
    
#     def _create_category_pie(self):
#         """Crée le graphique en secteurs par catégorie"""
#         self.category_chart_card = ModernCard(title="🏷️ Répartition par catégorie", icon="📊")
        
#         self.category_chart = QChart()
#         self.category_chart.setBackgroundVisible(False)
#         self.category_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
#         view = QChartView(self.category_chart)
#         view.setRenderHint(QPainter.Antialiasing)
#         view.setMinimumHeight(250)
        
#         self.category_chart_card.add_widget(view)
    
#     def _create_zone_chart(self):
#         """Crée le graphique par zone"""
#         self.zone_chart_card = ModernCard(title="🗺️ Répartition par zone", icon="📍")
        
#         self.zone_chart = QChart()
#         self.zone_chart.setBackgroundVisible(False)
#         self.zone_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
#         view = QChartView(self.zone_chart)
#         view.setRenderHint(QPainter.Antialiasing)
#         view.setMinimumHeight(250)
        
#         self.zone_chart_card.add_widget(view)
    
#     def _create_energy_chart(self):
#         """Crée le graphique par énergie"""
#         self.energy_chart_card = ModernCard(title="⛽ Répartition par énergie", icon="🔋")
        
#         self.energy_chart = QChart()
#         self.energy_chart.setBackgroundVisible(False)
#         self.energy_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
#         view = QChartView(self.energy_chart)
#         view.setRenderHint(QPainter.Antialiasing)
#         view.setMinimumHeight(250)
        
#         self.energy_chart_card.add_widget(view)
    
#     def _create_top_vehicles(self):
#         """Crée le tableau des top véhicules"""
#         self.top_vehicles_card = ModernCard(title="🚗 Top 10 véhicules", icon="🏆")
        
#         self.top_vehicles_table = QTableWidget()
#         self.top_vehicles_table.setColumnCount(4)
#         self.top_vehicles_table.setHorizontalHeaderLabels(["#", "Immatriculation", "Marque", "Contrats"])
#         self.top_vehicles_table.setAlternatingRowColors(True)
#         self.top_vehicles_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
#         self.top_vehicles_table.verticalHeader().setVisible(False)
#         self.top_vehicles_table.setFixedHeight(200)
        
#         self.top_vehicles_card.add_widget(self.top_vehicles_table)
    
#     def _create_top_clients(self):
#         """Crée le tableau des top clients"""
#         self.top_clients_card = ModernCard(title="👥 Top 10 clients", icon="🏆")
        
#         self.top_clients_table = QTableWidget()
#         self.top_clients_table.setColumnCount(4)
#         self.top_clients_table.setHorizontalHeaderLabels(["#", "Client", "Véhicules", "Prime totale"])
#         self.top_clients_table.setAlternatingRowColors(True)
#         self.top_clients_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
#         self.top_clients_table.verticalHeader().setVisible(False)
#         self.top_clients_table.setFixedHeight(200)
        
#         self.top_clients_card.add_widget(self.top_clients_table)
    
#     def _create_detailed_report(self):
#         """Crée le rapport détaillé"""
#         self.detailed_report_card = ModernCard(title="📋 Rapport détaillé", icon="📄")
        
#         self.report_text = QTextEdit()
#         self.report_text.setReadOnly(True)
#         self.report_text.setStyleSheet("""
#             QTextEdit {
#                 font-family: 'Courier New', monospace;
#                 font-size: 12px;
#                 background: #f8fafc;
#                 border: 1px solid #e2e8f0;
#                 border-radius: 8px;
#                 padding: 16px;
#             }
#         """)
#         self.report_text.setMinimumHeight(200)
        
#         self.detailed_report_card.add_widget(self.report_text)
    
#     def _create_status_bar(self):
#         """Crée la barre de statut"""
#         self.status_bar = QFrame()
#         self.status_bar.setStyleSheet(f"""
#             QFrame {{
#                 background: white;
#                 border-top: 1px solid #e2e8f0;
#                 padding: 8px 16px;
#             }}
#         """)
        
#         layout = QHBoxLayout(self.status_bar)
        
#         self.status_label = QLabel("✅ Prêt")
#         self.status_label.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; font-size: 12px;")
#         layout.addWidget(self.status_label)
        
#         layout.addStretch()
        
#         self.last_update_label = QLabel(f"Dernière mise à jour: {datetime.now().strftime('%H:%M:%S')}")
#         self.last_update_label.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 11px;")
#         layout.addWidget(self.last_update_label)
        
#         self.progress_bar = QProgressBar()
#         self.progress_bar.setFixedWidth(150)
#         self.progress_bar.setVisible(False)
#         layout.addWidget(self.progress_bar)
    
#     def load_dashboard(self):
#         """Charge le tableau de bord"""
#         self.generate_report()
    
#     def generate_report(self):
#         """Génère le rapport complet"""
#         self.status_label.setText("🔄 Génération du rapport...")
#         self.progress_bar.setVisible(True)
#         self.progress_bar.setRange(0, 100)
        
#         try:
#             start = self.start_date.date().toPython()
#             end = self.end_date.date().toPython()
            
#             # Simuler la progression
#             self.progress_bar.setValue(20)
            
#             # Mettre à jour les KPI
#             self._update_kpis()
            
#             self.progress_bar.setValue(40)
            
#             # Mettre à jour les graphiques
#             self._update_contracts_chart()
#             self._update_revenue_chart()
#             self._update_category_pie()
#             self._update_zone_chart()
#             self._update_energy_chart()
            
#             self.progress_bar.setValue(70)
            
#             # Mettre à jour les tableaux
#             self._update_top_vehicles()
#             self._update_top_clients()
#             self._update_detailed_report()
            
#             self.progress_bar.setValue(100)
            
#             self.status_label.setText(f"✅ Rapport généré: {start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}")
#             self.last_update_label.setText(f"Dernière mise à jour: {datetime.now().strftime('%H:%M:%S')}")
            
#         except Exception as e:
#             self.status_label.setText(f"❌ Erreur: {str(e)}")
#             QMessageBox.warning(self, "Erreur", f"Erreur lors de la génération: {e}")
#         finally:
#             QTimer.singleShot(500, lambda: self.progress_bar.setVisible(False))
    
#     def _update_kpis(self):
#         """Met à jour les KPI"""
#         # Simuler des données (à remplacer par des données réelles)
#         kpi_data = {
#             "Chiffre d'affaires": ("12 450 000 XAF", "+12.5%"),
#             "Contrats actifs": ("156", "+8.3%"),
#             "Véhicules": ("342", "+5.7%"),
#             "Clients": ("89", "+15.2%"),
#             "Taux de satisfaction": ("94%", "+2.1%"),
#             "Taux de conversion": ("68%", "-1.2%")
#         }
        
#         for title, (value, trend) in kpi_data.items():
#             if title in self.kpi_cards:
#                 card = self.kpi_cards[title]
#                 card.value_label.setText(value)
    
#     def _update_contracts_chart(self):
#         """Met à jour le graphique des contrats"""
#         self.contracts_chart.removeAllSeries()
        
#         # Données simulées
#         months = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Aoû", "Sep", "Oct", "Nov", "Déc"]
#         values = [45, 52, 48, 55, 62, 58, 65, 72, 78, 85, 82, 90]
        
#         # Ligne
#         line_series = QLineSeries()
#         line_series.setName("Contrats")
#         for i, value in enumerate(values):
#             line_series.append(i, value)
        
#         pen = QPen(QColor(Colors.PRIMARY))
#         pen.setWidth(2)
#         line_series.setPen(pen)
        
#         self.contracts_chart.addSeries(line_series)
        
#         # Barres
#         bar_set = QBarSet("Contrats")
#         for value in values:
#             bar_set.append(value)
#         bar_set.setColor(QColor(Colors.PRIMARY))
        
#         bar_series = QBarSeries()
#         bar_series.append(bar_set)
#         bar_series.setBarWidth(0.4)
        
#         self.contracts_chart.addSeries(bar_series)
        
#         # Axes
#         axis_x = QBarCategoryAxis()
#         axis_x.append(months)
#         self.contracts_chart.addAxis(axis_x, Qt.AlignBottom)
#         bar_series.attachAxis(axis_x)
#         line_series.attachAxis(axis_x)
        
#         axis_y = QValueAxis()
#         axis_y.setRange(0, 100)
#         axis_y.setTitleText("Nombre de contrats")
#         self.contracts_chart.addAxis(axis_y, Qt.AlignLeft)
#         bar_series.attachAxis(axis_y)
#         line_series.attachAxis(axis_y)
        
#         self.contracts_chart.setTitle("Évolution des contrats 2025")
    
#     def _update_revenue_chart(self):
#         """Met à jour le graphique des revenus"""
#         self.revenue_chart.removeAllSeries()
        
#         months = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Aoû", "Sep", "Oct", "Nov", "Déc"]
#         values = [1250000, 1380000, 1520000, 1620000, 1750000, 1850000, 1920000, 1980000, 2050000, 2120000, 2150000, 2200000]
        
#         bar_set = QBarSet("Revenus (XAF)")
#         for value in values:
#             bar_set.append(value)
#         bar_set.setColor(QColor(Colors.SUCCESS))
        
#         series = QBarSeries()
#         series.append(bar_set)
#         series.setBarWidth(0.6)
        
#         self.revenue_chart.addSeries(series)
        
#         axis_x = QBarCategoryAxis()
#         axis_x.append(months)
#         self.revenue_chart.addAxis(axis_x, Qt.AlignBottom)
#         series.attachAxis(axis_x)
        
#         axis_y = QValueAxis()
#         axis_y.setRange(0, 2500000)
#         axis_y.setTitleText("Revenus (XAF)")
#         self.revenue_chart.addAxis(axis_y, Qt.AlignLeft)
#         series.attachAxis(axis_y)
        
#         self.revenue_chart.setTitle("Revenus mensuels 2025")
    
#     def _update_category_pie(self):
#         """Met à jour le graphique en secteurs par catégorie"""
#         self.category_chart.removeAllSeries()
        
#         categories = [
#             ("VP - Particuliers", 45),
#             ("VU - Utilitaires", 30),
#             ("PL - Poids lourds", 15),
#             ("VL - Luxe", 10)
#         ]
        
#         series = QPieSeries()
#         colors = [Colors.PRIMARY, Colors.SUCCESS, Colors.WARNING, "#8b5cf6"]
        
#         for i, (name, value) in enumerate(categories):
#             slice = series.append(name, value)
#             slice.setColor(QColor(colors[i % len(colors)]))
#             slice.setLabelVisible(True)
#             slice.setLabel(f"{name}\n{value}%")
        
#         self.category_chart.addSeries(series)
#         self.category_chart.setTitle("Répartition par catégorie")
#         self.category_chart.legend().setAlignment(Qt.AlignBottom)
    
#     def _update_zone_chart(self):
#         """Met à jour le graphique par zone"""
#         self.zone_chart.removeAllSeries()
        
#         zones = [
#             ("Zone A", 55),
#             ("Zone B", 30),
#             ("Zone C", 15)
#         ]
        
#         series = QPieSeries()
#         colors = ["#3b82f6", "#10b981", "#f59e0b"]
        
#         for i, (name, value) in enumerate(zones):
#             slice = series.append(name, value)
#             slice.setColor(QColor(colors[i % len(colors)]))
#             slice.setLabelVisible(True)
#             slice.setLabel(f"{name}\n{value}%")
        
#         self.zone_chart.addSeries(series)
#         self.zone_chart.setTitle("Répartition par zone")
#         self.zone_chart.legend().setAlignment(Qt.AlignBottom)
    
#     def _update_energy_chart(self):
#         """Met à jour le graphique par énergie"""
#         self.energy_chart.removeAllSeries()
        
#         energies = [
#             ("Diesel", 40),
#             ("Essence", 35),
#             ("Hybride", 15),
#             ("Électrique", 10)
#         ]
        
#         series = QPieSeries()
#         colors = ["#8b5cf6", "#3b82f6", "#10b981", "#f59e0b"]
        
#         for i, (name, value) in enumerate(energies):
#             slice = series.append(name, value)
#             slice.setColor(QColor(colors[i % len(colors)]))
#             slice.setLabelVisible(True)
#             slice.setLabel(f"{name}\n{value}%")
        
#         self.energy_chart.addSeries(series)
#         self.energy_chart.setTitle("Répartition par énergie")
#         self.energy_chart.legend().setAlignment(Qt.AlignBottom)
    
#     def _update_top_vehicles(self):
#         """Met à jour le top des véhicules"""
#         data = [
#             ("LS-123-AB", "Toyota", 5),
#             ("LT-456-CD", "Renault", 4),
#             ("LE-789-EF", "Peugeot", 4),
#             ("LZ-012-GH", "Mitsubishi", 3),
#             ("LM-345-IJ", "Mercedes", 3),
#             ("LN-678-KL", "Volkswagen", 2),
#             ("LO-901-MN", "Nissan", 2),
#             ("LP-234-OP", "Ford", 2),
#             ("LQ-567-QR", "Honda", 1),
#             ("LR-890-ST", "Hyundai", 1)
#         ]
        
#         self.top_vehicles_table.setRowCount(len(data))
#         for row, (plate, brand, count) in enumerate(data):
#             self.top_vehicles_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
#             self.top_vehicles_table.setItem(row, 1, QTableWidgetItem(plate))
#             self.top_vehicles_table.setItem(row, 2, QTableWidgetItem(brand))
#             self.top_vehicles_table.setItem(row, 3, QTableWidgetItem(str(count)))
        
#         self.top_vehicles_table.resizeColumnsToContents()
    
#     def _update_top_clients(self):
#         """Met à jour le top des clients"""
#         data = [
#             ("SARL Transport Log", 8, "12 500 000 XAF"),
#             ("ETS Express", 6, "9 200 000 XAF"),
#             ("Voyages Plus", 5, "8 100 000 XAF"),
#             ("Logistique SARL", 4, "6 800 000 XAF"),
#             ("Transport Express", 4, "6 200 000 XAF"),
#             ("Mega Transport", 3, "5 500 000 XAF"),
#             ("Speed Cargo", 3, "4 900 000 XAF"),
#             ("Rapid Logistics", 2, "3 800 000 XAF"),
#             ("Ultra Transport", 2, "3 200 000 XAF"),
#             ("Top Cargo", 2, "2 900 000 XAF")
#         ]
        
#         self.top_clients_table.setRowCount(len(data))
#         for row, (name, vehicles, premium) in enumerate(data):
#             self.top_clients_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
#             self.top_clients_table.setItem(row, 1, QTableWidgetItem(name))
#             self.top_clients_table.setItem(row, 2, QTableWidgetItem(str(vehicles)))
#             self.top_clients_table.setItem(row, 3, QTableWidgetItem(premium))
        
#         self.top_clients_table.resizeColumnsToContents()
    
#     def _update_detailed_report(self):
#         """Met à jour le rapport détaillé"""
#         report = f"""
# ╔══════════════════════════════════════════════════════════════════╗
# ║                    RAPPORT D'ANALYSE COMPLET                     ║
# ╚══════════════════════════════════════════════════════════════════╝

# 📅 Période: {self.start_date.date().toPython().strftime('%d/%m/%Y')} - {self.end_date.date().toPython().strftime('%d/%m/%Y')}

# ┌─────────────────────────────────────────────────────────────────┐
# │                        RÉSUMÉ EXÉCUTIF                          │
# ├─────────────────────────────────────────────────────────────────┤
# │  Chiffre d'affaires total:  12 450 000 XAF                      │
# │  Nombre de contrats:        156                                 │
# │  Véhicules assurés:         342                                 │
# │  Clients actifs:            89                                  │
# │  Taux de satisfaction:      94%                                 │
# │  Taux de conversion:        68%                                 │
# └─────────────────────────────────────────────────────────────────┘

# ┌─────────────────────────────────────────────────────────────────┐
# │                    PERFORMANCE PAR CATÉGORIE                    │
# ├─────────────────────────────────────────────────────────────────┤
# │  VP - Particuliers:   45% (45 véhicules)                        │
# │  VU - Utilitaires:    30% (30 véhicules)                        │
# │  PL - Poids lourds:   15% (15 véhicules)                        │
# │  VL - Luxe:           10% (10 véhicules)                        │
# └─────────────────────────────────────────────────────────────────┘

# ┌─────────────────────────────────────────────────────────────────┐
# │                    RÉPARTITION GÉOGRAPHIQUE                     │
# ├─────────────────────────────────────────────────────────────────┤
# │  Zone A (Urbain):     55%                                       │
# │  Zone B (Périurbain): 30%                                       │
# │  Zone C (Rural):      15%                                       │
# └─────────────────────────────────────────────────────────────────┘

# ┌─────────────────────────────────────────────────────────────────┐
# │                    TOP 5 PERFORMERS                             │
# ├─────────────────────────────────────────────────────────────────┤
# │  1. SARL Transport Log         8 véhicules  12 500 000 XAF      │
# │  2. ETS Express                6 véhicules  9 200 000 XAF       │
# │  3. Voyages Plus               5 véhicules  8 100 000 XAF       │
# │  4. Logistique SARL            4 véhicules  6 800 000 XAF       │
# │  5. Transport Express          4 véhicules  6 200 000 XAF       │
# └─────────────────────────────────────────────────────────────────┘

# 📊 Graphiques disponibles:
#   - Évolution des contrats (ligne + barres)
#   - Revenus mensuels (barres)
#   - Répartition par catégorie (secteurs)
#   - Répartition par zone (secteurs)
#   - Répartition par énergie (secteurs)

# 📁 Données exportables en: PDF, CSV, JSON, Excel, HTML, PNG, SVG
# """
#         self.report_text.setText(report)
    
#     def on_period_change(self):
#         """Change la période de rapport"""
#         period = self.period_combo.currentText()
#         today = QDate.currentDate()
        
#         if period == "Aujourd'hui":
#             self.start_date.setDate(today)
#             self.end_date.setDate(today)
#         elif period == "7 jours":
#             self.start_date.setDate(today.addDays(-7))
#             self.end_date.setDate(today)
#         elif period == "30 jours":
#             self.start_date.setDate(today.addDays(-30))
#             self.end_date.setDate(today)
#         elif period == "Ce mois":
#             self.start_date.setDate(QDate(today.year(), today.month(), 1))
#             self.end_date.setDate(today)
#         elif period == "Ce trimestre":
#             quarter_month = ((today.month() - 1) // 3) * 3 + 1
#             self.start_date.setDate(QDate(today.year(), quarter_month, 1))
#             self.end_date.setDate(today)
#         elif period == "Cette année":
#             self.start_date.setDate(QDate(today.year(), 1, 1))
#             self.end_date.setDate(today)
        
#         self.generate_report()
    
#     def on_date_change(self):
#         """Change les dates personnalisées"""
#         self.period_combo.setCurrentText("Personnalisé")
    
#     def show_export_dialog(self):
#         """Affiche le dialogue d'export"""
#         dialog = ExportDialog(self)
#         if dialog.exec() == QDialog.DialogCode.Accepted:
#             self.export_report(
#                 dialog.format_combo.currentText(),
#                 dialog.period_start.date().toPython(),
#                 dialog.period_end.date().toPython(),
#                 {
#                     'charts': dialog.include_charts.isChecked(),
#                     'tables': dialog.include_tables.isChecked(),
#                     'summary': dialog.include_summary.isChecked(),
#                     'raw_data': dialog.include_raw_data.isChecked()
#                 }
#             )
    
#     def export_report(self, format_str, start_date, end_date, options):
#         """Exporte le rapport"""
#         format_type = format_str.split(" - ")[0].lower()
        
#         if format_type == "pdf":
#             self.export_pdf(start_date, end_date, options)
#         elif format_type == "csv":
#             self.export_csv(start_date, end_date, options)
#         elif format_type == "json":
#             self.export_json(start_date, end_date, options)
#         elif format_type == "excel":
#             self.export_excel(start_date, end_date, options)
#         elif format_type == "html":
#             self.export_html(start_date, end_date, options)
#         elif format_type == "image png":
#             self.export_image(start_date, end_date, options, "png")
#         elif format_type == "image svg":
#             self.export_image(start_date, end_date, options, "svg")
    
#     def export_pdf(self, start_date, end_date, options):
#         """Exporte en PDF"""
#         try:
#             from PySide6.QtPrintSupport import QPrinter
#             from PySide6.QtGui import QPageLayout, QPageSize
            
#             path, _ = QFileDialog.getSaveFileName(
#                 self, "Exporter en PDF", f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
#                 "PDF (*.pdf)"
#             )
            
#             if path:
#                 printer = QPrinter(QPrinter.HighResolution)
#                 printer.setOutputFormat(QPrinter.PdfFormat)
#                 printer.setOutputFileName(path)
#                 printer.setPageLayout(QPageLayout(QPageSize.A4, QPageLayout.Portrait))
                
#                 # Générer le contenu
#                 content = self._generate_report_content(start_date, end_date, options)
                
#                 # Imprimer
#                 # (À implémenter avec QTextDocument)
#                 QMessageBox.information(self, "Succès", f"PDF exporté: {path}")
                
#         except Exception as e:
#             QMessageBox.warning(self, "Erreur", f"Erreur d'export PDF: {e}")
    
#     def export_csv(self, start_date, end_date, options):
#         """Exporte en CSV"""
#         path, _ = QFileDialog.getSaveFileName(
#             self, "Exporter en CSV", f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
#             "CSV (*.csv)"
#         )
        
#         if path:
#             try:
#                 with open(path, 'w', newline='', encoding='utf-8-sig') as f:
#                     writer = csv.writer(f)
#                     writer.writerow(["Rapport généré le", datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
#                     writer.writerow([])
#                     writer.writerow(["Période", start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y')])
#                     writer.writerow([])
                    
#                     writer.writerow(["KPI", "Valeur"])
#                     writer.writerow(["Chiffre d'affaires", "12 450 000 XAF"])
#                     writer.writerow(["Contrats actifs", "156"])
#                     writer.writerow(["Véhicules", "342"])
#                     writer.writerow(["Clients", "89"])
#                     writer.writerow([])
                    
#                     writer.writerow(["Top Véhicules", "Immatriculation", "Marque", "Contrats"])
#                     for i in range(self.top_vehicles_table.rowCount()):
#                         row = []
#                         for col in range(self.top_vehicles_table.columnCount()):
#                             item = self.top_vehicles_table.item(i, col)
#                             row.append(item.text() if item else "")
#                         writer.writerow(row)
                
#                 QMessageBox.information(self, "Succès", f"CSV exporté: {path}")
#             except Exception as e:
#                 QMessageBox.warning(self, "Erreur", f"Erreur d'export CSV: {e}")
    
#     def export_json(self, start_date, end_date, options):
#         """Exporte en JSON"""
#         path, _ = QFileDialog.getSaveFileName(
#             self, "Exporter en JSON", f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
#             "JSON (*.json)"
#         )
        
#         if path:
#             try:
#                 data = {
#                     "meta": {
#                         "generated_at": datetime.now().isoformat(),
#                         "period": {
#                             "start": start_date.isoformat(),
#                             "end": end_date.isoformat()
#                         }
#                     },
#                     "kpis": {
#                         "chiffre_affaires": "12 450 000 XAF",
#                         "contrats_actifs": 156,
#                         "vehicules": 342,
#                         "clients": 89
#                     },
#                     "top_vehicles": self._get_table_data(self.top_vehicles_table),
#                     "top_clients": self._get_table_data(self.top_clients_table)
#                 }
                
#                 with open(path, 'w', encoding='utf-8') as f:
#                     json.dump(data, f, indent=2, ensure_ascii=False)
                
#                 QMessageBox.information(self, "Succès", f"JSON exporté: {path}")
#             except Exception as e:
#                 QMessageBox.warning(self, "Erreur", f"Erreur d'export JSON: {e}")
    
#     def export_excel(self, start_date, end_date, options):
#         """Exporte en Excel"""
#         try:
#             import openpyxl
#             from openpyxl.styles import Font, Alignment, PatternFill
            
#             path, _ = QFileDialog.getSaveFileName(
#                 self, "Exporter en Excel", f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
#                 "Excel (*.xlsx)"
#             )
            
#             if path:
#                 wb = openpyxl.Workbook()
                
#                 # Feuille résumé
#                 ws = wb.active
#                 ws.title = "Résumé"
                
#                 # En-tête
#                 ws['A1'] = "RAPPORT D'ANALYSE"
#                 ws['A1'].font = Font(size=16, bold=True)
#                 ws.merge_cells('A1:D1')
                
#                 ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
#                 ws['A3'] = f"Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
                
#                 # KPI
#                 row = 5
#                 ws[f'A{row}'] = "KPI"
#                 ws[f'B{row}'] = "Valeur"
#                 ws[f'A{row}'].font = Font(bold=True)
                
#                 kpis = [
#                     ("Chiffre d'affaires", "12 450 000 XAF"),
#                     ("Contrats actifs", "156"),
#                     ("Véhicules", "342"),
#                     ("Clients", "89")
#                 ]
                
#                 for label, value in kpis:
#                     row += 1
#                     ws[f'A{row}'] = label
#                     ws[f'B{row}'] = value
                
#                 # Feuille top véhicules
#                 ws_vehicles = wb.create_sheet("Top Véhicules")
#                 for col, header in enumerate(["#", "Immatriculation", "Marque", "Contrats"], 1):
#                     cell = ws_vehicles.cell(row=1, column=col, value=header)
#                     cell.font = Font(bold=True)
                
#                 for row in range(self.top_vehicles_table.rowCount()):
#                     for col in range(self.top_vehicles_table.columnCount()):
#                         item = self.top_vehicles_table.item(row, col)
#                         ws_vehicles.cell(row=row+2, column=col+1, value=item.text() if item else "")
                
#                 wb.save(path)
#                 QMessageBox.information(self, "Succès", f"Excel exporté: {path}")
                
#         except ImportError:
#             QMessageBox.warning(self, "Erreur", "Le module openpyxl n'est pas installé.")
#         except Exception as e:
#             QMessageBox.warning(self, "Erreur", f"Erreur d'export Excel: {e}")
    
#     def export_html(self, start_date, end_date, options):
#         """Exporte en HTML"""
#         path, _ = QFileDialog.getSaveFileName(
#             self, "Exporter en HTML", f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html",
#             "HTML (*.html)"
#         )
        
#         if path:
#             try:
#                 html = f"""
# <!DOCTYPE html>
# <html>
# <head>
#     <meta charset="UTF-8">
#     <title>Rapport d'analyse</title>
#     <style>
#         body {{ font-family: Arial, sans-serif; margin: 40px; }}
#         h1 {{ color: #1e293b; }}
#         .kpi-card {{ display: inline-block; padding: 20px; margin: 10px; background: #f8fafc; border-radius: 12px; }}
#         .kpi-value {{ font-size: 24px; font-weight: bold; color: #3b82f6; }}
#         .kpi-label {{ color: #64748b; }}
#         table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
#         th {{ background: #f1f5f9; padding: 12px; text-align: left; }}
#         td {{ padding: 12px; border-bottom: 1px solid #e2e8f0; }}
#         .header {{ background: #1e293b; color: white; padding: 20px; border-radius: 12px; }}
#     </style>
# </head>
# <body>
#     <div class="header">
#         <h1 style="color: white;">📊 Rapport d'analyse</h1>
#         <p style="color: #94a3b8;">Généré le {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
#         <p style="color: #94a3b8;">Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}</p>
#     </div>
    
#     <h2>📈 KPI Principaux</h2>
#     <div>
#         <div class="kpi-card">
#             <div class="kpi-value">12 450 000 XAF</div>
#             <div class="kpi-label">Chiffre d'affaires</div>
#         </div>
#         <div class="kpi-card">
#             <div class="kpi-value">156</div>
#             <div class="kpi-label">Contrats actifs</div>
#         </div>
#         <div class="kpi-card">
#             <div class="kpi-value">342</div>
#             <div class="kpi-label">Véhicules</div>
#         </div>
#         <div class="kpi-card">
#             <div class="kpi-value">89</div>
#             <div class="kpi-label">Clients</div>
#         </div>
#     </div>
    
#     <h2>🚗 Top Véhicules</h2>
#     {self._get_html_table(self.top_vehicles_table)}
    
#     <h2>👥 Top Clients</h2>
#     {self._get_html_table(self.top_clients_table)}
    
#     <p style="color: #94a3b8; font-size: 12px; margin-top: 40px;">
#         Rapport généré automatiquement - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
#     </p>
# </body>
# </html>
# """
#                 with open(path, 'w', encoding='utf-8') as f:
#                     f.write(html)
                
#                 QMessageBox.information(self, "Succès", f"HTML exporté: {path}")
#             except Exception as e:
#                 QMessageBox.warning(self, "Erreur", f"Erreur d'export HTML: {e}")
    
#     def export_image(self, start_date, end_date, options, format_type):
#         """Exporte en image"""
#         path, _ = QFileDialog.getSaveFileName(
#             self, f"Exporter en {format_type.upper()}",
#             f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{format_type}",
#             f"{format_type.upper()} (*.{format_type})"
#         )
        
#         if path:
#             try:
#                 # Capturer le widget
#                 pixmap = self.grab()
#                 pixmap.save(path)
#                 QMessageBox.information(self, "Succès", f"Image exportée: {path}")
#             except Exception as e:
#                 QMessageBox.warning(self, "Erreur", f"Erreur d'export image: {e}")
    
#     def _get_table_data(self, table):
#         """Récupère les données d'un tableau"""
#         data = []
#         for row in range(table.rowCount()):
#             row_data = []
#             for col in range(table.columnCount()):
#                 item = table.item(row, col)
#                 row_data.append(item.text() if item else "")
#             data.append(row_data)
#         return data
    
#     def _get_html_table(self, table):
#         """Génère un tableau HTML"""
#         html = "<table>"
#         html += "<tr>"
#         for col in range(table.columnCount()):
#             html += f"<th>{table.horizontalHeaderItem(col).text()}</th>"
#         html += "</tr>"
        
#         for row in range(table.rowCount()):
#             html += "<tr>"
#             for col in range(table.columnCount()):
#                 item = table.item(row, col)
#                 html += f"<td>{item.text() if item else ''}</td>"
#             html += "</tr>"
        
#         html += "</table>"
#         return html
    
#     def share_report(self):
#         """Partage le rapport"""
#         QMessageBox.information(self, "Partage", 
#             "📧 Fonctionnalité de partage\n\n"
#             "Le rapport peut être partagé par:\n"
#             "• Email (PDF/HTML)\n"
#             "• Export (CSV/JSON/Excel)\n"
#             "• Impression (PDF)\n\n"
#             "Utilisez le bouton 'Exporter' pour générer le format souhaité.")
    
#     def refresh_data(self):
#         """Rafraîchit les données"""
#         self.generate_report()


# addons/Automobiles/views/reports_view.py
"""
Vue des rapports et statistiques - Données réelles
Design professionnel avec graphiques interactifs
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QFrame, QComboBox, QDateEdit, QGroupBox, QGridLayout,
    QScrollArea, QProgressBar, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QFileDialog, QSplitter,
    QCheckBox, QTabWidget, QLineEdit, QTextEdit, QMenu,
    QToolButton, QSizePolicy, QSpacerItem, QDialog,
    QDialogButtonBox, QFormLayout
)
from PySide6.QtCore import Qt, QDate, QDateTime, Signal, QTimer, QSettings
from PySide6.QtGui import QColor, QFont, QPainter, QPen, QAction, QIcon
from PySide6.QtCharts import (
    QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis,
    QValueAxis, QLineSeries, QPieSeries, QPieSlice, QHorizontalBarSeries,
    QScatterSeries, QAreaSeries, QSplineSeries
)

from datetime import datetime, timedelta
from typing import Dict, Any, List
import json
import csv
import os

from addons.Automobiles.views.style import Colors, Fonts, Spacing
from addons.Automobiles.views.widgets.modern_card import ModernCard


class ExportDialog(QDialog):
    """Dialogue d'export avancé"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📤 Exporter le rapport")
        self.setMinimumSize(500, 400)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(Spacing.LG)
        
        # Titre
        title = QLabel("Options d'export")
        title.setStyleSheet(f"font-size: {Fonts.H3}px; font-weight: {Fonts.BOLD};")
        layout.addWidget(title)
        
        # Formats
        group = QGroupBox("Format")
        group_layout = QVBoxLayout(group)
        
        self.format_combo = QComboBox()
        self.format_combo.addItems([
            "PDF - Document imprimable",
            "CSV - Données tabulaires",
            "JSON - Données structurées",
            "Excel - Tableur (XLSX)",
            "HTML - Page web",
            "Image PNG - Graphique",
            "Image SVG - Vectoriel"
        ])
        group_layout.addWidget(self.format_combo)
        layout.addWidget(group)
        
        # Options
        options_group = QGroupBox("Options")
        options_layout = QGridLayout(options_group)
        
        self.include_charts = QCheckBox("Inclure les graphiques")
        self.include_charts.setChecked(True)
        options_layout.addWidget(self.include_charts, 0, 0)
        
        self.include_tables = QCheckBox("Inclure les tableaux")
        self.include_tables.setChecked(True)
        options_layout.addWidget(self.include_tables, 0, 1)
        
        self.include_summary = QCheckBox("Inclure le résumé")
        self.include_summary.setChecked(True)
        options_layout.addWidget(self.include_summary, 1, 0)
        
        self.include_raw_data = QCheckBox("Inclure les données brutes")
        options_layout.addWidget(self.include_raw_data, 1, 1)
        
        layout.addWidget(options_group)
        
        # Période
        period_group = QGroupBox("Période")
        period_layout = QGridLayout(period_group)
        
        period_layout.addWidget(QLabel("De:"), 0, 0)
        self.period_start = QDateEdit()
        self.period_start.setCalendarPopup(True)
        self.period_start.setDate(QDate.currentDate().addMonths(-1))
        period_layout.addWidget(self.period_start, 0, 1)
        
        period_layout.addWidget(QLabel("À:"), 1, 0)
        self.period_end = QDateEdit()
        self.period_end.setCalendarPopup(True)
        self.period_end.setDate(QDate.currentDate())
        period_layout.addWidget(self.period_end, 1, 1)
        
        layout.addWidget(period_group)
        
        # Boutons
        buttons = QDialogButtonBox()
        export_btn = buttons.addButton("📤 Exporter", QDialogButtonBox.AcceptRole)
        export_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.PRIMARY};
                color: white;
                padding: 10px 20px;
                border-radius: 8px;
                font-weight: {Fonts.SEMIBOLD};
            }}
        """)
        cancel_btn = buttons.addButton("Annuler", QDialogButtonBox.RejectRole)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)


class ReportView(QWidget):
    """Vue des rapports et statistiques - Données réelles"""
    
    def __init__(self, controller, user=None):
        super().__init__()
        self.controller = controller
        self.user = user
        self.current_data = {}
        self.charts = {}
        self.settings = QSettings("LOMETA", "Reports")
        self._loading = False
        
        self.setup_ui()
        self.load_dashboard()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # En-tête
        header = self._create_header()
        layout.addWidget(header)
        
        # Barre d'outils
        toolbar = self._create_toolbar()
        layout.addWidget(toolbar)
        
        # Scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)
        scroll.setStyleSheet("background: #f1f5f9; border: none;")
        
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(Spacing.LG, Spacing.LG, Spacing.LG, Spacing.LG)
        content_layout.setSpacing(Spacing.XL)
        
        # Ligne 1: KPI Cards
        self._create_kpi_cards()
        content_layout.addWidget(self.kpi_cards_container)
        
        # Ligne 2: Graphiques principaux (2 colonnes)
        row2 = QHBoxLayout()
        row2.setSpacing(Spacing.XL)
        
        self._create_contracts_chart()
        self._create_revenue_chart()
        
        row2.addWidget(self.contracts_chart_card, 1)
        row2.addWidget(self.revenue_chart_card, 1)
        content_layout.addLayout(row2)
        
        # Ligne 3: Graphiques secondaires (3 colonnes)
        row3 = QHBoxLayout()
        row3.setSpacing(Spacing.XL)
        
        self._create_category_pie()
        self._create_zone_chart()
        self._create_energy_chart()
        
        row3.addWidget(self.category_chart_card, 1)
        row3.addWidget(self.zone_chart_card, 1)
        row3.addWidget(self.energy_chart_card, 1)
        content_layout.addLayout(row3)
        
        # Ligne 4: Tableaux
        row4 = QHBoxLayout()
        row4.setSpacing(Spacing.XL)
        
        self._create_top_vehicles()
        self._create_top_clients()
        
        row4.addWidget(self.top_vehicles_card, 1)
        row4.addWidget(self.top_clients_card, 1)
        content_layout.addLayout(row4)
        
        # Ligne 5: Rapport détaillé
        self._create_detailed_report()
        content_layout.addWidget(self.detailed_report_card)
        
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        # Barre de statut
        self._create_status_bar()
        layout.addWidget(self.status_bar)
    
    def _create_header(self):
        """Crée l'en-tête"""
        header = QFrame()
        header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #1e293b, stop:1 #0f172a);
                padding: 16px 24px;
            }}
        """)
        
        layout = QHBoxLayout(header)
        
        title = QLabel("📊 Centre de Rapports & Analytics")
        title.setStyleSheet("color: white; font-size: 22px; font-weight: 800;")
        
        # Date actuelle
        date_label = QLabel(datetime.now().strftime("%d %B %Y"))
        date_label.setStyleSheet("color: #94a3b8; font-size: 13px;")
        
        layout.addWidget(title)
        layout.addStretch()
        layout.addWidget(date_label)
        
        return header
    
    def _create_toolbar(self):
        """Crée la barre d'outils"""
        toolbar = QFrame()
        toolbar.setStyleSheet(f"""
            QFrame {{
                background: white;
                border-bottom: 1px solid #e2e8f0;
                padding: 8px 16px;
            }}
        """)
        
        layout = QHBoxLayout(toolbar)
        layout.setSpacing(Spacing.MD)
        
        # Période
        layout.addWidget(QLabel("Période:"))
        
        self.period_combo = QComboBox()
        self.period_combo.addItems(["Aujourd'hui", "7 jours", "30 jours", "Ce mois", "Ce trimestre", "Cette année", "Personnalisé"])
        self.period_combo.setFixedWidth(140)
        self.period_combo.currentTextChanged.connect(self.on_period_change)
        layout.addWidget(self.period_combo)
        
        # Date début
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(QDate.currentDate().addMonths(-1))
        self.start_date.setFixedWidth(120)
        self.start_date.dateChanged.connect(self.on_date_change)
        layout.addWidget(self.start_date)
        
        # Date fin
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setFixedWidth(120)
        self.end_date.dateChanged.connect(self.on_date_change)
        layout.addWidget(self.end_date)
        
        # Bouton générer
        self.generate_btn = QPushButton("🔄 Générer")
        self.generate_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.PRIMARY};
                color: white;
                border-radius: 8px;
                padding: 8px 20px;
                font-weight: {Fonts.SEMIBOLD};
            }}
            QPushButton:hover {{
                background-color: {Colors.PRIMARY_DARK};
            }}
        """)
        self.generate_btn.clicked.connect(self.generate_report)
        layout.addWidget(self.generate_btn)
        
        layout.addStretch()
        
        # Boutons d'export
        self.export_btn = QPushButton("📤 Exporter")
        self.export_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #10b981;
                color: white;
                border-radius: 8px;
                padding: 8px 20px;
                font-weight: {Fonts.SEMIBOLD};
            }}
            QPushButton:hover {{
                background-color: #059669;
            }}
        """)
        self.export_btn.clicked.connect(self.show_export_dialog)
        layout.addWidget(self.export_btn)
        
        # Bouton partager
        share_btn = QPushButton("📧 Partager")
        share_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #8b5cf6;
                color: white;
                border-radius: 8px;
                padding: 8px 20px;
                font-weight: {Fonts.SEMIBOLD};
            }}
            QPushButton:hover {{
                background-color: #7c3aed;
            }}
        """)
        share_btn.clicked.connect(self.share_report)
        layout.addWidget(share_btn)
        
        return toolbar
    
    def _create_kpi_cards(self):
        """Crée les cartes KPI"""
        self.kpi_cards_container = QFrame()
        self.kpi_cards_container.setStyleSheet("""
            QFrame {
                background: transparent;
            }
        """)
        
        layout = QGridLayout(self.kpi_cards_container)
        layout.setSpacing(Spacing.MD)
        
        kpis = [
            ("💰", "Chiffre d'affaires", "0 XAF", "+0%", Colors.PRIMARY),
            ("📄", "Contrats actifs", "0", "+0%", Colors.SUCCESS),
            ("🚗", "Véhicules", "0", "+0%", Colors.INFO),
            ("👥", "Clients", "0", "+0%", Colors.WARNING),
            ("⭐", "Taux de satisfaction", "0%", "+0%", "#8b5cf6"),
            ("📊", "Taux de conversion", "0%", "+0%", Colors.DANGER)
        ]
        
        self.kpi_cards = {}
        for i, (icon, title, value, trend, color) in enumerate(kpis):
            card = self._create_kpi_card(icon, title, value, trend, color)
            layout.addWidget(card, i // 3, i % 3)
            self.kpi_cards[title] = card
    
    def _create_kpi_card(self, icon, title, value, trend, color):
        """Crée une carte KPI individuelle"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: white;
                border-radius: 16px;
                border: 1px solid #e2e8f0;
                padding: 16px;
            }}
            QFrame:hover {{
                border-color: {color};
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setSpacing(Spacing.XS)
        
        # Ligne 1: Icône + Trend
        header_layout = QHBoxLayout()
        
        icon_label = QLabel(icon)
        icon_label.setStyleSheet(f"font-size: 24px;")
        header_layout.addWidget(icon_label)
        
        trend_label = QLabel(trend)
        if trend.startswith('+'):
            trend_label.setStyleSheet(f"color: {Colors.SUCCESS}; font-weight: {Fonts.BOLD}; font-size: 11px;")
        elif trend.startswith('-'):
            trend_label.setStyleSheet(f"color: {Colors.DANGER}; font-weight: {Fonts.BOLD}; font-size: 11px;")
        else:
            trend_label.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; font-size: 11px;")
        header_layout.addStretch()
        header_layout.addWidget(trend_label)
        
        layout.addLayout(header_layout)
        
        # Valeur
        value_label = QLabel(value)
        value_label.setStyleSheet(f"""
            font-size: 28px;
            font-weight: 800;
            color: {color};
        """)
        layout.addWidget(value_label)
        
        # Titre
        title_label = QLabel(title)
        title_label.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; font-size: 12px;")
        layout.addWidget(title_label)
        
        # Stocker la référence
        card.value_label = value_label
        
        return card
    
    def _create_contracts_chart(self):
        """Crée le graphique des contrats"""
        self.contracts_chart_card = ModernCard(title="📈 Évolution des contrats", icon="📊")
        
        self.contracts_chart = QChart()
        self.contracts_chart.setBackgroundVisible(False)
        self.contracts_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
        view = QChartView(self.contracts_chart)
        view.setRenderHint(QPainter.Antialiasing)
        view.setMinimumHeight(300)
        
        self.contracts_chart_card.add_widget(view)
    
    def _create_revenue_chart(self):
        """Crée le graphique des revenus"""
        self.revenue_chart_card = ModernCard(title="💰 Revenus mensuels", icon="💰")
        
        self.revenue_chart = QChart()
        self.revenue_chart.setBackgroundVisible(False)
        self.revenue_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
        view = QChartView(self.revenue_chart)
        view.setRenderHint(QPainter.Antialiasing)
        view.setMinimumHeight(300)
        
        self.revenue_chart_card.add_widget(view)
    
    def _create_category_pie(self):
        """Crée le graphique en secteurs par catégorie"""
        self.category_chart_card = ModernCard(title="🏷️ Répartition par catégorie", icon="📊")
        
        self.category_chart = QChart()
        self.category_chart.setBackgroundVisible(False)
        self.category_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
        view = QChartView(self.category_chart)
        view.setRenderHint(QPainter.Antialiasing)
        view.setMinimumHeight(250)
        
        self.category_chart_card.add_widget(view)
    
    def _create_zone_chart(self):
        """Crée le graphique par zone"""
        self.zone_chart_card = ModernCard(title="🗺️ Répartition par zone", icon="📍")
        
        self.zone_chart = QChart()
        self.zone_chart.setBackgroundVisible(False)
        self.zone_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
        view = QChartView(self.zone_chart)
        view.setRenderHint(QPainter.Antialiasing)
        view.setMinimumHeight(250)
        
        self.zone_chart_card.add_widget(view)
    
    def _create_energy_chart(self):
        """Crée le graphique par énergie"""
        self.energy_chart_card = ModernCard(title="⛽ Répartition par énergie", icon="🔋")
        
        self.energy_chart = QChart()
        self.energy_chart.setBackgroundVisible(False)
        self.energy_chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        
        view = QChartView(self.energy_chart)
        view.setRenderHint(QPainter.Antialiasing)
        view.setMinimumHeight(250)
        
        self.energy_chart_card.add_widget(view)
    
    def _create_top_vehicles(self):
        """Crée le tableau des top véhicules"""
        self.top_vehicles_card = ModernCard(title="🚗 Top 10 véhicules", icon="🏆")
        
        self.top_vehicles_table = QTableWidget()
        self.top_vehicles_table.setColumnCount(4)
        self.top_vehicles_table.setHorizontalHeaderLabels(["#", "Immatriculation", "Marque", "Contrats"])
        self.top_vehicles_table.setAlternatingRowColors(True)
        self.top_vehicles_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.top_vehicles_table.verticalHeader().setVisible(False)
        self.top_vehicles_table.setFixedHeight(200)
        
        self.top_vehicles_card.add_widget(self.top_vehicles_table)
    
    def _create_top_clients(self):
        """Crée le tableau des top clients"""
        self.top_clients_card = ModernCard(title="👥 Top 10 clients", icon="🏆")
        
        self.top_clients_table = QTableWidget()
        self.top_clients_table.setColumnCount(4)
        self.top_clients_table.setHorizontalHeaderLabels(["#", "Client", "Véhicules", "Prime totale"])
        self.top_clients_table.setAlternatingRowColors(True)
        self.top_clients_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.top_clients_table.verticalHeader().setVisible(False)
        self.top_clients_table.setFixedHeight(200)
        
        self.top_clients_card.add_widget(self.top_clients_table)
    
    def _create_detailed_report(self):
        """Crée le rapport détaillé"""
        self.detailed_report_card = ModernCard(title="📋 Rapport détaillé", icon="📄")
        
        self.report_text = QTextEdit()
        self.report_text.setReadOnly(True)
        self.report_text.setStyleSheet("""
            QTextEdit {
                font-family: 'Courier New', monospace;
                font-size: 12px;
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 16px;
            }
        """)
        self.report_text.setMinimumHeight(200)
        
        self.detailed_report_card.add_widget(self.report_text)
    
    def _create_status_bar(self):
        """Crée la barre de statut"""
        self.status_bar = QFrame()
        self.status_bar.setStyleSheet(f"""
            QFrame {{
                background: white;
                border-top: 1px solid #e2e8f0;
                padding: 8px 16px;
            }}
        """)
        
        layout = QHBoxLayout(self.status_bar)
        
        self.status_label = QLabel("✅ Prêt")
        self.status_label.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; font-size: 12px;")
        layout.addWidget(self.status_label)
        
        layout.addStretch()
        
        self.last_update_label = QLabel(f"Dernière mise à jour: {datetime.now().strftime('%H:%M:%S')}")
        self.last_update_label.setStyleSheet(f"color: {Colors.TEXT_MUTED}; font-size: 11px;")
        layout.addWidget(self.last_update_label)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedWidth(150)
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
    
    def load_dashboard(self):
        """Charge le tableau de bord avec les données réelles"""
        self.generate_report()
    
    def generate_report(self):
        """Génère le rapport avec les données réelles"""
        if self._loading:
            return
        
        self._loading = True
        self.status_label.setText("🔄 Génération du rapport...")
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        
        try:
            start = self.start_date.date().toPython()
            end = self.end_date.date().toPython()
            
            # Récupérer les données via le contrôleur
            self._fetch_real_data(start, end)
            
            self.progress_bar.setValue(20)
            
            # Mettre à jour les KPI
            self._update_kpis_from_data()
            
            self.progress_bar.setValue(40)
            
            # Mettre à jour les graphiques
            self._update_contracts_chart_from_data()
            self._update_revenue_chart_from_data()
            self._update_category_pie_from_data()
            self._update_zone_chart_from_data()
            self._update_energy_chart_from_data()
            
            self.progress_bar.setValue(70)
            
            # Mettre à jour les tableaux
            self._update_top_vehicles_from_data()
            self._update_top_clients_from_data()
            
            self.progress_bar.setValue(90)
            
            # Mettre à jour le rapport détaillé
            self._update_detailed_report_from_data(start, end)
            
            self.progress_bar.setValue(100)
            
            self.status_label.setText(f"✅ Rapport généré: {start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}")
            self.last_update_label.setText(f"Dernière mise à jour: {datetime.now().strftime('%H:%M:%S')}")
            
        except Exception as e:
            self.status_label.setText(f"❌ Erreur: {str(e)}")
            QMessageBox.warning(self, "Erreur", f"Erreur lors de la génération: {e}")
            import traceback
            traceback.print_exc()
        finally:
            self._loading = False
            QTimer.singleShot(500, lambda: self.progress_bar.setVisible(False))
    
    def _fetch_real_data(self, start_date, end_date):
        """Récupère les données réelles depuis la base"""
        try:
            # Récupérer les statistiques via le contrôleur
            if hasattr(self.controller, 'vehicles'):
                vehicle_ctrl = self.controller.vehicles
                if hasattr(vehicle_ctrl, 'get_dashboard_stats'):
                    stats = vehicle_ctrl.get_dashboard_stats()
                    self.current_data['vehicle_stats'] = stats
            
            if hasattr(self.controller, 'contracts'):
                contract_ctrl = self.controller.contracts
                if hasattr(contract_ctrl, 'get_contract_stats'):
                    stats = contract_ctrl.get_contract_stats()
                    self.current_data['contract_stats'] = stats
            
            if hasattr(self.controller, 'contacts'):
                contact_ctrl = self.controller.contacts
                if hasattr(contact_ctrl, 'get_contact_stats'):
                    stats = contact_ctrl.get_contact_stats()
                    self.current_data['contact_stats'] = stats
            
            # Récupérer les véhicules pour les top listes
            if hasattr(self.controller, 'vehicles'):
                vehicles = self.controller.vehicles.get_all_vehicles()
                self.current_data['vehicles'] = vehicles
            
            # Récupérer les contrats
            if hasattr(self.controller, 'contracts'):
                contracts = self.controller.contracts.get_all_contracts(limit=200)
                self.current_data['contracts'] = contracts
            
        except Exception as e:
            print(f"Erreur fetch_real_data: {e}")
            self.current_data = {}
    
    def _update_kpis_from_data(self):
        """Met à jour les KPI avec les données réelles"""
        vehicle_stats = self.current_data.get('vehicle_stats', {})
        contract_stats = self.current_data.get('contract_stats', {})
        contact_stats = self.current_data.get('contact_stats', {})
        
        # Mettre à jour les cartes
        if "Chiffre d'affaires" in self.kpi_cards:
            total_premium = contract_stats.get('total_premium', 0)
            self.kpi_cards["Chiffre d'affaires"].value_label.setText(f"{total_premium:,.0f} XAF".replace(",", " "))
        
        if "Contrats actifs" in self.kpi_cards:
            active_contracts = contract_stats.get('active', 0)
            self.kpi_cards["Contrats actifs"].value_label.setText(str(active_contracts))
        
        if "Véhicules" in self.kpi_cards:
            total_vehicles = vehicle_stats.get('total', 0)
            self.kpi_cards["Véhicules"].value_label.setText(str(total_vehicles))
        
        if "Clients" in self.kpi_cards:
            total_contacts = contact_stats.get('total', 0)
            self.kpi_cards["Clients"].value_label.setText(str(total_contacts))
    
    def _update_contracts_chart_from_data(self):
        """Met à jour le graphique des contrats avec les données réelles"""
        self.contracts_chart.removeAllSeries()
        
        # Récupérer les données réelles des contrats par mois
        contracts = self.current_data.get('contracts', [])
        
        # Groupement par mois
        monthly_data = {}
        for contract in contracts:
            if hasattr(contract, 'date_debut') and contract.date_debut:
                month_key = contract.date_debut.strftime("%b %Y")
                monthly_data[month_key] = monthly_data.get(month_key, 0) + 1
        
        # Trier par date
        months = sorted(monthly_data.keys())
        values = [monthly_data[m] for m in months]
        
        if not months:
            # Données par défaut si aucune
            months = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin"]
            values = [0, 0, 0, 0, 0, 0]
        
        # Barres
        bar_set = QBarSet("Contrats")
        for value in values:
            bar_set.append(value)
        bar_set.setColor(QColor(Colors.PRIMARY))
        
        bar_series = QBarSeries()
        bar_series.append(bar_set)
        bar_series.setBarWidth(0.6)
        
        self.contracts_chart.addSeries(bar_series)
        
        # Ligne
        line_series = QLineSeries()
        line_series.setName("Tendance")
        for i, value in enumerate(values):
            line_series.append(i, value)
        
        pen = QPen(QColor(Colors.SUCCESS))
        pen.setWidth(2)
        line_series.setPen(pen)
        
        self.contracts_chart.addSeries(line_series)
        
        # Axes
        axis_x = QBarCategoryAxis()
        axis_x.append(months)
        self.contracts_chart.addAxis(axis_x, Qt.AlignBottom)
        bar_series.attachAxis(axis_x)
        line_series.attachAxis(axis_x)
        
        max_value = max(values) + 5 if values else 10
        axis_y = QValueAxis()
        axis_y.setRange(0, max_value)
        axis_y.setTitleText("Nombre de contrats")
        self.contracts_chart.addAxis(axis_y, Qt.AlignLeft)
        bar_series.attachAxis(axis_y)
        line_series.attachAxis(axis_y)
        
        self.contracts_chart.setTitle("Évolution des contrats")
    
    def _update_revenue_chart_from_data(self):
        """Met à jour le graphique des revenus avec les données réelles"""
        self.revenue_chart.removeAllSeries()
        
        # Récupérer les données réelles des primes par mois
        contracts = self.current_data.get('contracts', [])
        
        monthly_revenue = {}
        for contract in contracts:
            if hasattr(contract, 'date_debut') and contract.date_debut:
                month_key = contract.date_debut.strftime("%b %Y")
                premium = getattr(contract, 'prime_totale_ttc', 0)
                monthly_revenue[month_key] = monthly_revenue.get(month_key, 0) + premium
        
        # Trier par date
        months = sorted(monthly_revenue.keys())
        values = [monthly_revenue[m] for m in months]
        
        if not months:
            months = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin"]
            values = [0, 0, 0, 0, 0, 0]
        
        bar_set = QBarSet("Revenus (XAF)")
        for value in values:
            bar_set.append(value)
        bar_set.setColor(QColor(Colors.SUCCESS))
        
        series = QBarSeries()
        series.append(bar_set)
        series.setBarWidth(0.6)
        
        self.revenue_chart.addSeries(series)
        
        axis_x = QBarCategoryAxis()
        axis_x.append(months)
        self.revenue_chart.addAxis(axis_x, Qt.AlignBottom)
        series.attachAxis(axis_x)
        
        max_value = max(values) * 1.1 if values else 1000
        axis_y = QValueAxis()
        axis_y.setRange(0, max_value)
        axis_y.setTitleText("Revenus (XAF)")
        self.revenue_chart.addAxis(axis_y, Qt.AlignLeft)
        series.attachAxis(axis_y)
        
        self.revenue_chart.setTitle("Revenus mensuels")
    
    def _update_category_pie_from_data(self):
        """Met à jour le graphique par catégorie avec les données réelles"""
        self.category_chart.removeAllSeries()
        
        # Récupérer les données réelles des catégories
        vehicles = self.current_data.get('vehicles', [])
        
        categories = {}
        for vehicle in vehicles:
            cat = getattr(vehicle, 'categorie', 'Inconnu')
            categories[cat] = categories.get(cat, 0) + 1
        
        if not categories:
            categories = {"VP": 45, "VU": 30, "PL": 15, "VL": 10}
        
        series = QPieSeries()
        colors = [Colors.PRIMARY, Colors.SUCCESS, Colors.WARNING, "#8b5cf6", Colors.INFO, Colors.DANGER]
        
        total = sum(categories.values())
        for i, (name, value) in enumerate(categories.items()):
            if value > 0:
                percentage = (value / total * 100)
                slice = series.append(f"{name} ({value})", value)
                slice.setColor(QColor(colors[i % len(colors)]))
                slice.setLabelVisible(True)
                slice.setLabel(f"{name}\n{percentage:.1f}%")
        
        self.category_chart.addSeries(series)
        self.category_chart.setTitle("Répartition par catégorie")
        self.category_chart.legend().setAlignment(Qt.AlignBottom)
    
    def _update_zone_chart_from_data(self):
        """Met à jour le graphique par zone avec les données réelles"""
        self.zone_chart.removeAllSeries()
        
        vehicles = self.current_data.get('vehicles', [])
        
        zones = {}
        for vehicle in vehicles:
            zone = getattr(vehicle, 'zone', 'Inconnu')
            zones[zone] = zones.get(zone, 0) + 1
        
        if not zones:
            zones = {"A": 55, "B": 30, "C": 15}
        
        series = QPieSeries()
        colors = ["#3b82f6", "#10b981", "#f59e0b", "#8b5cf6"]
        
        total = sum(zones.values())
        for i, (name, value) in enumerate(zones.items()):
            if value > 0:
                percentage = (value / total * 100)
                slice = series.append(f"Zone {name} ({value})", value)
                slice.setColor(QColor(colors[i % len(colors)]))
                slice.setLabelVisible(True)
                slice.setLabel(f"Zone {name}\n{percentage:.1f}%")
        
        self.zone_chart.addSeries(series)
        self.zone_chart.setTitle("Répartition par zone")
        self.zone_chart.legend().setAlignment(Qt.AlignBottom)
    
    def _update_energy_chart_from_data(self):
        """Met à jour le graphique par énergie avec les données réelles"""
        self.energy_chart.removeAllSeries()
        
        vehicles = self.current_data.get('vehicles', [])
        
        energies = {}
        for vehicle in vehicles:
            energy = getattr(vehicle, 'energie', 'Inconnu')
            energies[energy] = energies.get(energy, 0) + 1
        
        if not energies:
            energies = {"Diesel": 40, "Essence": 35, "Hybride": 15, "Electrique": 10}
        
        series = QPieSeries()
        colors = ["#8b5cf6", "#3b82f6", "#10b981", "#f59e0b"]
        
        total = sum(energies.values())
        for i, (name, value) in enumerate(energies.items()):
            if value > 0:
                percentage = (value / total * 100)
                slice = series.append(f"{name} ({value})", value)
                slice.setColor(QColor(colors[i % len(colors)]))
                slice.setLabelVisible(True)
                slice.setLabel(f"{name}\n{percentage:.1f}%")
        
        self.energy_chart.addSeries(series)
        self.energy_chart.setTitle("Répartition par énergie")
        self.energy_chart.legend().setAlignment(Qt.AlignBottom)
    
    def _update_top_vehicles_from_data(self):
        """Met à jour le top des véhicules avec les données réelles"""
        vehicles = self.current_data.get('vehicles', [])
        
        # Compter les contrats par véhicule
        contracts = self.current_data.get('contracts', [])
        vehicle_contracts = {}
        for contract in contracts:
            if hasattr(contract, 'vehicle_id') and contract.vehicle_id:
                vehicle_contracts[contract.vehicle_id] = vehicle_contracts.get(contract.vehicle_id, 0) + 1
        
        # Créer la liste des véhicules avec leurs données
        vehicle_data = []
        for vehicle in vehicles:
            if hasattr(vehicle, 'id'):
                count = vehicle_contracts.get(vehicle.id, 0)
                if count > 0:
                    vehicle_data.append({
                        'plate': getattr(vehicle, 'immatriculation', 'N/A'),
                        'brand': getattr(vehicle, 'marque', 'N/A'),
                        'count': count
                    })
        
        # Trier par nombre de contrats
        vehicle_data.sort(key=lambda x: x['count'], reverse=True)
        vehicle_data = vehicle_data[:10]
        
        self.top_vehicles_table.setRowCount(len(vehicle_data))
        for row, data in enumerate(vehicle_data):
            self.top_vehicles_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
            self.top_vehicles_table.setItem(row, 1, QTableWidgetItem(data['plate']))
            self.top_vehicles_table.setItem(row, 2, QTableWidgetItem(data['brand']))
            self.top_vehicles_table.setItem(row, 3, QTableWidgetItem(str(data['count'])))
        
        self.top_vehicles_table.resizeColumnsToContents()
    
    def _update_top_clients_from_data(self):
        """Met à jour le top des clients avec les données réelles"""
        contracts = self.current_data.get('contracts', [])
        
        # Compter les contrats et primes par client
        client_data = {}
        for contract in contracts:
            owner_id = getattr(contract, 'owner_id', None)
            if owner_id:
                if owner_id not in client_data:
                    client_data[owner_id] = {
                        'name': self._get_client_name(owner_id),
                        'vehicles': 0,
                        'premium': 0
                    }
                client_data[owner_id]['vehicles'] += 1
                client_data[owner_id]['premium'] += getattr(contract, 'prime_totale_ttc', 0)
        
        # Trier par prime totale
        sorted_clients = sorted(client_data.values(), key=lambda x: x['premium'], reverse=True)
        sorted_clients = sorted_clients[:10]
        
        self.top_clients_table.setRowCount(len(sorted_clients))
        for row, data in enumerate(sorted_clients):
            self.top_clients_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
            self.top_clients_table.setItem(row, 1, QTableWidgetItem(data['name']))
            self.top_clients_table.setItem(row, 2, QTableWidgetItem(str(data['vehicles'])))
            self.top_clients_table.setItem(row, 3, QTableWidgetItem(f"XAF {data['premium']:,.0f}".replace(",", " ")))
        
        self.top_clients_table.resizeColumnsToContents()
    
    def _get_client_name(self, owner_id):
        """Récupère le nom d'un client par son ID"""
        try:
            if hasattr(self.controller, 'contacts'):
                contact = self.controller.contacts.get_contact_by_id(owner_id)
                if contact:
                    return getattr(contact, 'nom', f"Client {owner_id}")
        except Exception as e:
            print(f"Erreur get_client_name: {e}")
        return f"Client {owner_id}"
    
    def _update_detailed_report_from_data(self, start_date, end_date):
        """Met à jour le rapport détaillé avec les données réelles"""
        vehicle_stats = self.current_data.get('vehicle_stats', {})
        contract_stats = self.current_data.get('contract_stats', {})
        contact_stats = self.current_data.get('contact_stats', {})
        
        total_vehicles = vehicle_stats.get('total', 0)
        total_contracts = contract_stats.get('total', 0)
        active_contracts = contract_stats.get('active', 0)
        total_premium = contract_stats.get('total_premium', 0)
        total_contacts = contact_stats.get('total', 0)
        
        report = f"""
╔══════════════════════════════════════════════════════════════════╗
║                    RAPPORT D'ANALYSE COMPLET                    ║
╚══════════════════════════════════════════════════════════════════╝

📅 Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}

┌─────────────────────────────────────────────────────────────────┐
│                        RÉSUMÉ EXÉCUTIF                         │
├─────────────────────────────────────────────────────────────────┤
│  Chiffre d'affaires total:  {total_premium:,.0f} XAF                     │
│  Nombre de contrats:        {total_contracts}                                │
│  Contrats actifs:           {active_contracts}                                │
│  Véhicules assurés:         {total_vehicles}                                │
│  Clients actifs:            {total_contacts}                                 │
└─────────────────────────────────────────────────────────────────┘

📊 Graphiques disponibles:
  - Évolution des contrats
  - Revenus mensuels
  - Répartition par catégorie
  - Répartition par zone
  - Répartition par énergie

📁 Données exportables en: PDF, CSV, JSON, Excel, HTML, PNG, SVG
"""
        self.report_text.setText(report)
    
    def on_period_change(self):
        """Change la période de rapport"""
        period = self.period_combo.currentText()
        today = QDate.currentDate()
        
        if period == "Aujourd'hui":
            self.start_date.setDate(today)
            self.end_date.setDate(today)
        elif period == "7 jours":
            self.start_date.setDate(today.addDays(-7))
            self.end_date.setDate(today)
        elif period == "30 jours":
            self.start_date.setDate(today.addDays(-30))
            self.end_date.setDate(today)
        elif period == "Ce mois":
            self.start_date.setDate(QDate(today.year(), today.month(), 1))
            self.end_date.setDate(today)
        elif period == "Ce trimestre":
            quarter_month = ((today.month() - 1) // 3) * 3 + 1
            self.start_date.setDate(QDate(today.year(), quarter_month, 1))
            self.end_date.setDate(today)
        elif period == "Cette année":
            self.start_date.setDate(QDate(today.year(), 1, 1))
            self.end_date.setDate(today)
        
        self.generate_report()
    
    def on_date_change(self):
        """Change les dates personnalisées"""
        self.period_combo.setCurrentText("Personnalisé")
    
    def show_export_dialog(self):
        """Affiche le dialogue d'export"""
        dialog = ExportDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.export_report(
                dialog.format_combo.currentText(),
                dialog.period_start.date().toPython(),
                dialog.period_end.date().toPython(),
                {
                    'charts': dialog.include_charts.isChecked(),
                    'tables': dialog.include_tables.isChecked(),
                    'summary': dialog.include_summary.isChecked(),
                    'raw_data': dialog.include_raw_data.isChecked()
                }
            )
    
    def export_report(self, format_str, start_date, end_date, options):
        """Exporte le rapport"""
        format_type = format_str.split(" - ")[0].lower()
        
        if format_type == "pdf":
            self.export_pdf(start_date, end_date, options)
        elif format_type == "csv":
            self.export_csv(start_date, end_date, options)
        elif format_type == "json":
            self.export_json(start_date, end_date, options)
        elif format_type == "excel":
            self.export_excel(start_date, end_date, options)
        elif format_type == "html":
            self.export_html(start_date, end_date, options)
        elif format_type == "image png":
            self.export_image(start_date, end_date, options, "png")
        elif format_type == "image svg":
            self.export_image(start_date, end_date, options, "svg")
    
    def export_csv(self, start_date, end_date, options):
        """Exporte en CSV"""
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter en CSV", f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV (*.csv)"
        )
        
        if path:
            try:
                with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Rapport généré le", datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
                    writer.writerow([])
                    writer.writerow(["Période", start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y')])
                    writer.writerow([])
                    
                    # Top véhicules
                    writer.writerow(["Top Véhicules", "Immatriculation", "Marque", "Contrats"])
                    for row in range(self.top_vehicles_table.rowCount()):
                        row_data = []
                        for col in range(self.top_vehicles_table.columnCount()):
                            item = self.top_vehicles_table.item(row, col)
                            row_data.append(item.text() if item else "")
                        writer.writerow(row_data)
                    
                    writer.writerow([])
                    
                    # Top clients
                    writer.writerow(["Top Clients", "Client", "Véhicules", "Prime totale"])
                    for row in range(self.top_clients_table.rowCount()):
                        row_data = []
                        for col in range(self.top_clients_table.columnCount()):
                            item = self.top_clients_table.item(row, col)
                            row_data.append(item.text() if item else "")
                        writer.writerow(row_data)
                
                QMessageBox.information(self, "Succès", f"CSV exporté: {path}")
            except Exception as e:
                QMessageBox.warning(self, "Erreur", f"Erreur d'export CSV: {e}")
    
    def export_json(self, start_date, end_date, options):
        """Exporte en JSON"""
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter en JSON", f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "JSON (*.json)"
        )
        
        if path:
            try:
                data = {
                    "meta": {
                        "generated_at": datetime.now().isoformat(),
                        "period": {
                            "start": start_date.isoformat(),
                            "end": end_date.isoformat()
                        }
                    },
                    "kpis": {
                        "total_vehicles": len(self.current_data.get('vehicles', [])),
                        "total_contracts": len(self.current_data.get('contracts', [])),
                    },
                    "top_vehicles": self._get_table_data(self.top_vehicles_table),
                    "top_clients": self._get_table_data(self.top_clients_table)
                }
                
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                
                QMessageBox.information(self, "Succès", f"JSON exporté: {path}")
            except Exception as e:
                QMessageBox.warning(self, "Erreur", f"Erreur d'export JSON: {e}")
    
    def export_excel(self, start_date, end_date, options):
        """Exporte en Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            path, _ = QFileDialog.getSaveFileName(
                self, "Exporter en Excel", f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel (*.xlsx)"
            )
            
            if path:
                wb = openpyxl.Workbook()
                
                # Feuille résumé
                ws = wb.active
                ws.title = "Résumé"
                
                ws['A1'] = "RAPPORT D'ANALYSE"
                ws['A1'].font = Font(size=16, bold=True)
                ws.merge_cells('A1:D1')
                
                ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                ws['A3'] = f"Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
                
                # Top véhicules
                ws_vehicles = wb.create_sheet("Top Véhicules")
                for col, header in enumerate(["#", "Immatriculation", "Marque", "Contrats"], 1):
                    cell = ws_vehicles.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                
                for row in range(self.top_vehicles_table.rowCount()):
                    for col in range(self.top_vehicles_table.columnCount()):
                        item = self.top_vehicles_table.item(row, col)
                        ws_vehicles.cell(row=row+2, column=col+1, value=item.text() if item else "")
                
                # Top clients
                ws_clients = wb.create_sheet("Top Clients")
                for col, header in enumerate(["#", "Client", "Véhicules", "Prime totale"], 1):
                    cell = ws_clients.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                
                for row in range(self.top_clients_table.rowCount()):
                    for col in range(self.top_clients_table.columnCount()):
                        item = self.top_clients_table.item(row, col)
                        ws_clients.cell(row=row+2, column=col+1, value=item.text() if item else "")
                
                wb.save(path)
                QMessageBox.information(self, "Succès", f"Excel exporté: {path}")
                
        except ImportError:
            QMessageBox.warning(self, "Erreur", "Le module openpyxl n'est pas installé.")
        except Exception as e:
            QMessageBox.warning(self, "Erreur", f"Erreur d'export Excel: {e}")
    
    def export_html(self, start_date, end_date, options):
        """Exporte en HTML"""
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter en HTML", f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html",
            "HTML (*.html)"
        )
        
        if path:
            try:
                html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Rapport d'analyse</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; }}
        h1 {{ color: #1e293b; }}
        .kpi-card {{ display: inline-block; padding: 20px; margin: 10px; background: #f8fafc; border-radius: 12px; }}
        .kpi-value {{ font-size: 24px; font-weight: bold; color: #3b82f6; }}
        .kpi-label {{ color: #64748b; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th {{ background: #f1f5f9; padding: 12px; text-align: left; }}
        td {{ padding: 12px; border-bottom: 1px solid #e2e8f0; }}
        .header {{ background: #1e293b; color: white; padding: 20px; border-radius: 12px; }}
    </style>
</head>
<body>
    <div class="header">
        <h1 style="color: white;">📊 Rapport d'analyse</h1>
        <p style="color: #94a3b8;">Généré le {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
        <p style="color: #94a3b8;">Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}</p>
    </div>
    
    <h2>📈 KPI Principaux</h2>
    <div>
        <div class="kpi-card">
            <div class="kpi-value">{len(self.current_data.get('vehicles', []))}</div>
            <div class="kpi-label">Véhicules</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-value">{len(self.current_data.get('contracts', []))}</div>
            <div class="kpi-label">Contrats</div>
        </div>
    </div>
    
    <h2>🚗 Top Véhicules</h2>
    {self._get_html_table(self.top_vehicles_table)}
    
    <h2>👥 Top Clients</h2>
    {self._get_html_table(self.top_clients_table)}
    
    <p style="color: #94a3b8; font-size: 12px; margin-top: 40px;">
        Rapport généré automatiquement - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
    </p>
</body>
</html>
"""
                with open(path, 'w', encoding='utf-8') as f:
                    f.write(html)
                
                QMessageBox.information(self, "Succès", f"HTML exporté: {path}")
            except Exception as e:
                QMessageBox.warning(self, "Erreur", f"Erreur d'export HTML: {e}")
    
    def export_image(self, start_date, end_date, options, format_type):
        """Exporte en image"""
        path, _ = QFileDialog.getSaveFileName(
            self, f"Exporter en {format_type.upper()}",
            f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{format_type}",
            f"{format_type.upper()} (*.{format_type})"
        )
        
        if path:
            try:
                pixmap = self.grab()
                pixmap.save(path)
                QMessageBox.information(self, "Succès", f"Image exportée: {path}")
            except Exception as e:
                QMessageBox.warning(self, "Erreur", f"Erreur d'export image: {e}")
    
    def _get_table_data(self, table):
        """Récupère les données d'un tableau"""
        data = []
        for row in range(table.rowCount()):
            row_data = []
            for col in range(table.columnCount()):
                item = table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        return data
    
    def _get_html_table(self, table):
        """Génère un tableau HTML"""
        html = "<table>"
        html += "<tr>"
        for col in range(table.columnCount()):
            html += f"<th>{table.horizontalHeaderItem(col).text()}</th>"
        html += "</tr>"
        
        for row in range(table.rowCount()):
            html += "<tr>"
            for col in range(table.columnCount()):
                item = table.item(row, col)
                html += f"<td>{item.text() if item else ''}</td>"
            html += "</tr>"
        
        html += "</table>"
        return html
    
    def export_pdf(self, start_date, end_date, options):
        """Exporte en PDF"""
        try:
            from PySide6.QtPrintSupport import QPrinter
            from PySide6.QtGui import QPageLayout, QPageSize
            from PySide6.QtWidgets import QTextEdit
            
            path, _ = QFileDialog.getSaveFileName(
                self, "Exporter en PDF", f"rapport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                "PDF (*.pdf)"
            )
            
            if path:
                # Créer un document temporaire
                doc = QTextEdit()
                doc.setHtml(self._generate_html_content(start_date, end_date))
                
                printer = QPrinter(QPrinter.HighResolution)
                printer.setOutputFormat(QPrinter.PdfFormat)
                printer.setOutputFileName(path)
                printer.setPageLayout(QPageLayout(QPageSize.A4, QPageLayout.Portrait))
                
                doc.print_(printer)
                
                QMessageBox.information(self, "Succès", f"PDF exporté: {path}")
                
        except Exception as e:
            QMessageBox.warning(self, "Erreur", f"Erreur d'export PDF: {e}")
    
    def _generate_html_content(self, start_date, end_date):
        """Génère le contenu HTML pour le PDF"""
        return f"""
        <h1>Rapport d'analyse</h1>
        <p>Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}</p>
        <p>Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
        <h2>Top Véhicules</h2>
        {self._get_html_table(self.top_vehicles_table)}
        <h2>Top Clients</h2>
        {self._get_html_table(self.top_clients_table)}
        """
    
    def share_report(self):
        """Partage le rapport"""
        QMessageBox.information(self, "Partage", 
            "📧 Fonctionnalité de partage\n\n"
            "Le rapport peut être partagé par:\n"
            "• Email (PDF/HTML)\n"
            "• Export (CSV/JSON/Excel)\n"
            "• Impression (PDF)\n\n"
            "Utilisez le bouton 'Exporter' pour générer le format souhaité.")
    
    def refresh_data(self):
        """Rafraîchit les données"""
        self.generate_report()